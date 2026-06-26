from __future__ import annotations

from copy import copy
from datetime import UTC, datetime
from html import escape
from itertools import zip_longest
import json
import os
from pathlib import Path
import re
import shutil
import subprocess
import sys
import tempfile
from typing import Any, Callable
from uuid import uuid4

from openpyxl import load_workbook

from .contracts import TABLE_BUILD_INTENT_KIND, validate_table_build_intent
from .excel_engine import ExcelEngineError, resource_script_path, sample_workbook_cells
from .sheets_api import (
    build_drive_copy_url,
    build_drive_delete_url,
    build_metadata_url,
    build_spreadsheet_batch_update_url,
    build_values_update_url,
    build_values_window_url,
    first_visible_sheet_title,
    google_delete_json,
    google_get_json,
    google_post_json,
    qualify_ranges,
    quote_sheet_title,
    sheet_title_for_gid,
)


DEFAULT_BUILDER_PACKAGE_ROOT = Path("review-packages/spreadsheet-table-builder/mcp")
DEFAULT_FORMULA_TEMPLATE = "=IFERROR(SUMIFS({measure_range},{row_label_range},{row_label_cell},{column_label_range},{column_label_cell}),0)"
ERROR_PREFIXES = ("#REF!", "#VALUE!", "#N/A", "#DIV/0!", "#ERROR!", "#NAME?", "#NUM!")
DEFAULT_OUTPUT_FORMAT = {
    "header_bold": True,
    "freeze_header_rows": 2,
    "auto_resize_columns": True,
    "protect_created_sheet": False,
}
DEFAULT_EXCEL_DESKTOP_WRITE_THRESHOLD_BYTES = 20 * 1024 * 1024
MAC_EXCEL_CONTAINER_DOCUMENTS = Path.home() / "Library/Containers/com.microsoft.Excel/Data/Documents"
DEFAULT_INTENT_PACKAGE_ROOT = Path("review-packages/spreadsheet-table-builder/intents")


def build_table_builder_ui(
    *,
    spreadsheet_id: str = "",
    access_token: str = "",
    workbook_path: str = "",
    sheet_name: str = "",
    gid: str = "",
    source_range: str = "",
    source_preview: dict[str, Any] | None = None,
    max_rows: int = 200,
    max_columns: int = 30,
    package_root: Path | str = DEFAULT_BUILDER_PACKAGE_ROOT,
    now: datetime | None = None,
    transport: Callable[[str, str], dict[str, Any]] | None = None,
) -> dict[str, Any]:
    if not spreadsheet_id and not workbook_path and not source_preview:
        raise ValueError("spreadsheet_id or workbook_path is required")
    if max_rows < 1 or max_columns < 1:
        raise ValueError("max_rows and max_columns must be positive")

    created_at = (now or datetime.now(UTC)).isoformat()
    if source_preview:
        source = _source_descriptor_from_preview(
            preview=source_preview,
            created_at=created_at,
            max_rows=max_rows,
            max_columns=max_columns,
        )
        source_node = source["source"]
        preview_id = str(source.get("spreadsheet_id") or source_node.get("sheet_title") or "source")
        package = write_table_builder_package(
            source=source,
            package_root=package_root,
            request_id=f"table-builder-preview-{_safe_id(preview_id)}-{_safe_id(source_node['sheet_title'])}",
            now=now,
        )
        return {
            "operation": "table_builder.ui",
            "artifact_type": source["artifact_type"],
            "spreadsheet_id": source.get("spreadsheet_id", ""),
            "source_range": source_node["qualified_range"],
            "summary": {
                "workbook_title": source["workbook_title"],
                "sheet_title": source_node["sheet_title"],
                "row_count": len(source["grid"]),
                "column_count": max((len(row) for row in source["grid"]), default=0),
                "source_authority": "sanitized_preview",
            },
            "package": package,
            "app_source": _app_source_payload(source, package),
        }
    if workbook_path:
        source = _excel_source_descriptor(
            workbook_path=workbook_path,
            sheet_name=sheet_name,
            source_range=source_range,
            max_rows=max_rows,
            max_columns=max_columns,
            created_at=created_at,
        )
        package = write_table_builder_package(
            source=source,
            package_root=package_root,
            request_id=f"table-builder-excel-{_safe_id(Path(workbook_path).stem)}-{_safe_id(source['source']['sheet_title'])}",
            now=now,
        )
        return {
            "operation": "table_builder.ui",
            "artifact_type": "excel_workbook",
            "workbook_path": str(Path(workbook_path).expanduser().resolve()),
            "source_range": source["source"]["qualified_range"],
            "summary": {
                "workbook_title": source["workbook_title"],
                "sheet_title": source["source"]["sheet_title"],
                "row_count": len(source["grid"]),
                "column_count": max((len(row) for row in source["grid"]), default=0),
            },
            "package": package,
            "app_source": _app_source_payload(source, package),
        }

    transport = transport or google_get_json
    metadata = transport(build_metadata_url(spreadsheet_id), access_token)
    qualified_range = _qualified_source_range(
        metadata=metadata,
        gid=gid,
        source_range=source_range,
        max_rows=max_rows,
        max_columns=max_columns,
    )
    sheet_title = _sheet_title_from_qualified_range(qualified_range)
    values_snapshot = transport(
        build_values_window_url(
            spreadsheet_id=spreadsheet_id,
            ranges=[qualified_range],
            value_render_option="FORMATTED_VALUE",
        ),
        access_token,
    )
    formulas_snapshot = transport(
        build_values_window_url(
            spreadsheet_id=spreadsheet_id,
            ranges=[qualified_range],
            value_render_option="FORMULA",
        ),
        access_token,
    )
    values = _first_value_matrix(values_snapshot)
    formulas = _first_value_matrix(formulas_snapshot)
    source = _source_descriptor(
        artifact_type="google_sheets",
        spreadsheet_id=spreadsheet_id,
        workbook_title=str((metadata.get("properties") or {}).get("title", "")),
        qualified_range=qualified_range,
        sheet_title=sheet_title,
        values=values,
        formulas=formulas,
        created_at=created_at,
    )
    package = write_table_builder_package(
        source=source,
        package_root=package_root,
        request_id=f"table-builder-{spreadsheet_id[:12]}-{_safe_id(sheet_title)}",
        now=now,
    )
    return {
        "operation": "table_builder.ui",
        "artifact_type": "google_sheets",
        "spreadsheet_id": spreadsheet_id,
        "source_range": qualified_range,
        "summary": {
            "workbook_title": source["workbook_title"],
            "sheet_title": sheet_title,
            "row_count": len(values),
            "column_count": max((len(row) for row in values), default=0),
        },
        "package": package,
        "app_source": _app_source_payload(source, package),
    }


def save_table_build_intent(
    *,
    intent: dict[str, Any],
    package_root: Path | str = DEFAULT_INTENT_PACKAGE_ROOT,
    now: datetime | None = None,
) -> dict[str, Any]:
    if not isinstance(intent, dict):
        raise ValueError("intent must be a JSON object")
    output_canvas = _normalize_output_canvas(intent.get("output_canvas"))
    llm_prompt = str(intent.get("llm_prompt") or "").strip()
    if not output_canvas:
        raise ValueError("intent.output_canvas is required")
    if not llm_prompt:
        raise ValueError("intent.llm_prompt is required")

    created_at = (now or datetime.now(UTC)).isoformat()
    intent_id = _safe_id(str(intent.get("intent_id") or f"table-build-intent-{created_at}-{uuid4().hex[:8]}"))
    package_dir = _unique_package_dir(package_root, created_at, intent_id)
    intent_path = package_dir / "intent.json"
    manifest_path = package_dir / "manifest.json"
    handoff_path = package_dir / "mcp-handoff.json"
    artifact_type = str(intent.get("artifact_type") or "")
    raw_output = intent.get("output") if isinstance(intent.get("output"), dict) else {}
    default_creation_mode = "copy" if artifact_type == "excel_workbook" else "sheet"
    output = {
        **raw_output,
        "creation_mode": str(raw_output.get("creation_mode") or default_creation_mode),
    }
    normalized_intent = {
        "schema_version": "1.0",
        "intent_kind": TABLE_BUILD_INTENT_KIND,
        "intent_id": package_dir.name,
        "created_at": created_at,
        "source": intent.get("source") if isinstance(intent.get("source"), dict) else {},
        "source_package": intent.get("source_package") if isinstance(intent.get("source_package"), dict) else {},
        "artifact_type": artifact_type,
        "output_canvas": output_canvas,
        "llm_prompt": llm_prompt,
        "source_hints": intent.get("source_hints") if isinstance(intent.get("source_hints"), dict) else {},
        "fields": intent.get("fields") if isinstance(intent.get("fields"), dict) else {},
        "formula": intent.get("formula") if isinstance(intent.get("formula"), dict) else {},
        "output": output,
        "review_state": {
            "status": "submitted",
            "next_action": "Generate a TableBuildPlan from this intent and ask the user to confirm the interpreted table shape.",
        },
    }
    validate_table_build_intent(normalized_intent)
    _write_json(intent_path, normalized_intent)
    _write_manifest(
        manifest_path=manifest_path,
        handoff_path=handoff_path,
        created_at=created_at,
        request_id=package_dir.name,
        package_dir=package_dir,
        primary_kind="table_build_intent",
        primary_path=intent_path,
        extra_artifacts=[],
    )
    return {
        "operation": "table_builder.save_intent",
        "intent_id": package_dir.name,
        "intent": normalized_intent,
        "package": {
            "package_dir": str(package_dir.resolve()),
            "manifest_path": str(manifest_path.resolve()),
            "intent_path": str(intent_path.resolve()),
            "mcp_handoff_path": str(handoff_path.resolve()),
        },
        "next_prompt": (
            "이 TableBuildIntent를 바탕으로 원본 evidence를 읽고, "
            "사용자에게 확인할 TableBuildPlan을 생성해주세요. "
            f"intent_path={intent_path.resolve()}"
        ),
    }


def create_formula_table_from_spec(
    *,
    spec: dict[str, Any],
    access_token: str,
    dry_run: bool = False,
    package_root: Path | str = DEFAULT_BUILDER_PACKAGE_ROOT,
    now: datetime | None = None,
    transport: Callable[[str, str], dict[str, Any]] | None = None,
    write_transport: Callable[[str, str, dict[str, Any]], dict[str, Any]] | None = None,
) -> dict[str, Any]:
    transport = transport or google_get_json
    write_transport = write_transport or google_post_json
    created_at = (now or datetime.now(UTC)).isoformat()
    normalized = normalize_formula_table_spec(spec)
    if normalized["source"]["artifact_type"] == "excel_workbook":
        return _create_excel_formula_table_from_spec(
            spec=normalized,
            dry_run=dry_run,
            package_root=package_root,
            now=now,
        )

    source_spreadsheet_id = normalized["spreadsheet_id"]
    creation_mode = normalized["output"]["creation_mode"]
    source_metadata = transport(build_metadata_url(source_spreadsheet_id), access_token)
    metadata = source_metadata
    target_spreadsheet_id = source_spreadsheet_id
    drive_copy_response: dict[str, Any] = {}
    if creation_mode == "copy" and not dry_run:
        source_title = str((source_metadata.get("properties") or {}).get("title", ""))
        copy_title = normalized["output"]["copy_title"] or (f"{source_title} - formula table" if source_title else "Formula table copy")
        drive_copy_response = write_transport(
            build_drive_copy_url(source_spreadsheet_id),
            access_token,
            {"name": copy_title},
        )
        target_spreadsheet_id = str(drive_copy_response.get("id") or "")
        if not target_spreadsheet_id:
            raise ValueError("Drive copy response did not include a copied spreadsheet id")
        metadata = transport(build_metadata_url(target_spreadsheet_id), access_token)
    source_values = transport(
        build_values_window_url(
            spreadsheet_id=target_spreadsheet_id,
            ranges=[normalized["source"]["qualified_range"]],
            value_render_option="FORMATTED_VALUE",
        ),
        access_token,
    )
    labels = _extract_layout_labels(normalized, _first_value_matrix(source_values))
    new_sheet_title = _unique_sheet_title(metadata, normalized["output"]["sheet_title"])
    new_sheet_id = _unique_sheet_id(metadata)
    write_grid = build_formula_table_grid(
        spec=normalized,
        row_labels=labels["row_labels"],
        column_labels=labels["column_labels"],
        output_sheet_title=new_sheet_title,
    )
    structural_requests = build_formula_table_batch_requests(
        new_sheet_id=new_sheet_id,
        new_sheet_title=new_sheet_title,
        row_count=len(write_grid),
        column_count=max((len(row) for row in write_grid), default=1),
        format_options=normalized["output"]["format"],
    )
    write_range = f"{quote_sheet_title(new_sheet_title)}!A1:{_column_label(max((len(row) for row in write_grid), default=1))}{len(write_grid)}"
    result: dict[str, Any] = {
        "schema_version": "1.0",
        "artifact_kind": "spreadsheet_formula_table_result",
        "operation": "table_builder.create_formula_table",
        "generated_at": created_at,
        "artifact_type": "google_sheets",
        "creation_mode": creation_mode,
        "source_spreadsheet_id": source_spreadsheet_id,
        "spreadsheet_id": target_spreadsheet_id,
        "dry_run": dry_run,
        "source_range": normalized["source"]["qualified_range"],
        "formula_template": normalized["formula"]["template"],
        "output_format": normalized["output"]["format"],
        "new_sheet_title": new_sheet_title,
        "new_sheet_id": new_sheet_id,
        "new_sheet_url": _sheet_url(target_spreadsheet_id, new_sheet_id),
        "row_label_count": len(labels["row_labels"]),
        "column_label_count": len(labels["column_labels"]),
        "label_source": str(labels.get("source") or "source_values"),
        "llm_prompt": normalized.get("llm_prompt", ""),
        "formula_cell_count": _count_formula_cells(write_grid),
        "write_range": write_range,
        "drive_copy_response": {
            "id": drive_copy_response.get("id", ""),
            "name": drive_copy_response.get("name", ""),
            "web_view_link": drive_copy_response.get("webViewLink", ""),
        } if drive_copy_response else {},
        "calculation_boundary": [
            "The output table cells are spreadsheet formulas that reference source ranges.",
            "The LLM does not calculate output values.",
            "Deterministic code extracts visible row/column labels only to size the requested output table.",
        ],
        "rollback": (
            {
                "type": "delete_copied_spreadsheet",
                "file_id": target_spreadsheet_id,
                "spreadsheet_id": target_spreadsheet_id,
                "reason": "Rollback is deleting the copied spreadsheet that contains the generated formula table sheet.",
            }
            if creation_mode == "copy"
            else {
                "type": "delete_created_sheet",
                "spreadsheet_id": target_spreadsheet_id,
                "sheet_id": new_sheet_id,
                "sheet_title": new_sheet_title,
                "reason": "The original data is not modified; rollback is removing the newly created formula table sheet.",
            }
        ),
    }
    if dry_run:
        if creation_mode == "copy":
            result["planned_copy"] = {
                "source_spreadsheet_id": source_spreadsheet_id,
                "copy_title": normalized["output"]["copy_title"] or str((source_metadata.get("properties") or {}).get("title", "")) + " - formula table",
            }
        result["validation"] = {"status": "not_run", "reason": "dry_run=true"}
        result["package"] = write_formula_table_package(result=result, package_root=package_root, now=now)
        return result

    batch_response = write_transport(
        build_spreadsheet_batch_update_url(target_spreadsheet_id),
        access_token,
        {"requests": structural_requests},
    )
    values_response = write_transport(
        build_values_update_url(target_spreadsheet_id),
        access_token,
        {
            "valueInputOption": "USER_ENTERED",
            "data": [{"range": write_range, "majorDimension": "ROWS", "values": write_grid}],
        },
    )
    readback = transport(
        build_values_window_url(
            spreadsheet_id=target_spreadsheet_id,
            ranges=[write_range],
            value_render_option="FORMATTED_VALUE",
        ),
        access_token,
    )
    result["batch_update_response"] = {"reply_count": len(batch_response.get("replies", []) or [])}
    result["values_update_response"] = {
        "total_updated_cells": values_response.get("totalUpdatedCells", 0),
        "total_updated_rows": values_response.get("totalUpdatedRows", 0),
        "total_updated_columns": values_response.get("totalUpdatedColumns", 0),
    }
    result["validation"] = _formula_table_readback_validation(_first_value_matrix(readback), len(write_grid), max((len(row) for row in write_grid), default=1))
    result["package"] = write_formula_table_package(result=result, package_root=package_root, now=now)
    return result


def rollback_created_artifact(
    *,
    rollback: dict[str, Any],
    access_token: str = "",
    dry_run: bool = False,
    write_transport: Callable[[str, str, dict[str, Any]], dict[str, Any]] | None = None,
    delete_transport: Callable[[str, str], dict[str, Any]] | None = None,
) -> dict[str, Any]:
    if not isinstance(rollback, dict):
        raise ValueError("rollback must be a JSON object")
    rollback_type = str(rollback.get("type") or rollback.get("operation") or "")
    if not rollback_type:
        raise ValueError("rollback.type is required")
    result: dict[str, Any] = {
        "schema_version": "1.0",
        "artifact_kind": "spreadsheet_rollback_result",
        "rollback_type": rollback_type,
        "dry_run": dry_run,
    }
    if rollback_type == "delete_created_sheet":
        spreadsheet_id = str(rollback.get("spreadsheet_id") or "")
        sheet_id = int(rollback.get("sheet_id") or 0)
        if not spreadsheet_id or not sheet_id:
            raise ValueError("delete_created_sheet rollback requires spreadsheet_id and sheet_id")
        result.update({"artifact_type": "google_sheets", "spreadsheet_id": spreadsheet_id, "sheet_id": sheet_id})
        if dry_run:
            result["status"] = "planned"
            return result
        write_transport = write_transport or google_post_json
        response = write_transport(
            build_spreadsheet_batch_update_url(spreadsheet_id),
            access_token,
            {"requests": [{"deleteSheet": {"sheetId": sheet_id}}]},
        )
        result["status"] = "rolled_back"
        result["response"] = {"reply_count": len(response.get("replies", []) or [])}
        return result
    if rollback_type == "delete_copied_spreadsheet":
        file_id = str(rollback.get("file_id") or rollback.get("spreadsheet_id") or "")
        if not file_id:
            raise ValueError("delete_copied_spreadsheet rollback requires file_id")
        result.update({"artifact_type": "google_sheets", "file_id": file_id})
        if dry_run:
            result["status"] = "planned"
            return result
        delete_transport = delete_transport or google_delete_json
        delete_transport(build_drive_delete_url(file_id), access_token)
        result["status"] = "rolled_back"
        return result
    if rollback_type == "delete_output_workbook_copy":
        raw_path = str(rollback.get("workbook_path") or "")
        if not raw_path:
            raise ValueError("delete_output_workbook_copy rollback requires workbook_path")
        path = Path(raw_path).expanduser().resolve()
        result.update({"artifact_type": "excel_workbook", "workbook_path": str(path)})
        if dry_run:
            result["status"] = "planned"
            result["exists"] = path.exists()
            return result
        if path.exists():
            path.unlink()
        result["status"] = "rolled_back"
        result["exists"] = path.exists()
        return result
    if rollback_type == "delete_created_worksheet":
        raw_path = str(rollback.get("workbook_path") or "")
        sheet_title = str(rollback.get("sheet_title") or "")
        if not raw_path or not sheet_title:
            raise ValueError("delete_created_worksheet rollback requires workbook_path and sheet_title")
        path = Path(raw_path).expanduser().resolve()
        result.update({"artifact_type": "excel_workbook", "workbook_path": str(path), "sheet_title": sheet_title})
        if dry_run:
            result["status"] = "planned"
            result["exists"] = path.exists()
            return result
        wb = load_workbook(path)
        try:
            if sheet_title in wb.sheetnames:
                wb.remove(wb[sheet_title])
                wb.save(path)
                result["status"] = "rolled_back"
            else:
                result["status"] = "not_found"
        finally:
            wb.close()
        return result
    raise ValueError(f"unsupported rollback type: {rollback_type}")


def validate_excel_formula_results(
    *,
    workbook_path: str,
    worksheet: str,
    cells: list[str],
    run_excel_engine: bool = True,
    timeout_seconds: int = 180,
    runner: Callable[..., subprocess.CompletedProcess[str]] | None = None,
) -> dict[str, Any]:
    if not workbook_path:
        raise ValueError("workbook_path is required")
    path = Path(workbook_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"missing workbook: {path}")
    if not worksheet:
        raise ValueError("worksheet is required")
    if not cells:
        raise ValueError("cells are required")
    static_scan = _excel_formula_error_scan(path)
    result: dict[str, Any] = {
        "schema_version": "1.0",
        "artifact_kind": "excel_formula_result_validation",
        "workbook_path": str(path),
        "worksheet": worksheet,
        "cells": cells,
        "static_scan": static_scan,
        "excel_engine": {"status": "not_run"},
    }
    command = _excel_engine_sample_command(path, worksheet, cells)
    result["excel_engine"]["command"] = command
    if not run_excel_engine:
        result["status"] = "review_required" if static_scan["summary"]["error_count"] else "static_passed"
        result["excel_engine"]["reason"] = "run_excel_engine=false"
        return result
    if runner is not None:
        completed = runner(command, check=False, capture_output=True, text=True, timeout=timeout_seconds)
        if completed.returncode != 0:
            result["status"] = "failed"
            result["excel_engine"] = {
                "status": "failed",
                "command": command,
                "returncode": completed.returncode,
                "stdout": completed.stdout,
                "stderr": completed.stderr,
            }
            return result
        values = _parse_excel_engine_output(completed.stdout)
    else:
        try:
            values = sample_workbook_cells(path, worksheet, cells, timeout=timeout_seconds)
        except (ExcelEngineError, OSError, ValueError) as error:
            result["status"] = "failed"
            result["excel_engine"] = {
                "status": "failed",
                "command": command,
                "error": str(error),
            }
            return result
    result["excel_engine"] = {
        "status": "passed",
        "command": command,
        "values": values,
    }
    result["status"] = "passed" if not static_scan["summary"]["error_count"] else "review_required"
    return result


def normalize_formula_table_spec(spec: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(spec, dict):
        raise ValueError("spec must be a JSON object")
    source = spec.get("source") if isinstance(spec.get("source"), dict) else {}
    fields = spec.get("fields") if isinstance(spec.get("fields"), dict) else {}
    output = spec.get("output") if isinstance(spec.get("output"), dict) else {}
    artifact_type = str(spec.get("artifact_type") or source.get("artifact_type") or source.get("type") or "").strip()
    if not artifact_type:
        artifact_type = "excel_workbook" if source.get("workbook_path") else "google_sheets"
    if artifact_type not in {"google_sheets", "excel_workbook"}:
        raise ValueError("spec source artifact_type must be google_sheets or excel_workbook")
    spreadsheet_id = str(spec.get("spreadsheet_id") or source.get("spreadsheet_id") or "")
    workbook_path = str(spec.get("workbook_path") or source.get("workbook_path") or "")
    if artifact_type == "google_sheets" and not spreadsheet_id:
        raise ValueError("spec.spreadsheet_id is required for google_sheets")
    if artifact_type == "excel_workbook" and not workbook_path:
        raise ValueError("spec.source.workbook_path is required for excel_workbook")
    creation_mode = str(output.get("creation_mode") or "").strip().lower()
    if not creation_mode:
        creation_mode = "copy" if artifact_type == "excel_workbook" else "sheet"
    if creation_mode not in {"copy", "sheet"}:
        raise ValueError("spec.output.creation_mode must be copy or sheet")
    if artifact_type == "google_sheets" and output.get("workbook_path"):
        raise ValueError("output.workbook_path is only valid for excel_workbook creation_mode='copy'")
    if artifact_type == "excel_workbook" and creation_mode == "sheet" and output.get("workbook_path"):
        raise ValueError("output.workbook_path is only valid when output.creation_mode='copy'")
    qualified_range = str(source.get("qualified_range") or "")
    if "!" not in qualified_range:
        raise ValueError("spec.source.qualified_range is required")
    start_col, start_row, end_col, end_row = _range_bounds(qualified_range)
    header_row = int(source.get("header_row") or 1)
    if header_row < 1 or start_row + header_row > end_row:
        raise ValueError("spec.source.header_row must be inside the source range")
    formula_template = _formula_template(spec)
    output_format = _normalize_output_format(output.get("format") if isinstance(output.get("format"), dict) else {})
    output_canvas = _normalize_output_canvas(spec.get("output_canvas"))
    normalized_fields = {
        role: _normalize_field(fields.get(role), role)
        for role in ("row_label", "column_label", "measure")
    }
    for role, field in normalized_fields.items():
        column_index = _column_index(field["column"])
        if column_index < start_col or column_index > end_col:
            raise ValueError(f"spec.fields.{role}.column is outside the source range")
    return {
        "schema_version": "1.0",
        "spec_kind": "formula_table_apply_v1",
        "artifact_type": artifact_type,
        "spreadsheet_id": spreadsheet_id,
        "source": {
            "artifact_type": artifact_type,
            "spreadsheet_id": spreadsheet_id,
            "workbook_path": str(Path(workbook_path).expanduser().resolve()) if workbook_path else "",
            "sheet_title": str(source.get("sheet_title") or _sheet_title_from_qualified_range(qualified_range)),
            "qualified_range": qualified_range,
            "header_row": header_row,
        },
        "fields": normalized_fields,
        "output_canvas": output_canvas,
        "llm_prompt": str(spec.get("llm_prompt") or ""),
        "formula": {"template": formula_template},
        "output": {
            "sheet_title": str(output.get("sheet_title") or f"MCP_TABLE_{datetime.now(UTC).strftime('%Y%m%d_%H%M')}"),
            "title": str(output.get("title") or "MCP Formula Table"),
            "creation_mode": creation_mode,
            "workbook_path": str(output.get("workbook_path") or ""),
            "copy_title": str(output.get("copy_title") or ""),
            "format": output_format,
        },
    }


def build_formula_table_grid(
    *,
    spec: dict[str, Any],
    row_labels: list[str],
    column_labels: list[str],
    output_sheet_title: str,
) -> list[list[Any]]:
    source = spec["source"]
    fields = spec["fields"]
    formula_template = spec["formula"]["template"]
    source_sheet = quote_sheet_title(source["sheet_title"])
    start_col, start_row, _end_col, end_row = _range_bounds(source["qualified_range"])
    data_start_row = start_row + int(source["header_row"])
    row_range = _bounded_column_range(source_sheet, fields["row_label"]["column"], data_start_row, end_row)
    column_range = _bounded_column_range(source_sheet, fields["column_label"]["column"], data_start_row, end_row)
    measure_range = _bounded_column_range(source_sheet, fields["measure"]["column"], data_start_row, end_row)
    grid: list[list[Any]] = [
        [spec["output"]["title"], "", "Generated by Spreadsheet formula-table builder"],
        [fields["row_label"]["header"], *column_labels],
    ]
    for row_index, row_label in enumerate(row_labels, start=3):
        formula_row = [row_label]
        for column_index, _column_label_value in enumerate(column_labels, start=2):
            formula_row.append(
                _render_formula_template(
                    template=formula_template,
                    measure_range=measure_range,
                    row_range=row_range,
                    column_range=column_range,
                    row_criteria=f"$A{row_index}",
                    column_criteria=f"{_column_label(column_index)}$2",
                    row_value=row_label,
                    column_value=_column_label_value,
                    source_sheet=source_sheet,
                    source_range=source["qualified_range"],
                    output_sheet=quote_sheet_title(output_sheet_title),
                )
            )
        grid.append(formula_row)
    if len(grid) == 2:
        grid.append(["No matching labels found"])
    return grid


def build_formula_table_batch_requests(
    *,
    new_sheet_id: int,
    new_sheet_title: str,
    row_count: int,
    column_count: int,
    format_options: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    options = _normalize_output_format(format_options or {})
    requests = [
        {
            "addSheet": {
                "properties": {
                    "sheetId": new_sheet_id,
                    "title": new_sheet_title,
                    "gridProperties": {
                        "rowCount": max(row_count + 10, 20),
                        "columnCount": max(column_count + 5, 8),
                    },
                }
            }
        }
    ]
    if options["freeze_header_rows"]:
        requests.append(
            {
                "updateSheetProperties": {
                    "properties": {
                        "sheetId": new_sheet_id,
                        "gridProperties": {"frozenRowCount": min(int(options["freeze_header_rows"]), max(row_count, 1))},
                    },
                    "fields": "gridProperties.frozenRowCount",
                }
            }
        )
    if options["header_bold"]:
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": new_sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": min(2, max(row_count, 1)),
                    "startColumnIndex": 0,
                    "endColumnIndex": max(column_count, 1),
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": {"red": 0.92, "green": 0.96, "blue": 1.0},
                        "textFormat": {"bold": True},
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat)",
            }
        })
    if options["auto_resize_columns"]:
        requests.append(
            {
                "autoResizeDimensions": {
                    "dimensions": {
                        "sheetId": new_sheet_id,
                        "dimension": "COLUMNS",
                        "startIndex": 0,
                        "endIndex": max(column_count, 1),
                    }
                }
            }
        )
    if options["protect_created_sheet"]:
        requests.append(
            {
                "addProtectedRange": {
                    "protectedRange": {
                        "description": "Generated formula table review surface",
                        "range": {
                            "sheetId": new_sheet_id,
                            "startRowIndex": 0,
                            "endRowIndex": max(row_count, 1),
                            "startColumnIndex": 0,
                            "endColumnIndex": max(column_count, 1),
                        },
                        "warningOnly": True,
                    }
                }
            }
        )
    return requests


def write_table_builder_package(
    *,
    source: dict[str, Any],
    package_root: Path | str = DEFAULT_BUILDER_PACKAGE_ROOT,
    request_id: str | None = None,
    now: datetime | None = None,
) -> dict[str, Any]:
    created_at = (now or datetime.now(UTC)).isoformat()
    package_dir = _unique_package_dir(package_root, created_at, request_id or "table-builder-ui")
    source_path = package_dir / "builder-source.json"
    html_path = package_dir / "index.html"
    manifest_path = package_dir / "manifest.json"
    handoff_path = package_dir / "mcp-handoff.json"
    _write_json(source_path, source)
    html_path.write_text(_build_table_builder_html(source), encoding="utf-8")
    _write_manifest(
        manifest_path=manifest_path,
        handoff_path=handoff_path,
        created_at=created_at,
        request_id=package_dir.name,
        package_dir=package_dir,
        primary_kind="table_builder_ui",
        primary_path=html_path,
        extra_artifacts=[{"kind": "table_builder_source", "path": str(source_path.resolve())}],
    )
    return {
        "package_dir": str(package_dir.resolve()),
        "manifest_path": str(manifest_path.resolve()),
        "html_path": str(html_path.resolve()),
        "source_path": str(source_path.resolve()),
        "mcp_handoff_path": str(handoff_path.resolve()),
    }


def build_table_builder_mcp_app_html() -> str:
    template = _read_table_builder_ui_resource("mcp_app.html")
    host_adapter_js = _read_table_builder_ui_resource("host_adapter.js")
    return template.replace(
        "<!-- SHEETS_BRIDGE_HOST_ADAPTER_BUNDLE -->",
        f"<script>\n{host_adapter_js}\n</script>",
    )


def _read_table_builder_ui_resource(name: str) -> str:
    return _table_builder_ui_resource_path(name).read_text(encoding="utf-8")


def _table_builder_ui_resource_path(name: str) -> Path:
    candidates: list[Path] = []
    bundle_root = getattr(sys, "_MEIPASS", None)
    if bundle_root:
        candidates.append(Path(bundle_root) / "sheets_bridge" / "ui" / "table_builder" / name)
    candidates.append(Path(__file__).resolve().parent / "ui" / "table_builder" / name)
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def _app_source_payload(source: dict[str, Any], package: dict[str, Any]) -> dict[str, Any]:
    return {
        **source,
        "package": {
            "manifest_path": package.get("manifest_path", ""),
            "source_path": package.get("source_path", ""),
            "html_path": package.get("html_path", ""),
            "mcp_handoff_path": package.get("mcp_handoff_path", ""),
        },
    }


def write_formula_table_package(
    *,
    result: dict[str, Any],
    package_root: Path | str = DEFAULT_BUILDER_PACKAGE_ROOT,
    request_id: str | None = None,
    now: datetime | None = None,
) -> dict[str, Any]:
    created_at = (now or datetime.now(UTC)).isoformat()
    package_dir = _unique_package_dir(package_root, created_at, request_id or "formula-table-result")
    result_path = package_dir / "result.json"
    html_path = package_dir / "index.html"
    manifest_path = package_dir / "manifest.json"
    handoff_path = package_dir / "mcp-handoff.json"
    result_for_file = {key: value for key, value in result.items() if key != "package"}
    _write_json(result_path, result_for_file)
    html_path.write_text(_build_formula_table_result_html(result_for_file), encoding="utf-8")
    _write_manifest(
        manifest_path=manifest_path,
        handoff_path=handoff_path,
        created_at=created_at,
        request_id=package_dir.name,
        package_dir=package_dir,
        primary_kind="formula_table_result",
        primary_path=result_path,
        extra_artifacts=[{"kind": "formula_table_result_html", "path": str(html_path.resolve())}],
    )
    return {
        "package_dir": str(package_dir.resolve()),
        "manifest_path": str(manifest_path.resolve()),
        "result_path": str(result_path.resolve()),
        "html_path": str(html_path.resolve()),
        "mcp_handoff_path": str(handoff_path.resolve()),
    }


def _create_excel_formula_table_from_spec(
    *,
    spec: dict[str, Any],
    dry_run: bool,
    package_root: Path | str,
    now: datetime | None,
) -> dict[str, Any]:
    created_at = (now or datetime.now(UTC)).isoformat()
    source_path = Path(spec["source"]["workbook_path"]).expanduser().resolve()
    if not source_path.exists():
        raise FileNotFoundError(f"missing workbook: {source_path}")
    source_values = _excel_values_for_range(
        workbook_path=source_path,
        sheet_title=spec["source"]["sheet_title"],
        qualified_range=spec["source"]["qualified_range"],
    )
    labels = _extract_layout_labels(spec, source_values)
    creation_mode = str(spec["output"].get("creation_mode") or "copy")
    target_path = (
        _excel_output_path(source_path, str(spec["output"].get("workbook_path") or ""))
        if creation_mode == "copy"
        else source_path
    )

    probe = load_workbook(source_path, read_only=True, data_only=False)
    try:
        new_sheet_title = _unique_excel_sheet_title(probe.sheetnames, spec["output"]["sheet_title"])
    finally:
        probe.close()
    write_grid = build_formula_table_grid(
        spec=spec,
        row_labels=labels["row_labels"],
        column_labels=labels["column_labels"],
        output_sheet_title=new_sheet_title,
    )
    result: dict[str, Any] = {
        "schema_version": "1.0",
        "artifact_kind": "spreadsheet_formula_table_result",
        "operation": "table_builder.create_formula_table",
        "generated_at": created_at,
        "artifact_type": "excel_workbook",
        "creation_mode": creation_mode,
        "dry_run": dry_run,
        "source_workbook_path": str(source_path),
        "workbook_path": str(target_path),
        "source_range": spec["source"]["qualified_range"],
        "formula_template": spec["formula"]["template"],
        "output_format": spec["output"]["format"],
        "new_sheet_title": new_sheet_title,
        "row_label_count": len(labels["row_labels"]),
        "column_label_count": len(labels["column_labels"]),
        "label_source": str(labels.get("source") or "source_values"),
        "llm_prompt": spec.get("llm_prompt", ""),
        "formula_cell_count": _count_formula_cells(write_grid),
        "write_range": f"{quote_sheet_title(new_sheet_title)}!A1:{_column_label(max((len(row) for row in write_grid), default=1))}{len(write_grid)}",
        "calculation_boundary": [
            "The output table cells are Excel formulas that reference source ranges.",
            "The LLM does not calculate output values.",
            "Formula-result validation requires Microsoft Excel recalculation.",
        ],
        "rollback": (
            {
                "type": "delete_output_workbook_copy",
                "workbook_path": str(target_path),
                "reason": "Rollback is removing the generated workbook copy.",
            }
            if creation_mode == "copy"
            else {
                "type": "delete_created_worksheet",
                "workbook_path": str(target_path),
                "sheet_title": new_sheet_title,
                "reason": "Rollback is removing the worksheet created for this formula table.",
            }
        ),
    }
    if dry_run:
        result["validation"] = {"status": "not_run", "reason": "dry_run=true"}
        result["package"] = write_formula_table_package(result=result, package_root=package_root, now=now)
        return result

    write_engine = "openpyxl"
    try:
        if creation_mode == "copy":
            shutil.copy2(source_path, target_path)
            write_engine = _excel_copy_write_engine(source_path)
            if write_engine == "desktop_excel":
                _write_excel_formula_table_with_desktop_excel(
                    target_path=target_path,
                    sheet_title=new_sheet_title,
                    write_grid=write_grid,
                    format_options=spec["output"]["format"],
                )
            else:
                _write_excel_formula_table_with_openpyxl(
                    workbook_path=target_path,
                    sheet_title=new_sheet_title,
                    write_grid=write_grid,
                    format_options=spec["output"]["format"],
                )
        else:
            _write_excel_formula_table_with_openpyxl(
                workbook_path=target_path,
                sheet_title=new_sheet_title,
                write_grid=write_grid,
                format_options=spec["output"]["format"],
            )
        result["write_engine"] = write_engine
        result["new_sheet_title"] = new_sheet_title
        result["formula_cell_count"] = _count_formula_cells(write_grid)
        result["write_range"] = f"{quote_sheet_title(new_sheet_title)}!A1:{_column_label(max((len(row) for row in write_grid), default=1))}{len(write_grid)}"
        if result["rollback"]["type"] == "delete_created_worksheet":
            result["rollback"]["sheet_title"] = new_sheet_title
    except Exception:
        if creation_mode == "copy" and target_path.exists():
            try:
                target_path.unlink()
            except OSError:
                pass
        raise

    result["validation"] = _excel_formula_table_static_validation(
        workbook_path=target_path,
        sheet_title=str(result["new_sheet_title"]),
        expected_formula_count=result["formula_cell_count"],
    )
    result["package"] = write_formula_table_package(result=result, package_root=package_root, now=now)
    return result


def _source_descriptor(
    *,
    artifact_type: str,
    spreadsheet_id: str,
    workbook_title: str,
    qualified_range: str,
    sheet_title: str,
    values: list[list[Any]],
    formulas: list[list[Any]],
    created_at: str,
) -> dict[str, Any]:
    start_col, start_row, _end_col, _end_row = _range_bounds(qualified_range)
    sheet_suffix = created_at[:16].replace("-", "").replace(":", "").replace("T", "_")
    default_creation_mode = "copy" if artifact_type == "excel_workbook" else "sheet"
    return {
        "schema_version": "1.0",
        "artifact_kind": "spreadsheet_table_builder_source",
        "created_at": created_at,
        "artifact_type": artifact_type,
        "spreadsheet_id": spreadsheet_id,
        "workbook_title": workbook_title,
        "source": {
            "artifact_type": artifact_type,
            "spreadsheet_id": spreadsheet_id,
            "workbook_path": "",
            "sheet_title": sheet_title,
            "qualified_range": qualified_range,
            "range_start_row": start_row,
            "range_start_column": _column_label(start_col),
            "header_row": 1,
        },
        "grid": _grid_cells(values=values, formulas=formulas, start_col=start_col, start_row=start_row),
        "default_spec": {
            "schema_version": "1.0",
            "spec_kind": "formula_table_apply_v1",
            "artifact_type": artifact_type,
            "spreadsheet_id": spreadsheet_id,
            "source": {
                "artifact_type": artifact_type,
                "spreadsheet_id": spreadsheet_id,
                "workbook_path": "",
                "sheet_title": sheet_title,
                "qualified_range": qualified_range,
                "header_row": 1,
            },
            "fields": {},
            "formula": {"template": DEFAULT_FORMULA_TEMPLATE},
            "output": {
                "sheet_title": f"MCP_TABLE_{sheet_suffix}",
                "title": "MCP Formula Table",
                "creation_mode": default_creation_mode,
                "workbook_path": "",
                "copy_title": f"{workbook_title} - formula table" if workbook_title else "",
                "format": dict(DEFAULT_OUTPUT_FORMAT),
            },
        },
    }


def _source_descriptor_from_preview(
    *,
    preview: dict[str, Any],
    created_at: str,
    max_rows: int,
    max_columns: int,
) -> dict[str, Any]:
    if not isinstance(preview, dict):
        raise ValueError("source_preview must be a JSON object")
    source_node = preview.get("source") if isinstance(preview.get("source"), dict) else {}
    artifact_type = str(preview.get("artifact_type") or source_node.get("artifact_type") or "google_sheets")
    if artifact_type not in {"google_sheets", "excel_workbook"}:
        raise ValueError("source_preview.artifact_type must be google_sheets or excel_workbook")
    spreadsheet_id = str(preview.get("spreadsheet_id") or source_node.get("spreadsheet_id") or "")
    if artifact_type == "google_sheets" and not spreadsheet_id:
        raise ValueError("source_preview.spreadsheet_id is required for google_sheets")
    qualified_range = str(
        preview.get("qualified_range")
        or preview.get("source_range")
        or source_node.get("qualified_range")
        or ""
    )
    if "!" not in qualified_range:
        raise ValueError("source_preview.qualified_range is required")
    sheet_title = str(preview.get("sheet_title") or source_node.get("sheet_title") or _sheet_title_from_qualified_range(qualified_range))
    workbook_title = str(preview.get("workbook_title") or preview.get("title") or "")
    values = _limited_matrix(
        preview.get("values") if isinstance(preview.get("values"), list) else _values_from_preview_grid(preview.get("grid")),
        max_rows=max_rows,
        max_columns=max_columns,
    )
    formulas = _limited_matrix(
        preview.get("formulas") if isinstance(preview.get("formulas"), list) else _formulas_from_preview_grid(preview.get("grid")),
        max_rows=max_rows,
        max_columns=max_columns,
    )
    if not values and not formulas:
        raise ValueError("source_preview.values or source_preview.grid is required")
    return _source_descriptor(
        artifact_type=artifact_type,
        spreadsheet_id=spreadsheet_id,
        workbook_title=workbook_title,
        qualified_range=qualified_range,
        sheet_title=sheet_title,
        values=values,
        formulas=formulas,
        created_at=created_at,
    )


def _limited_matrix(value: Any, *, max_rows: int, max_columns: int) -> list[list[Any]]:
    if not isinstance(value, list):
        return []
    result: list[list[Any]] = []
    for row in value[:max_rows]:
        if not isinstance(row, list):
            result.append([])
            continue
        result.append(row[:max_columns])
    return result


def _values_from_preview_grid(grid: Any) -> list[list[Any]]:
    if not isinstance(grid, list):
        return []
    values: list[list[Any]] = []
    for row in grid:
        if not isinstance(row, list):
            values.append([])
            continue
        values.append([
            _cell_text(cell.get("formatted_value", cell.get("value", ""))) if isinstance(cell, dict) else _cell_text(cell)
            for cell in row
        ])
    return values


def _formulas_from_preview_grid(grid: Any) -> list[list[Any]]:
    if not isinstance(grid, list):
        return []
    formulas: list[list[Any]] = []
    for row in grid:
        if not isinstance(row, list):
            formulas.append([])
            continue
        formulas.append([
            _cell_text(cell.get("formula", "")) if isinstance(cell, dict) else ""
            for cell in row
        ])
    return formulas


def _grid_cells(
    *,
    values: list[list[Any]],
    formulas: list[list[Any]],
    start_col: int,
    start_row: int,
) -> list[list[dict[str, Any]]]:
    row_count = max(len(values), len(formulas))
    result = []
    for row_offset in range(row_count):
        value_row = values[row_offset] if row_offset < len(values) and isinstance(values[row_offset], list) else []
        formula_row = formulas[row_offset] if row_offset < len(formulas) and isinstance(formulas[row_offset], list) else []
        column_count = max(len(value_row), len(formula_row))
        cells = []
        for column_offset in range(column_count):
            absolute_column = start_col + column_offset
            absolute_row = start_row + row_offset
            value = value_row[column_offset] if column_offset < len(value_row) else ""
            formula = formula_row[column_offset] if column_offset < len(formula_row) else ""
            cells.append(
                {
                    "a1": f"{_column_label(absolute_column)}{absolute_row}",
                    "row": absolute_row,
                    "column": _column_label(absolute_column),
                    "formatted_value": _cell_text(value),
                    "formula": _cell_text(formula) if _cell_text(formula).startswith("=") else "",
                }
            )
        result.append(cells)
    return result


def _build_table_builder_html(source: dict[str, Any]) -> str:
    data_json = json.dumps(source, ensure_ascii=False)
    grid_rows = []
    for row in source["grid"]:
        row_number = row[0]["row"] if row else ""
        cells = [
            f'<th class="row-head">{escape(str(row_number))}</th>',
            *[
                f'<td data-a1="{escape(cell["a1"])}" data-column="{escape(cell["column"])}" data-value="{escape(cell["formatted_value"])}"><span>{escape(cell["formatted_value"]) or "&nbsp;"}</span>{"<b>fx</b>" if cell["formula"] else ""}</td>'
                for cell in row
            ],
        ]
        grid_rows.append("<tr>" + "".join(cells) + "</tr>")
    columns = source["grid"][0] if source["grid"] else []
    column_headers = "".join(f'<th>{escape(cell["column"])}</th>' for cell in columns)
    rows_html = "\n".join(grid_rows)
    canvas_column_headers = "".join(f"<th>{_column_label(index)}</th>" for index in range(1, 8))
    canvas_rows = []
    for row_index in range(1, 9):
        cells = [
            f'<th class="row-head">{row_index}</th>',
            *[
                f'<td contenteditable="true" data-canvas-cell="{_column_label(column_index)}{row_index}" aria-label="{_column_label(column_index)}{row_index}"></td>'
                for column_index in range(1, 8)
            ],
        ]
        canvas_rows.append("<tr>" + "".join(cells) + "</tr>")
    canvas_rows_html = "\n".join(canvas_rows)
    return f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta http-equiv="Cache-Control" content="no-store">
  <meta http-equiv="Pragma" content="no-cache">
  <meta http-equiv="Expires" content="0">
  <title>새 표 만들기</title>
  <style>
    *{{box-sizing:border-box}} body{{margin:0;background:#f6f8fb;color:#172033;font-family:Arial,'Apple SD Gothic Neo',sans-serif}} main{{display:grid;grid-template-columns:520px 1fr;gap:18px;min-height:100vh;padding:18px}} aside,section{{background:#fff;border:1px solid #d8e1ea;border-radius:8px}} aside{{padding:18px;overflow:auto;max-height:calc(100vh - 36px)}} h1{{font-size:26px;margin:0 0 6px}} h2{{font-size:16px;margin:18px 0 8px}} p{{color:#5d6a7a;line-height:1.45}} button,select,input{{font:inherit}} .role{{display:grid;grid-template-columns:48px 1fr;gap:12px;align-items:center;width:100%;padding:12px;margin:8px 0;border:1px solid #cbd7e3;border-radius:8px;background:#fff;text-align:left;cursor:pointer}} .role strong{{display:block;font-size:15px}} .role small{{display:block;margin-top:3px;color:#5d6a7a;line-height:1.35}} .role.active{{border-color:#1b7895;background:#e8f7fb;box-shadow:0 0 0 2px rgba(27,120,149,.12)}} .role-icon{{width:44px;height:44px;border:1px solid #cbd7e3;border-radius:8px;background-color:#fff;box-shadow:inset 0 0 0 1px rgba(255,255,255,.7)}} .role-icon.row{{background:linear-gradient(90deg,#1b7895 0 32%,transparent 32%),linear-gradient(#eef3f8 0 32%,transparent 32%),linear-gradient(90deg,transparent 31%,#d7dde8 31%,#d7dde8 34%,transparent 34%,transparent 65%,#d7dde8 65%,#d7dde8 68%,transparent 68%),linear-gradient(transparent 31%,#d7dde8 31%,#d7dde8 34%,transparent 34%,transparent 65%,#d7dde8 65%,#d7dde8 68%,transparent 68%)}} .role-icon.column{{background:linear-gradient(#d79216 0 32%,transparent 32%),linear-gradient(90deg,#eef3f8 0 32%,transparent 32%),linear-gradient(90deg,transparent 31%,#d7dde8 31%,#d7dde8 34%,transparent 34%,transparent 65%,#d7dde8 65%,#d7dde8 68%,transparent 68%),linear-gradient(transparent 31%,#d7dde8 31%,#d7dde8 34%,transparent 34%,transparent 65%,#d7dde8 65%,#d7dde8 68%,transparent 68%)}} .role-icon.measure{{background:linear-gradient(90deg,#eef3f8 0 32%,transparent 32%),linear-gradient(#eef3f8 0 32%,transparent 32%),radial-gradient(circle at 67% 67%,#2f9e55 0 18%,transparent 19%),linear-gradient(90deg,transparent 31%,#d7dde8 31%,#d7dde8 34%,transparent 34%,transparent 65%,#d7dde8 65%,#d7dde8 68%,transparent 68%),linear-gradient(transparent 31%,#d7dde8 31%,#d7dde8 34%,transparent 34%,transparent 65%,#d7dde8 65%,#d7dde8 68%,transparent 68%)}} label{{display:block;margin:10px 0 4px;color:#5d6a7a;font-weight:700}} input,select,textarea{{width:100%;border:1px solid #cbd7e3;border-radius:8px;padding:10px;background:#fff}} textarea{{height:220px;font-family:ui-monospace,SFMono-Regular,Menlo,monospace;font-size:12px}} textarea.prompt{{height:92px;font-family:inherit;font-size:14px}} textarea.formula{{height:96px}} details{{margin-top:12px;border:1px solid #d8e1ea;border-radius:8px;background:#fbfdff}} summary{{padding:12px;font-weight:700;cursor:pointer}} details .panel{{padding:0 12px 12px}} .actions{{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:10px}} .actions.one{{grid-template-columns:1fr}} .actions button{{border:0;border-radius:8px;background:#1b7895;color:white;font-weight:700;padding:11px;cursor:pointer}} .actions button.secondary{{background:#eef3f8;color:#172033}} .output-canvas-wrap{{border:1px solid #cbd7e3;border-radius:8px;background:#fff;overflow:auto;max-height:310px}} table.output-canvas{{min-width:0;width:max-content;border-collapse:separate;border-spacing:0}} .output-canvas th,.output-canvas td{{min-width:96px;max-width:180px;height:34px;padding:6px 8px;border-right:1px solid #d7dde8;border-bottom:1px solid #d7dde8;background:#fff;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}} .output-canvas th{{position:sticky;top:0;background:#eef3f8;text-align:center;z-index:2}} .output-canvas .row-head{{position:sticky;left:0;z-index:1;min-width:48px;background:#eef3f8;text-align:right}} .output-canvas td:focus{{outline:2px solid #1b7895;outline-offset:-2px;background:#f7fcff}} .preview{{margin-top:12px;border:1px solid #d8e1ea;border-radius:8px;background:#fff}} .preview-head{{display:flex;justify-content:space-between;gap:8px;padding:10px 12px;border-bottom:1px solid #e5ebf2;font-weight:700}} .preview-head small{{font-weight:400;color:#5d6a7a}} .preview-title{{display:grid;gap:3px}} .preview-tools{{display:inline-flex;align-self:start;border:1px solid #cbd7e3;border-radius:8px;overflow:hidden;background:#fff}} .preview-tools button{{border:0;border-right:1px solid #cbd7e3;background:#fff;color:#172033;padding:6px 9px;font-size:12px;cursor:pointer}} .preview-tools button:last-child{{border-right:0}} .preview-tools button.active{{background:#1b7895;color:#fff;font-weight:700}} .preview-empty{{padding:12px;color:#5d6a7a}} .plan{{padding:12px;border-top:1px solid #e5ebf2}} .plan h3{{font-size:14px;margin:0 0 8px}} .plan ul{{margin:8px 0 0;padding-left:18px;color:#374151;line-height:1.45}} .plan-note{{margin:8px 0 0;padding:10px;border-radius:8px;background:#fff7ed;color:#7c2d12}} .preview-scroll{{overflow:auto;max-height:240px}} .preview table{{min-width:0;width:max-content;max-width:none}} .preview th,.preview td{{position:static;min-width:82px;max-width:220px;height:32px;font-size:12px;background:#fff}} .preview th{{background:#f1f5f9}} .preview td{{text-align:right}} .preview td.formula-cell{{font-family:ui-monospace,SFMono-Regular,Menlo,monospace;text-align:left;color:#334155;max-width:360px}} .preview td:first-child,.preview th:first-child{{text-align:left;position:sticky;left:0;z-index:1}} .sheet{{overflow:auto;max-height:calc(100vh - 36px)}} table{{border-collapse:separate;border-spacing:0;min-width:900px;width:max-content}} th,td{{border-right:1px solid #d7dde8;border-bottom:1px solid #d7dde8;min-width:118px;max-width:220px;height:36px;padding:6px 8px;background:#fff;vertical-align:middle;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}} th{{position:sticky;top:0;z-index:2;background:#eef3f8;text-align:center;font-weight:700}} .corner{{left:0;z-index:3}} .row-head{{position:sticky;left:0;z-index:1;min-width:54px;background:#eef3f8;text-align:right}} td{{cursor:pointer}} td:hover{{outline:2px solid #8cc8dc;outline-offset:-2px}} td.selected-row{{background:#e8f7fb}} td.selected-column{{background:#fff4d6}} td.selected-measure{{background:#eaf8ef}} td b{{float:right;color:#1b7895;font-size:11px}} .pill{{display:block;border:1px solid #cbd7e3;border-radius:8px;padding:8px 10px;margin:6px 0;background:#f8fafc}} .pill.missing{{color:#7a8796}} .hint{{font-size:13px}} .note{{padding:10px;border-radius:8px;background:#eef7fb}} .hidden{{display:none}} @media(max-width:980px){{main{{grid-template-columns:1fr}} aside{{max-height:none}} .sheet{{max-height:70vh}}}}
  </style>
</head>
<body>
<main>
  <aside>
    <h1>새 표 만들기</h1>
    <p class="hint">{escape(source["source"]["sheet_title"])} · {escape(source["source"]["qualified_range"])}</p>
    <h2>1. 원하는 결과표 직접 입력</h2>
    <p class="hint">빈 시트에 만들고 싶은 결과표의 좌측 레이블과 상단 레이블을 직접 입력하세요. 원본에서 고르고 싶을 때만 아래 보조 기능을 열면 됩니다.</p>
    <div class="output-canvas-wrap">
      <table id="output-canvas" class="output-canvas"><thead><tr><th class="corner"></th>{canvas_column_headers}</tr></thead><tbody>{canvas_rows_html}</tbody></table>
    </div>
    <label for="llm-prompt">AI에게 요청할 내용</label>
    <textarea id="llm-prompt" class="prompt" placeholder="예: 이 표의 빈 칸에 브랜드별 결제수단별 결제·취소액 합계를 수식으로 채워줘. 취소 건은 음수로 반영하고, 결제상태가 완료인 건만 포함해줘."></textarea>
    <div class="actions one"><button id="submit-intent" type="button">AI에게 이해한 내용 확인하기</button></div>
    <p id="canvas-summary" class="hint note"></p>
    <div id="preview" class="preview" aria-live="polite"></div>
    <details>
      <summary>원본에서 열 찾아 넣기(선택)</summary>
      <div class="panel">
        <p class="hint">직접 입력한 결과표와 AI 요청만으로 부족할 때, 원본 표의 제목 셀을 눌러 AI가 참조할 열을 알려줄 수 있습니다.</p>
        <button class="role active" data-role="row_label" aria-label="왼쪽에 나열할 기준 선택"><span class="role-icon row"></span><span><strong>왼쪽 레이블과 연결할 원본 열</strong> <small>예: 브랜드, 날짜</small></span></button>
        <button class="role" data-role="column_label" aria-label="위쪽으로 펼칠 기준 선택"><span class="role-icon column"></span><span><strong>상단 레이블과 연결할 원본 열</strong> <small>예: 결제수단, 월</small></span></button>
        <button class="role" data-role="measure" aria-label="칸마다 계산할 값 선택"><span class="role-icon measure"></span><span><strong>계산할 값이 있는 원본 열</strong> <small>예: 결제액, 매출, 건수</small></span></button>
        <div id="picked"></div>
      </div>
    </details>
    <h2>2. 새 표 저장하기</h2>
    <label for="sheet-title">새 시트 이름</label>
    <input id="sheet-title" value="{escape(source["default_spec"]["output"]["sheet_title"])}">
    <label for="table-title">표 제목</label>
    <input id="table-title" value="MCP Formula Table">
    <label for="creation-mode">저장 방식</label>
    <select id="creation-mode"></select>
    <p id="mode-note" class="hint note"></p>
    <div id="workbook-path-group">
      <label for="output-workbook-path">복사본 파일 위치</label>
      <input id="output-workbook-path" value="{escape(source["default_spec"]["output"].get("workbook_path", ""))}">
    </div>
    <div id="copy-title-group">
      <label for="copy-title">복사본 이름</label>
      <input id="copy-title" value="{escape(source["default_spec"]["output"].get("copy_title", ""))}">
    </div>
    <details>
      <summary>고급 설정</summary>
      <div class="panel">
        <label for="formula-preset">계산 방식</label>
        <select id="formula-preset">
          <option value="custom">직접 입력</option>
          <option value="sumifs">합계 만들기</option>
          <option value="countifs">개수 세기</option>
          <option value="first_match">첫 번째 값 가져오기</option>
        </select>
        <label for="formula-template">수식 직접 입력</label>
        <textarea id="formula-template" class="formula">{escape(DEFAULT_FORMULA_TEMPLATE)}</textarea>
        <label for="header-row">제목이 있는 줄</label>
        <input id="header-row" type="number" min="1" value="1">
        <label><input id="header-bold" type="checkbox"> 제목 줄 굵게 표시</label>
        <label><input id="freeze-header" type="checkbox"> 제목 줄 고정</label>
        <label><input id="auto-resize" type="checkbox"> 열 너비 자동 맞춤</label>
        <label><input id="protect-created-sheet" type="checkbox"> 새 시트 수정 전 경고 표시</label>
      </div>
    </details>
    <details>
      <summary>AI에 전달할 설정 보기</summary>
      <div class="panel">
        <textarea id="spec"></textarea>
        <div class="actions"><button id="copy">설정 복사</button><button class="secondary" id="download">설정 파일 받기</button></div>
      </div>
    </details>
  </aside>
  <section class="sheet">
    <table id="grid"><thead><tr><th class="corner"></th>{column_headers}</tr></thead><tbody>{rows_html}</tbody></table>
  </section>
</main>
<script>
const DATA = {data_json};
const state = {{ activeRole: 'row_label', fields: {{}}, planSubmitted: false }};
const roleLabels = {{ row_label: '왼쪽 줄', column_label: '윗줄', measure: '계산값' }};
const roleEmptyText = {{
  row_label: '왼쪽에 나열할 기준을 고르세요',
  column_label: '위쪽으로 펼칠 기준을 고르세요',
  measure: '칸마다 계산할 값을 고르세요'
}};
const roleClasses = {{ row_label: 'selected-row', column_label: 'selected-column', measure: 'selected-measure' }};
const formulaPresets = {{
  sumifs: '{escape(DEFAULT_FORMULA_TEMPLATE)}',
  countifs: '=COUNTIFS({{row_label_range}},{{row_label_cell}},{{column_label_range}},{{column_label_cell}})',
  first_match: '=IFERROR(INDEX(FILTER({{measure_range}},{{row_label_range}}={{row_label_cell}},{{column_label_range}}={{column_label_cell}}),1),"")'
}};
const creationMode = document.getElementById('creation-mode');
const workbookPathGroup = document.getElementById('workbook-path-group');
const copyTitleGroup = document.getElementById('copy-title-group');
function setupCreationMode() {{
  const artifactType = DATA.artifact_type || (DATA.source && DATA.source.artifact_type) || 'google_sheets';
  const defaultMode = (DATA.default_spec.output && DATA.default_spec.output.creation_mode) || (artifactType === 'excel_workbook' ? 'copy' : 'sheet');
  const choices = artifactType === 'excel_workbook'
    ? [
        ['copy', '원본을 복사해서 만들기'],
        ['sheet', '이 파일에 새 시트 추가']
      ]
    : [
        ['sheet', '이 문서에 새 시트 추가'],
        ['copy', '문서를 복사해서 만들기']
      ];
  creationMode.innerHTML = choices.map(([value, label]) => `<option value="${{value}}">${{label}}</option>`).join('');
  creationMode.value = choices.some(([value]) => value === defaultMode) ? defaultMode : choices[0][0];
}}
function headerFor(column) {{
  const headerRow = Math.max(1, Number(document.getElementById('header-row').value || 1));
  const row = DATA.grid[headerRow - 1] || [];
  const cell = row.find(item => item.column === column);
  return (cell && cell.formatted_value) || column;
}}
function readOutputCanvas() {{
  const rawRows = Array.from(document.querySelectorAll('#output-canvas tbody tr')).map(row =>
    Array.from(row.querySelectorAll('td')).map(cell => cell.textContent.trim())
  );
  let maxRow = -1;
  let maxColumn = -1;
  rawRows.forEach((row, rowIndex) => {{
    row.forEach((value, columnIndex) => {{
      if (value) {{
        maxRow = Math.max(maxRow, rowIndex);
        maxColumn = Math.max(maxColumn, columnIndex);
      }}
    }});
  }});
  if (maxRow < 0 || maxColumn < 0) return [];
  return rawRows.slice(0, maxRow + 1).map(row => row.slice(0, maxColumn + 1));
}}
function canvasLabels() {{
  const canvas = readOutputCanvas();
  if (!canvas.length) return {{ canvas, rowLabels: [], columnLabels: [] }};
  const columnLabels = (canvas[0] || []).slice(1).map(value => String(value || '').trim()).filter(Boolean);
  const rowLabels = canvas.slice(1).map(row => String((row || [])[0] || '').trim()).filter(Boolean);
  return {{ canvas, rowLabels, columnLabels }};
}}
function refreshCanvasSummary(canvas) {{
  const summary = document.getElementById('canvas-summary');
  const llmPrompt = document.getElementById('llm-prompt').value.trim();
  const filledCells = canvas.flat().filter(Boolean).length;
  if (!filledCells && !llmPrompt) {{
    summary.textContent = '빈 표에 원하는 결과표의 좌측/상단 레이블을 입력하고, 아래에 AI에게 요청할 내용을 적어주세요.';
    return;
  }}
  const rowCount = canvas.length;
  const columnCount = Math.max(...canvas.map(row => row.length), 0);
  summary.textContent = `${{rowCount || 0}}행 x ${{columnCount || 0}}열 출력 모양, 입력된 셀 ${{filledCells}}개${{llmPrompt ? ' · AI 요청 입력됨' : ''}}`;
}}
function escapeHtml(value) {{
  return String(value ?? '').replace(/[&<>"']/g, char => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[char]));
}}
function previewHeaderHtml(title, note) {{
  return `<div class="preview-head"><div class="preview-title"><span>${{escapeHtml(title)}}</span><small>${{escapeHtml(note)}}</small></div></div>`;
}}
function shortList(values) {{
  const visible = values.slice(0, 8).map(escapeHtml).join(', ');
  const more = values.length > 8 ? ` 외 ${{values.length - 8}}개` : '';
  return visible ? `${{visible}}${{more}}` : '아직 명확히 입력되지 않았습니다';
}}
function planPreviewHtml(labelsFromCanvas, llmPrompt) {{
  const pickedFields = ['row_label','column_label','measure']
    .map(role => state.fields[role] ? `${{roleLabels[role]}}: ${{state.fields[role].header}}` : '')
    .filter(Boolean);
  const fieldText = pickedFields.length ? `<li>참고할 원본 열: ${{pickedFields.map(escapeHtml).join(' · ')}}</li>` : '<li>참고할 원본 열: AI가 원본 범위에서 필요한 열과 수식 근거를 찾습니다</li>';
  const promptText = llmPrompt ? escapeHtml(llmPrompt) : '추가 설명이 없습니다. 표 모양만 기준으로 AI가 해석합니다.';
  return `<div class="plan"><h3>제가 이해한 요청입니다. 이 모양 맞나요?</h3><ul><li>상단에 펼칠 항목: ${{shortList(labelsFromCanvas.columnLabels)}}</li><li>왼쪽에 나열할 항목: ${{shortList(labelsFromCanvas.rowLabels)}}</li><li>추가 설명: ${{promptText}}</li>${{fieldText}}<li>다음 작업: 원본 데이터와 기존 수식/참조를 찾아, 계산값을 AI가 직접 산출하지 않고 새 시트의 수식과 참조로 채웁니다</li></ul><p class="plan-note">맞다면 아래의 설정 보기를 열어 AI에게 전달할 JSON을 복사하거나 파일로 받을 수 있습니다. 틀리면 위 표나 설명을 고친 뒤 다시 확인하세요.</p></div>`;
}}
function buildPreview(spec) {{
  const preview = document.getElementById('preview');
  const labelsFromCanvas = canvasLabels();
  const llmPrompt = document.getElementById('llm-prompt').value.trim();
  if (!state.planSubmitted) {{
    preview.innerHTML = previewHeaderHtml('AI 작업 계획 미리보기', '결과표 모양과 설명을 입력한 뒤 버튼을 누르면 AI가 이해한 내용을 여기서 확인합니다.') + `<div class="preview-empty">아직 AI에게 확인 요청을 보내지 않았습니다.</div>`;
    return;
  }}
  if (!labelsFromCanvas.canvas.length && !llmPrompt) {{
    preview.innerHTML = previewHeaderHtml('입력이 필요합니다', '원하는 결과표 모양이나 AI에게 요청할 내용을 먼저 입력해주세요.') + `<div class="preview-empty">빈 표에 원하는 결과표를 직접 적고, 필요한 설명을 한두 문장으로 추가하세요.</div>`;
    return;
  }}
  const body = labelsFromCanvas.canvas.length
    ? outputCanvasPreviewHtml(labelsFromCanvas.canvas)
    : `<div class="preview-empty">표 모양은 비어 있습니다. AI가 설명만으로 필요한 표 모양을 제안해야 합니다.</div>`;
  preview.innerHTML = previewHeaderHtml('이 모양으로 만들까요?', 'AI가 이해한 작업 계획입니다. 표 모양과 요청 내용을 확인하세요.') + body + planPreviewHtml(labelsFromCanvas, llmPrompt);
}}
function outputCanvasPreviewHtml(canvas) {{
  const rows = canvas.slice(0, 8);
  const columnCount = Math.max(...rows.map(row => row.length), 1);
  const header = `<tr>${{Array.from({{length: columnCount}}, (_, index) => `<th>${{escapeHtml(index === 0 ? '좌측/상단' : '열 ' + index)}}</th>`).join('')}}</tr>`;
  const body = rows.map(row => `<tr>${{Array.from({{length: columnCount}}, (_, index) => `<td>${{escapeHtml(row[index] || '')}}</td>`).join('')}}</tr>`).join('');
  return `<div class="preview-scroll"><table><thead>${{header}}</thead><tbody>${{body}}</tbody></table></div>`;
}}
function refresh() {{
  const spec = JSON.parse(JSON.stringify(DATA.default_spec));
  const outputCanvas = readOutputCanvas();
  spec.source.header_row = Math.max(1, Number(document.getElementById('header-row').value || 1));
  spec.fields = state.fields;
  spec.output_canvas = outputCanvas;
  spec.llm_prompt = document.getElementById('llm-prompt').value.trim();
  spec.formula = {{ template: document.getElementById('formula-template').value }};
  spec.output.sheet_title = document.getElementById('sheet-title').value || DATA.default_spec.output.sheet_title;
  spec.output.title = document.getElementById('table-title').value || 'MCP Formula Table';
  spec.output.format = {{
    header_bold: document.getElementById('header-bold').checked,
    freeze_header_rows: document.getElementById('freeze-header').checked ? 2 : 0,
    auto_resize_columns: document.getElementById('auto-resize').checked,
    protect_created_sheet: document.getElementById('protect-created-sheet').checked
  }};
  spec.output.creation_mode = creationMode.value || DATA.default_spec.output.creation_mode || 'sheet';
  if (spec.output.creation_mode === 'copy') {{
    if (spec.artifact_type === 'excel_workbook') {{
      spec.output.workbook_path = document.getElementById('output-workbook-path').value || DATA.default_spec.output.workbook_path || '';
      delete spec.output.copy_title;
    }} else {{
      spec.output.copy_title = document.getElementById('copy-title').value || DATA.default_spec.output.copy_title || '';
      delete spec.output.workbook_path;
    }}
  }} else {{
    delete spec.output.workbook_path;
    delete spec.output.copy_title;
  }}
  workbookPathGroup.classList.toggle('hidden', spec.output.creation_mode !== 'copy' || spec.artifact_type !== 'excel_workbook');
  copyTitleGroup.classList.toggle('hidden', spec.output.creation_mode !== 'copy' || spec.artifact_type === 'excel_workbook');
  const modeNote = document.getElementById('mode-note');
  if (spec.output.creation_mode === 'copy') {{
    modeNote.textContent = spec.artifact_type === 'excel_workbook'
      ? '원본 파일은 그대로 두고, 복사본에 새 표 시트를 만듭니다.'
      : '원본 문서는 그대로 두고, 복사한 문서에 새 표 시트를 만듭니다.';
  }} else {{
    modeNote.textContent = '현재 파일 안에 새 시트를 추가합니다. 기존 시트의 내용은 바꾸지 않습니다.';
  }}
  document.getElementById('spec').value = JSON.stringify(spec, null, 2);
  document.getElementById('picked').innerHTML = ['row_label','column_label','measure'].map(role => {{
    const field = state.fields[role];
    const text = field ? `${{roleLabels[role]}}: ${{field.header}} (${{field.column}})` : roleEmptyText[role];
    return `<span class="pill ${{field ? '' : 'missing'}}">${{text}}</span>`;
  }}).join('');
  refreshCanvasSummary(outputCanvas);
  buildPreview(spec);
  document.querySelectorAll('td').forEach(td => {{
    td.classList.remove('selected-row','selected-column','selected-measure');
    for (const [role, field] of Object.entries(state.fields)) {{
      if (field.column === td.dataset.column) td.classList.add(roleClasses[role]);
    }}
  }});
}}
document.querySelectorAll('.role').forEach(button => {{
  button.addEventListener('click', () => {{
    document.querySelectorAll('.role').forEach(item => item.classList.remove('active'));
    button.classList.add('active');
    state.activeRole = button.dataset.role;
  }});
}});
document.querySelectorAll('td').forEach(td => {{
  td.addEventListener('click', () => {{
    state.fields[state.activeRole] = {{ column: td.dataset.column, header: headerFor(td.dataset.column), selected_cell: td.dataset.a1 }};
    state.planSubmitted = false;
    refresh();
  }});
}});
function markIntentDraft() {{
  state.planSubmitted = false;
  refresh();
}}
document.getElementById('submit-intent').addEventListener('click', () => {{
  state.planSubmitted = true;
  refresh();
}});
document.getElementById('output-canvas').addEventListener('input', markIntentDraft);
document.getElementById('llm-prompt').addEventListener('input', markIntentDraft);
['formula-template','header-row','sheet-title','table-title','output-workbook-path','copy-title'].forEach(id => document.getElementById(id).addEventListener('input', refresh));
['header-bold','freeze-header','auto-resize','protect-created-sheet'].forEach(id => document.getElementById(id).addEventListener('change', refresh));
document.getElementById('formula-preset').addEventListener('change', event => {{
  const value = event.target.value;
  if (formulaPresets[value]) {{
    document.getElementById('formula-template').value = formulaPresets[value];
  }}
  refresh();
}});
creationMode.addEventListener('change', refresh);
document.getElementById('copy').addEventListener('click', async () => {{
  await navigator.clipboard.writeText(document.getElementById('spec').value);
}});
document.getElementById('download').addEventListener('click', () => {{
  const blob = new Blob([document.getElementById('spec').value], {{type:'application/json'}});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = 'spreadsheet-table-spec.json';
  a.click();
  URL.revokeObjectURL(a.href);
}});
setupCreationMode();
document.getElementById('header-bold').checked = Boolean(DATA.default_spec.output.format && DATA.default_spec.output.format.header_bold);
document.getElementById('freeze-header').checked = Number(DATA.default_spec.output.format && DATA.default_spec.output.format.freeze_header_rows || 0) > 0;
document.getElementById('auto-resize').checked = Boolean(DATA.default_spec.output.format && DATA.default_spec.output.format.auto_resize_columns);
document.getElementById('protect-created-sheet').checked = Boolean(DATA.default_spec.output.format && DATA.default_spec.output.format.protect_created_sheet);
refresh();
</script>
</body></html>
"""


def _build_formula_table_result_html(result: dict[str, Any]) -> str:
    validation = result.get("validation", {}) if isinstance(result.get("validation"), dict) else {}
    artifact_type = str(result.get("artifact_type") or "")
    if artifact_type == "excel_workbook":
        location_label = "Workbook"
        location_html = f"<code>{escape(str(result.get('workbook_path', '')))}</code>"
        rollback = result.get("rollback") if isinstance(result.get("rollback"), dict) else {}
        if rollback.get("type") == "delete_output_workbook_copy":
            rollback_text = f"Delete workbook copy <code>{escape(str(rollback.get('workbook_path', '')))}</code>."
        else:
            rollback_text = f"Delete created worksheet <code>{escape(str(rollback.get('sheet_title', '')))}</code>."
    else:
        location_label = "URL"
        location = escape(str(result.get("new_sheet_url", "")))
        location_html = f'<a href="{location}">{location}</a>'
        rollback_text = f"Delete created sheet id <code>{escape(str((result.get('rollback') or {}).get('sheet_id', '')))}</code>."
    return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><title>Formula Table Result</title><style>body{{font-family:Arial,'Apple SD Gothic Neo',sans-serif;margin:32px;color:#172033}}main{{max-width:900px}}table{{border-collapse:collapse;width:100%}}th,td{{border-bottom:1px solid #d8e1ea;padding:10px;text-align:left}}th{{width:220px;background:#eef3f8}}code{{background:#eef3f8;padding:2px 5px;border-radius:4px}}</style></head>
<body><main><h1>Formula Table Result</h1><table>
<tr><th>New sheet</th><td>{escape(str(result.get("new_sheet_title", "")))}</td></tr>
<tr><th>{location_label}</th><td>{location_html}</td></tr>
<tr><th>Source range</th><td><code>{escape(str(result.get("source_range", "")))}</code></td></tr>
<tr><th>Formula template</th><td><code>{escape(str(result.get("formula_template", "")))}</code></td></tr>
<tr><th>Formula cells</th><td>{escape(str(result.get("formula_cell_count", "")))}</td></tr>
<tr><th>Rollback</th><td>{rollback_text}</td></tr>
</table><h2>Validation</h2><pre>{escape(json.dumps(validation, ensure_ascii=False, indent=2))}</pre></main></body></html>"""


def _excel_source_descriptor(
    *,
    workbook_path: str,
    sheet_name: str,
    source_range: str,
    max_rows: int,
    max_columns: int,
    created_at: str,
) -> dict[str, Any]:
    path = Path(workbook_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"missing workbook: {path}")
    wb_formula = load_workbook(path, read_only=True, data_only=False)
    wb_values = load_workbook(path, read_only=True, data_only=True)
    try:
        ws_formula = _excel_sheet(wb_formula, sheet_name)
        ws_values = wb_values[ws_formula.title]
        qualified_range = _excel_qualified_range(
            ws=ws_formula,
            source_range=source_range,
            max_rows=max_rows,
            max_columns=max_columns,
        )
        values = _excel_range_matrix(ws_values, qualified_range, prefer_formula=False)
        formulas = _excel_range_matrix(ws_formula, qualified_range, prefer_formula=True)
        start_col, start_row, _end_col, _end_row = _range_bounds(qualified_range)
        sheet_suffix = created_at[:16].replace("-", "").replace(":", "").replace("T", "_")
        return {
            "schema_version": "1.0",
            "artifact_kind": "spreadsheet_table_builder_source",
            "created_at": created_at,
            "artifact_type": "excel_workbook",
            "spreadsheet_id": "",
            "workbook_path": str(path),
            "workbook_title": path.name,
            "source": {
                "artifact_type": "excel_workbook",
                "spreadsheet_id": "",
                "workbook_path": str(path),
                "sheet_title": ws_formula.title,
                "qualified_range": qualified_range,
                "range_start_row": start_row,
                "range_start_column": _column_label(start_col),
                "header_row": 1,
            },
            "grid": _grid_cells(values=values, formulas=formulas, start_col=start_col, start_row=start_row),
            "default_spec": {
                "schema_version": "1.0",
                "spec_kind": "formula_table_apply_v1",
                "artifact_type": "excel_workbook",
                "spreadsheet_id": "",
                "source": {
                    "artifact_type": "excel_workbook",
                    "spreadsheet_id": "",
                    "workbook_path": str(path),
                    "sheet_title": ws_formula.title,
                    "qualified_range": qualified_range,
                    "header_row": 1,
                },
                "fields": {},
                "formula": {"template": DEFAULT_FORMULA_TEMPLATE},
                "output": {
                    "sheet_title": f"MCP_TABLE_{sheet_suffix}",
                    "title": "MCP Formula Table",
                    "creation_mode": "copy",
                    "workbook_path": str(_excel_output_path(path, "")),
                    "copy_title": "",
                    "format": dict(DEFAULT_OUTPUT_FORMAT),
                },
            },
        }
    finally:
        wb_formula.close()
        wb_values.close()


def _excel_sheet(workbook: Any, sheet_name: str) -> Any:
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"sheet not found in workbook: {sheet_name}")
        return workbook[sheet_name]
    return workbook.active


def _excel_qualified_range(*, ws: Any, source_range: str, max_rows: int, max_columns: int) -> str:
    if source_range:
        if "!" in source_range:
            title = _sheet_title_from_qualified_range(source_range)
            if title != ws.title:
                raise ValueError(f"source_range sheet does not match selected sheet: {title}")
            return source_range
        return f"{quote_sheet_title(ws.title)}!{source_range.replace('$', '').upper()}"
    rows = min(max(int(ws.max_row or 1), 1), max_rows)
    columns = min(max(int(ws.max_column or 1), 1), max_columns)
    return f"{quote_sheet_title(ws.title)}!A1:{_column_label(columns)}{rows}"


def _excel_range_matrix(ws: Any, qualified_range: str, *, prefer_formula: bool) -> list[list[Any]]:
    min_col, min_row, max_col, max_row = _range_bounds(qualified_range)
    matrix = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        values = []
        for cell in row:
            value = cell.value
            if prefer_formula:
                values.append(value if isinstance(value, str) and value.startswith("=") else "")
            else:
                values.append(value if value is not None else "")
        matrix.append(values)
    return matrix


def _apply_excel_output_format(ws: Any, grid: list[list[Any]], format_options: dict[str, Any]) -> None:
    options = _normalize_output_format(format_options)
    if options["header_bold"]:
        for row_index in range(1, min(2, len(grid)) + 1):
            for cell in ws[row_index]:
                font = copy(cell.font)
                font.bold = True
                cell.font = font
    freeze_rows = int(options["freeze_header_rows"])
    if freeze_rows > 0:
        ws.freeze_panes = f"A{min(freeze_rows, max(len(grid), 1)) + 1}"
    if options["auto_resize_columns"]:
        column_count = max((len(row) for row in grid), default=0)
        for column_index in range(1, column_count + 1):
            values = [str(ws.cell(row=row_index, column=column_index).value or "") for row_index in range(1, len(grid) + 1)]
            width = min(max(max((len(value) for value in values), default=8) + 2, 10), 48)
            ws.column_dimensions[_column_label(column_index)].width = width
    if options["protect_created_sheet"]:
        ws.protection.sheet = True


def _write_excel_formula_table_with_openpyxl(
    *,
    workbook_path: Path,
    sheet_title: str,
    write_grid: list[list[Any]],
    format_options: dict[str, Any],
) -> None:
    wb = load_workbook(workbook_path)
    try:
        if sheet_title in wb.sheetnames:
            raise ValueError(f"worksheet already exists: {sheet_title}")
        ws = wb.create_sheet(sheet_title)
        _populate_openpyxl_sheet(ws, write_grid, format_options)
        wb.save(workbook_path)
    finally:
        wb.close()


def _write_excel_formula_table_with_desktop_excel(
    *,
    target_path: Path,
    sheet_title: str,
    write_grid: list[list[Any]],
    format_options: dict[str, Any],
    timeout_seconds: int = 180,
    runner: Callable[..., subprocess.CompletedProcess[str]] | None = None,
) -> None:
    if not _desktop_excel_available():
        raise RuntimeError("desktop Excel automation is not available on this OS")
    runner = runner or subprocess.run
    script_path, command_prefix = _desktop_excel_copy_sheet_command()
    with tempfile.TemporaryDirectory(prefix="sheets_bridge_excel_write_", dir=str(_desktop_excel_temp_root())) as tmp_dir:
        tmp_root = Path(tmp_dir)
        temp_target = tmp_root / f"target-{uuid4().hex}{target_path.suffix}"
        temp_template = tmp_root / f"formula-table-template-{uuid4().hex}.xlsx"
        temp_formula_map = tmp_root / f"formula-rewrite-{uuid4().hex}.tsv"
        shutil.copy2(target_path, temp_target)
        _create_excel_formula_table_template(
            workbook_path=temp_template,
            sheet_title=sheet_title,
            write_grid=write_grid,
            format_options=format_options,
        )
        _write_excel_formula_rewrite_map(temp_formula_map, write_grid)
        command = [
            *command_prefix,
            str(script_path),
            *(
                [str(temp_target), str(temp_template), sheet_title, str(temp_formula_map)]
                if sys.platform == "darwin"
                else [
                    "-TargetWorkbook",
                    str(temp_target),
                    "-TemplateWorkbook",
                    str(temp_template),
                    "-TemplateSheetName",
                    sheet_title,
                    "-FormulaMap",
                    str(temp_formula_map),
                ]
            ),
        ]
        completed = runner(command, check=False, capture_output=True, text=True, timeout=timeout_seconds)
        if completed.returncode != 0:
            detail = (completed.stderr or completed.stdout or "").strip()
            raise RuntimeError(f"desktop Excel sheet copy failed: {detail}")
        shutil.copy2(temp_target, target_path)


def _create_excel_formula_table_template(
    *,
    workbook_path: Path,
    sheet_title: str,
    write_grid: list[list[Any]],
    format_options: dict[str, Any],
) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    try:
        ws = wb.active
        ws.title = sheet_title
        _populate_openpyxl_sheet(ws, write_grid, format_options)
        wb.save(workbook_path)
    finally:
        wb.close()


def _write_excel_formula_rewrite_map(path: Path, write_grid: list[list[Any]]) -> None:
    lines: list[str] = []
    for row_index, row in enumerate(write_grid, start=1):
        for column_index, value in enumerate(row, start=1):
            if isinstance(value, str) and value.startswith("="):
                if "\t" in value or "\r" in value or "\n" in value:
                    raise ValueError("desktop Excel formula rewrite map does not support formulas containing tabs or line breaks")
                lines.append(f"{_column_label(column_index)}{row_index}\t{value}")
    path.write_text(("\n".join(lines) + "\n") if lines else "", encoding="utf-8")


def _populate_openpyxl_sheet(ws: Any, write_grid: list[list[Any]], format_options: dict[str, Any]) -> None:
    for row_index, row in enumerate(write_grid, start=1):
        for column_index, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=column_index, value=value)
    _apply_excel_output_format(ws, write_grid, format_options)


def _excel_copy_write_engine(source_path: Path) -> str:
    requested = os.environ.get("SHEETS_BRIDGE_EXCEL_WRITE_ENGINE", "auto").strip().lower()
    aliases = {"desktop": "desktop_excel", "excel": "desktop_excel"}
    requested = aliases.get(requested, requested)
    if requested not in {"auto", "openpyxl", "desktop_excel"}:
        raise ValueError("SHEETS_BRIDGE_EXCEL_WRITE_ENGINE must be auto, openpyxl, or desktop")
    if requested != "auto":
        if requested == "desktop_excel" and not _desktop_excel_available():
            raise RuntimeError("desktop Excel automation is not available on this OS")
        return requested
    threshold = int(os.environ.get("SHEETS_BRIDGE_EXCEL_DESKTOP_WRITE_THRESHOLD_BYTES", DEFAULT_EXCEL_DESKTOP_WRITE_THRESHOLD_BYTES))
    if source_path.stat().st_size >= threshold:
        if not _desktop_excel_available():
            raise RuntimeError(
                "large Excel workbook copy mode requires desktop Excel automation; "
                "set SHEETS_BRIDGE_EXCEL_WRITE_ENGINE=openpyxl only if a full workbook round-trip is acceptable"
            )
        return "desktop_excel"
    return "openpyxl"


def _desktop_excel_available() -> bool:
    if sys.platform == "darwin":
        return shutil.which("osascript") is not None and _script_path("excel_copy_sheet_into_workbook.applescript").exists()
    if os.name == "nt":
        return (shutil.which("pwsh") or shutil.which("powershell.exe")) is not None and _script_path("excel_copy_sheet_into_workbook.ps1").exists()
    return False


def _desktop_excel_copy_sheet_command() -> tuple[Path, list[str]]:
    if sys.platform == "darwin":
        return _script_path("excel_copy_sheet_into_workbook.applescript"), ["/usr/bin/osascript"]
    if os.name == "nt":
        powershell = shutil.which("pwsh") or shutil.which("powershell.exe")
        if not powershell:
            raise RuntimeError("PowerShell executable not found")
        return _script_path("excel_copy_sheet_into_workbook.ps1"), [powershell, "-NoProfile", "-ExecutionPolicy", "Bypass", "-File"]
    raise RuntimeError("desktop Excel sheet copy is supported only on macOS or Windows")


def _desktop_excel_temp_root() -> Path:
    if sys.platform == "darwin" and MAC_EXCEL_CONTAINER_DOCUMENTS.exists():
        root = MAC_EXCEL_CONTAINER_DOCUMENTS / "excel_workbook_editing_write"
    else:
        root = Path(tempfile.gettempdir()) / "excel_workbook_editing_write"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _script_path(name: str) -> Path:
    return resource_script_path(name)


def _excel_values_for_range(*, workbook_path: Path, sheet_title: str, qualified_range: str) -> list[list[Any]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_title not in wb.sheetnames:
            raise ValueError(f"sheet not found in workbook: {sheet_title}")
        return _excel_range_matrix(wb[sheet_title], qualified_range, prefer_formula=False)
    finally:
        wb.close()


def _excel_output_path(source_path: Path, requested: str) -> Path:
    if requested:
        output_path = Path(requested).expanduser().resolve()
    else:
        output_path = source_path.with_name(f"{source_path.stem}-formula-table{source_path.suffix}")
    if output_path == source_path:
        raise ValueError("output workbook path must differ from the source workbook path when creation_mode='copy'")
    if output_path.exists():
        stem = output_path.stem
        suffix = output_path.suffix
        parent = output_path.parent
        for index in range(2, 1000):
            candidate = parent / f"{stem}-{index}{suffix}"
            if not candidate.exists():
                return candidate
        raise ValueError(f"could not allocate output workbook path from: {output_path}")
    return output_path


def _unique_excel_sheet_title(existing: list[str], desired: str) -> str:
    if desired not in existing:
        return desired
    for index in range(2, 100):
        candidate = f"{desired} ({index})"
        if candidate not in existing:
            return candidate
    raise ValueError(f"could not allocate unique worksheet title from: {desired}")


def _excel_formula_table_static_validation(
    *,
    workbook_path: Path,
    sheet_title: str,
    expected_formula_count: int,
) -> dict[str, Any]:
    wb = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        ws = wb[sheet_title]
        formula_count = 0
        formula_errors = []
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                    if any(token in value for token in ERROR_PREFIXES):
                        formula_errors.append({"cell": cell.coordinate, "formula": value})
        return {
            "status": "passed" if formula_count == expected_formula_count and not formula_errors else "review_required",
            "formula_text_verification": "passed" if formula_count == expected_formula_count else "review_required",
            "expected_formula_count": expected_formula_count,
            "actual_formula_count": formula_count,
            "formula_error_token_count": len(formula_errors),
            "sample_formula_error_tokens": formula_errors[:20],
            "formula_result_validation": "not_run",
            "formula_result_validation_note": "Excel formula results require Microsoft Excel recalculation; openpyxl only verifies written formula text.",
        }
    finally:
        wb.close()


def _excel_formula_error_scan(workbook_path: Path, *, sample_limit: int = 20) -> dict[str, Any]:
    wb_formula = load_workbook(workbook_path, read_only=True, data_only=False)
    wb_values = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheets = []
        for ws_formula in wb_formula.worksheets:
            ws_values = wb_values[ws_formula.title]
            formula_count = 0
            error_count = 0
            samples = []
            for formula_row, values_row in zip_longest(ws_formula.iter_rows(), ws_values.iter_rows(), fillvalue=()):
                for cell, value_cell in zip_longest(formula_row, values_row, fillvalue=None):
                    coordinate = getattr(cell, "coordinate", "") or getattr(value_cell, "coordinate", "")
                    if not coordinate:
                        continue
                    value = getattr(cell, "value", None)
                    is_formula = isinstance(value, str) and value.startswith("=")
                    cached_value = getattr(value_cell, "value", None)
                    has_error = _cell_text(value).startswith(ERROR_PREFIXES) or _cell_text(cached_value).startswith(ERROR_PREFIXES)
                    if is_formula:
                        formula_count += 1
                    if has_error:
                        error_count += 1
                        if len(samples) < sample_limit:
                            samples.append({"sheet": ws_formula.title, "cell": coordinate, "formula": value, "cached_value": cached_value})
            sheets.append({"sheet": ws_formula.title, "formula_count": formula_count, "error_count": error_count, "samples": samples})
        summary = {
            "formula_count": sum(item["formula_count"] for item in sheets),
            "error_count": sum(item["error_count"] for item in sheets),
        }
        summary["status"] = "passed" if summary["error_count"] == 0 else "failed"
        return {"summary": summary, "sheets": sheets}
    finally:
        wb_formula.close()
        wb_values.close()


def _parse_excel_engine_output(output: str) -> dict[str, str]:
    values = {}
    for line in str(output).splitlines():
        if "=" not in line:
            continue
        cell, value = line.split("=", 1)
        values[cell] = value
    return values


def _excel_engine_sample_command(path: Path, worksheet: str, cells: list[str]) -> list[str]:
    return [
        sys.executable,
        str(resource_script_path("excel_engine_sample.py")),
        str(path),
        worksheet,
        *cells,
    ]


def _qualified_source_range(
    *,
    metadata: dict[str, Any],
    gid: str,
    source_range: str,
    max_rows: int,
    max_columns: int,
) -> str:
    if source_range:
        return qualify_ranges(metadata, ranges=[source_range], gid=gid)[0]
    title = sheet_title_for_gid(metadata, gid) if gid else first_visible_sheet_title(metadata)
    sheet = _sheet_summary_for_title(metadata, title)
    rows = min(max(int(sheet.get("row_count") or 0), 1), max_rows)
    columns = min(max(int(sheet.get("column_count") or 0), 1), max_columns)
    return f"{quote_sheet_title(title)}!A1:{_column_label(columns)}{rows}"


def _sheet_summary_for_title(metadata: dict[str, Any], title: str) -> dict[str, Any]:
    for sheet in metadata.get("sheets", []) or []:
        props = sheet.get("properties", {}) if isinstance(sheet, dict) else {}
        if str(props.get("title", "")) == title:
            grid = props.get("gridProperties", {}) if isinstance(props.get("gridProperties"), dict) else {}
            return {
                "sheet_id": props.get("sheetId", 0),
                "title": props.get("title", ""),
                "row_count": grid.get("rowCount", 0),
                "column_count": grid.get("columnCount", 0),
            }
    raise ValueError(f"sheet title not found in metadata: {title}")


def _extract_layout_labels(spec: dict[str, Any], values: list[list[Any]]) -> dict[str, Any]:
    canvas_labels = _output_canvas_labels(spec.get("output_canvas"))
    if canvas_labels["row_labels"] and canvas_labels["column_labels"]:
        return {**canvas_labels, "source": "output_canvas"}

    source = spec["source"]
    start_col, _start_row, _end_col, _end_row = _range_bounds(source["qualified_range"])
    header_row = int(source["header_row"])
    row_column_index = _column_index(spec["fields"]["row_label"]["column"]) - start_col
    column_column_index = _column_index(spec["fields"]["column_label"]["column"]) - start_col
    row_labels = []
    column_labels = []
    for row in values[header_row:]:
        if not isinstance(row, list):
            continue
        row_label = _cell_text(row[row_column_index] if 0 <= row_column_index < len(row) else "")
        column_label = _cell_text(row[column_column_index] if 0 <= column_column_index < len(row) else "")
        if row_label and row_label not in row_labels:
            row_labels.append(row_label)
        if column_label and column_label not in column_labels:
            column_labels.append(column_label)
    return {"row_labels": row_labels, "column_labels": column_labels, "source": "source_values"}


def _normalize_output_canvas(value: Any) -> list[list[str]]:
    if not isinstance(value, list):
        return []
    rows: list[list[str]] = []
    max_column = -1
    for raw_row in value:
        if not isinstance(raw_row, list):
            continue
        row = [_cell_text(cell) for cell in raw_row]
        rows.append(row)
        for column_index, cell in enumerate(row):
            if cell:
                max_column = max(max_column, column_index)
    while rows and not any(rows[-1]):
        rows.pop()
    if not rows or max_column < 0:
        return []
    return [row[: max_column + 1] for row in rows]


def _output_canvas_labels(value: Any) -> dict[str, list[str]]:
    canvas = _normalize_output_canvas(value)
    if not canvas:
        return {"row_labels": [], "column_labels": []}
    column_labels = [
        label for label in (_cell_text(cell) for cell in canvas[0][1:]) if label
    ]
    row_labels = [
        label for label in (_cell_text(row[0] if row else "") for row in canvas[1:]) if label
    ]
    return {
        "row_labels": _dedupe_preserving_order(row_labels),
        "column_labels": _dedupe_preserving_order(column_labels),
    }


def _dedupe_preserving_order(values: list[str]) -> list[str]:
    deduped: list[str] = []
    for value in values:
        if value not in deduped:
            deduped.append(value)
    return deduped


def _normalize_field(value: Any, role: str) -> dict[str, str]:
    if not isinstance(value, dict):
        raise ValueError(f"spec.fields.{role} is required")
    column = str(value.get("column") or "").upper()
    if not re.match(r"^[A-Z]{1,4}$", column):
        raise ValueError(f"spec.fields.{role}.column must be a column label")
    return {
        "column": column,
        "header": str(value.get("header") or column),
        "selected_cell": str(value.get("selected_cell") or ""),
    }


def _formula_template(spec: dict[str, Any]) -> str:
    formula = spec.get("formula") if isinstance(spec.get("formula"), dict) else {}
    template = str(formula.get("template") or spec.get("formula_template") or "")
    if not template:
        template = DEFAULT_FORMULA_TEMPLATE
    if not template.strip().startswith("="):
        raise ValueError("formula template must start with '='")
    return template.strip()


def _normalize_output_format(value: dict[str, Any]) -> dict[str, Any]:
    return {
        "header_bold": bool(value.get("header_bold", DEFAULT_OUTPUT_FORMAT["header_bold"])),
        "freeze_header_rows": max(0, int(value.get("freeze_header_rows", DEFAULT_OUTPUT_FORMAT["freeze_header_rows"]) or 0)),
        "auto_resize_columns": bool(value.get("auto_resize_columns", DEFAULT_OUTPUT_FORMAT["auto_resize_columns"])),
        "protect_created_sheet": bool(value.get("protect_created_sheet", DEFAULT_OUTPUT_FORMAT["protect_created_sheet"])),
    }


def _render_formula_template(
    *,
    template: str,
    measure_range: str,
    row_range: str,
    column_range: str,
    row_criteria: str,
    column_criteria: str,
    row_value: str,
    column_value: str,
    source_sheet: str,
    source_range: str,
    output_sheet: str,
) -> str:
    replacements = {
        "measure_range": measure_range,
        "row_label_range": row_range,
        "column_label_range": column_range,
        "row_label_cell": row_criteria,
        "column_label_cell": column_criteria,
        "row_label_value": _formula_string(row_value),
        "column_label_value": _formula_string(column_value),
        "source_sheet": source_sheet,
        "source_range": source_range,
        "output_sheet": output_sheet,
    }
    rendered = template
    for key, value in replacements.items():
        rendered = rendered.replace("{" + key + "}", value)
    return rendered


def _bounded_column_range(sheet_ref: str, column: str, start_row: int, end_row: int) -> str:
    return f"{sheet_ref}!${column}${start_row}:${column}${end_row}"


def _formula_table_readback_validation(values: list[list[Any]], expected_rows: int, expected_columns: int) -> dict[str, Any]:
    errors = []
    for row_index, row in enumerate(values, start=1):
        if not isinstance(row, list):
            continue
        for column_index, value in enumerate(row, start=1):
            text = _cell_text(value)
            if text.startswith(ERROR_PREFIXES):
                errors.append({"row": row_index, "column": column_index, "value": text})
    return {
        "status": "passed" if not errors else "failed",
        "expected_rows": expected_rows,
        "expected_columns": expected_columns,
        "readback_rows": len(values),
        "error_count": len(errors),
        "sample_errors": errors[:20],
    }


def _write_manifest(
    *,
    manifest_path: Path,
    handoff_path: Path,
    created_at: str,
    request_id: str,
    package_dir: Path,
    primary_kind: str,
    primary_path: Path,
    extra_artifacts: list[dict[str, str]],
) -> None:
    handoff = {
        "schema_version": "1.0",
        "artifact_kind": "spreadsheet_table_builder_mcp_handoff",
        "request_id": request_id,
        "created_at": created_at,
        "package_dir": str(package_dir.resolve()),
        "manifest_path": str(manifest_path.resolve()),
        "primary_artifact": {"kind": primary_kind, "path": str(primary_path.resolve())},
        "mcp_prompt": f"이 Spreadsheet table-builder 패키지를 검토해줘: {manifest_path.resolve()}",
        "analysis_boundary": [
            "Read manifest.json first.",
            "Use only sanitized local artifacts referenced by the manifest.",
            "Use only credential-free MCP outputs and review artifacts.",
        ],
    }
    manifest = {
        "schema_version": "1.0",
        "artifact_kind": "spreadsheet_table_builder_mcp_package",
        "request_id": request_id,
        "created_at": created_at,
        "source": "mcp_spreadsheet_table_builder",
        "artifacts": [
            {"kind": primary_kind, "path": str(primary_path.resolve())},
            *extra_artifacts,
            {"kind": "mcp_handoff", "path": str(handoff_path.resolve())},
        ],
    }
    _write_json(handoff_path, handoff)
    _write_json(manifest_path, manifest)


def _unique_package_dir(package_root: Path | str, created_at: str, request_id: str) -> Path:
    root = Path(package_root)
    base = root / created_at[:10] / _safe_id(request_id)
    candidate = base
    index = 2
    while candidate.exists():
        candidate = Path(f"{base}-{index}")
        index += 1
    candidate.mkdir(parents=True, exist_ok=False)
    return candidate


def _write_json(path: Path, value: dict[str, Any]) -> None:
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _first_value_matrix(snapshot: dict[str, Any]) -> list[list[Any]]:
    ranges = snapshot.get("valueRanges", []) if isinstance(snapshot.get("valueRanges"), list) else []
    if not ranges:
        return []
    values = ranges[0].get("values", []) if isinstance(ranges[0], dict) else []
    return values if isinstance(values, list) else []


def _sheet_title_from_qualified_range(qualified_range: str) -> str:
    title = str(qualified_range).split("!", 1)[0]
    if title.startswith("'") and title.endswith("'"):
        return title[1:-1].replace("''", "'")
    return title


def _range_bounds(a1_range: str) -> tuple[int, int, int, int]:
    coordinate = str(a1_range).split("!", 1)[-1].replace("$", "").upper()
    start, separator, end = coordinate.partition(":")
    end = end if separator else start
    start_col, start_row = _cell_coordinate(start)
    end_col, end_row = _cell_coordinate(end)
    if end_col < start_col or end_row < start_row:
        raise ValueError(f"range must be bounded A1: {a1_range}")
    return start_col, start_row, end_col, end_row


def _cell_coordinate(cell: str) -> tuple[int, int]:
    match = re.match(r"^([A-Z]{1,4})([1-9][0-9]*)$", str(cell).upper())
    if not match:
        raise ValueError(f"range must be bounded A1: {cell}")
    return _column_index(match.group(1)), int(match.group(2))


def _column_index(column: str) -> int:
    index = 0
    for char in str(column).upper():
        if not ("A" <= char <= "Z"):
            raise ValueError(f"invalid column label: {column}")
        index = index * 26 + ord(char) - 64
    if index < 1:
        raise ValueError(f"invalid column label: {column}")
    return index


def _column_label(index: int) -> str:
    if index < 1:
        raise ValueError("column index must be positive")
    label = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        label = chr(65 + remainder) + label
    return label


def _unique_sheet_title(metadata: dict[str, Any], desired: str) -> str:
    existing = {
        str((sheet.get("properties") or {}).get("title", ""))
        for sheet in metadata.get("sheets", []) or []
        if isinstance(sheet, dict)
    }
    if desired not in existing:
        return desired
    for index in range(2, 100):
        candidate = f"{desired} ({index})"
        if candidate not in existing:
            return candidate
    raise ValueError(f"could not allocate unique sheet title from: {desired}")


def _unique_sheet_id(metadata: dict[str, Any]) -> int:
    existing = {
        int((sheet.get("properties") or {}).get("sheetId", 0))
        for sheet in metadata.get("sheets", []) or []
        if isinstance(sheet, dict)
    }
    while True:
        candidate = uuid4().int % 900_000_000 + 10_000_000
        if candidate not in existing:
            return candidate


def _count_formula_cells(values: list[list[Any]]) -> int:
    return sum(1 for row in values for value in row if str(value).startswith("="))


def _safe_id(value: object) -> str:
    raw = "".join(ch if ch.isalnum() or ch in "-_" else "-" for ch in str(value))
    return raw[:120] or "mcp-package"


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _formula_string(value: str) -> str:
    return '"' + str(value).replace('"', '""') + '"'


def _sheet_url(spreadsheet_id: str, sheet_id: int) -> str:
    return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit?gid={sheet_id}#gid={sheet_id}"
