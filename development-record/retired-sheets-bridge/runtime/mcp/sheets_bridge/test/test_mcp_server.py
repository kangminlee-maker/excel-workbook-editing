from __future__ import annotations

from datetime import UTC, datetime, timedelta
import json
from pathlib import Path
import subprocess
import sys
import tempfile
import unittest
from unittest import mock

from openpyxl import Workbook, load_workbook


REPO_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(REPO_ROOT / "mcp"))

from sheets_bridge import auth
from sheets_bridge import table_builder as table_builder_module
from sheets_bridge.chrome_resolver import resolve_current_sheet
from sheets_bridge.mcp_server import handle_message
from sheets_bridge.packages import write_inspection_package
from sheets_bridge.sheets_api import (
    apply_values_update,
    inspect_sheet,
    parse_spreadsheet_url,
    qualify_ranges,
)
from sheets_bridge.table_flow import (
    build_minimal_formula_write_requests,
    compare_value_matrices,
    refactor_minimal_formula_sheet,
    visualize_table_io,
)
from sheets_bridge.table_builder import (
    build_table_builder_ui,
    create_formula_table_from_spec,
    normalize_formula_table_spec,
    rollback_created_artifact,
    validate_excel_formula_results,
)
from sheets_bridge.version import __version__


class McpSheetsBridgeTests(unittest.TestCase):
    def test_parse_spreadsheet_url_extracts_gid_and_range(self) -> None:
        self.assertEqual(
            parse_spreadsheet_url(
                "https://docs.google.com/spreadsheets/d/sheet-id_123/edit?gid=1#gid=22&range=A1:E4"
            ),
            {"spreadsheet_id": "sheet-id_123", "gid": "22", "range": "A1:E4"},
        )

    def test_qualify_ranges_uses_gid_sheet_title(self) -> None:
        metadata = {
            "sheets": [
                {"properties": {"sheetId": 22, "title": "선택 탭", "gridProperties": {}}}
            ]
        }

        self.assertEqual(
            qualify_ranges(metadata, ranges=["A1:E4"], gid="22"),
            ["'선택 탭'!A1:E4"],
        )

    def test_inspect_sheet_reads_values_with_user_oauth(self) -> None:
        calls = []

        def transport(url: str, token: str) -> dict:
            calls.append((url, token))
            if "values:batchGet" in url:
                return {
                    "spreadsheetId": "spreadsheet-1",
                    "valueRanges": [
                        {
                            "range": "'Input'!A1:B1",
                            "majorDimension": "ROWS",
                            "values": [["1", "2"]],
                        }
                    ],
                }
            return {
                "spreadsheetId": "spreadsheet-1",
                "properties": {"title": "Ops", "locale": "ko_KR", "timeZone": "Asia/Seoul"},
                "sheets": [
                    {
                        "properties": {
                            "sheetId": 10,
                            "title": "Input",
                            "gridProperties": {"rowCount": 100, "columnCount": 10},
                        }
                    }
                ],
            }

        snapshot = inspect_sheet(
            spreadsheet_id="spreadsheet-1",
            access_token="oauth-access-token",
            operation="inspect.values_window",
            ranges=["A1:B1"],
            gid="10",
            transport=transport,
        )

        self.assertEqual(snapshot["operation"], "inspect.values_window")
        self.assertEqual(snapshot["requested_ranges"], ["'Input'!A1:B1"])
        self.assertEqual(snapshot["windows"][0]["values"], [["1", "2"]])
        self.assertEqual(calls[0][1], "oauth-access-token")

    def test_write_package_creates_mcp_handoff(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            result = write_inspection_package(
                snapshot={
                    "snapshot_id": "snapshot-1",
                    "operation": "inspect.metadata",
                    "spreadsheet_id": "spreadsheet-1",
                    "title": "Ops",
                },
                package_root=Path(tmpdir),
                request_id="request-1",
                now=datetime(2026, 6, 5, tzinfo=UTC),
            )

            self.assertTrue(Path(result["manifest_path"]).exists())
            self.assertTrue(Path(result["mcp_handoff_path"]).exists())
            handoff = json.loads(Path(result["mcp_handoff_path"]).read_text(encoding="utf-8"))
            self.assertIn("이 Sheets Bridge 패키지를 분석해줘", handoff["mcp_prompt"])

    def test_write_package_marks_apply_result_primary_artifact(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            result = write_inspection_package(
                snapshot={
                    "snapshot_id": "snapshot-apply",
                    "operation": "apply.values_update",
                    "spreadsheet_id": "spreadsheet-1",
                },
                package_root=Path(tmpdir),
                request_id="request-apply",
                now=datetime(2026, 6, 5, tzinfo=UTC),
            )

            manifest = json.loads(Path(result["manifest_path"]).read_text(encoding="utf-8"))
            self.assertEqual(manifest["artifacts"][0]["kind"], "apply_result")

    def test_apply_values_update_captures_rollback_and_readback(self) -> None:
        value_call_count = 0
        write_calls = []

        def transport(url: str, token: str) -> dict:
            nonlocal value_call_count
            if "values:batchGet" in url:
                value_call_count += 1
                values = [["old", "=A2"]] if value_call_count == 1 else [["new", "=A1"]]
                return {
                    "spreadsheetId": "spreadsheet-1",
                    "valueRanges": [
                        {
                            "range": "'Input'!A1:B1",
                            "majorDimension": "ROWS",
                            "values": values,
                        }
                    ],
                }
            return {
                "spreadsheetId": "spreadsheet-1",
                "properties": {"title": "Ops"},
                "sheets": [
                    {
                        "properties": {
                            "sheetId": 10,
                            "title": "Input",
                            "gridProperties": {"rowCount": 100, "columnCount": 10},
                        }
                    }
                ],
            }

        def write_transport(url: str, token: str, body: dict) -> dict:
            write_calls.append((url, token, body))
            return {
                "spreadsheetId": "spreadsheet-1",
                "totalUpdatedRows": 1,
                "totalUpdatedColumns": 2,
                "totalUpdatedCells": 2,
            }

        snapshot = apply_values_update(
            spreadsheet_id="spreadsheet-1",
            access_token="oauth-access-token",
            write_requests=[{"range": "A1:B1", "values": [["new", "=A1"]]}],
            gid="10",
            transport=transport,
            write_transport=write_transport,
        )

        self.assertEqual(snapshot["operation"], "apply.values_update")
        self.assertEqual(snapshot["requested_ranges"], ["'Input'!A1:B1"])
        self.assertEqual(snapshot["updated_cells"], 2)
        self.assertEqual(write_calls[0][2]["data"][0]["values"], [["new", "=A1"]])
        self.assertEqual(
            snapshot["rollback"]["write_requests"],
            [{"range": "'Input'!A1:B1", "values": [["old", "=A2"]]}],
        )
        self.assertEqual(snapshot["after"][0]["values"], [["new", "=A1"]])

    def test_mcp_lists_and_calls_auth_status_tool(self) -> None:
        listed = handle_message({"jsonrpc": "2.0", "id": 1, "method": "tools/list"})
        self.assertEqual(listed["result"]["tools"][0]["name"], "sheets_bridge_auth_status")
        tool_names = {item["name"] for item in listed["result"]["tools"]}
        self.assertIn("sheets_bridge_visualize_table_io", tool_names)
        self.assertIn("sheets_bridge_refactor_minimal_formula_sheet", tool_names)
        self.assertIn("spreadsheet_table_builder_ui", tool_names)
        self.assertIn("spreadsheet_table_builder_save_intent", tool_names)
        self.assertIn("spreadsheet_create_formula_table_from_spec", tool_names)
        self.assertIn("spreadsheet_rollback_created_artifact", tool_names)
        self.assertIn("spreadsheet_validate_excel_formula_results", tool_names)
        table_builder_tool = next(item for item in listed["result"]["tools"] if item["name"] == "spreadsheet_table_builder_ui")
        self.assertEqual(table_builder_tool["_meta"]["ui"]["resourceUri"], "ui://sheets-bridge/table-builder")
        self.assertEqual(table_builder_tool["_meta"]["ui/resourceUri"], "ui://sheets-bridge/table-builder")

        called = handle_message(
            {
                "jsonrpc": "2.0",
                "id": 2,
                "method": "tools/call",
                "params": {"name": "sheets_bridge_auth_status", "arguments": {}},
            },
            tool_caller=lambda _name, _args: {"authenticated": False},
        )
        self.assertFalse(called["result"]["structuredContent"]["authenticated"])

    def test_mcp_exposes_table_builder_app_resource(self) -> None:
        initialized = handle_message({"jsonrpc": "2.0", "id": 1, "method": "initialize", "params": {}})
        self.assertIn("resources", initialized["result"]["capabilities"])

        listed = handle_message({"jsonrpc": "2.0", "id": 2, "method": "resources/list"})
        resources = listed["result"]["resources"]
        self.assertEqual(resources[0]["uri"], "ui://sheets-bridge/table-builder")
        self.assertEqual(resources[0]["mimeType"], "text/html;profile=mcp-app")

        read = handle_message(
            {
                "jsonrpc": "2.0",
                "id": 3,
                "method": "resources/read",
                "params": {"uri": "ui://sheets-bridge/table-builder"},
            }
        )
        content = read["result"]["contents"][0]
        app_html_path = Path(table_builder_module.__file__).resolve().parent / "ui" / "table_builder" / "mcp_app.html"
        host_adapter_path = Path(table_builder_module.__file__).resolve().parent / "ui" / "table_builder" / "host_adapter.js"
        self.assertTrue(app_html_path.exists())
        self.assertTrue(host_adapter_path.exists())
        self.assertEqual(content["mimeType"], "text/html;profile=mcp-app")
        self.assertIn(host_adapter_path.read_text(encoding="utf-8").strip(), content["text"])
        self.assertNotIn("SHEETS_BRIDGE_HOST_ADAPTER_BUNDLE", content["text"])
        self.assertIn("AI에게 이해한 내용 확인하기", content["text"])
        self.assertIn("ui/initialize", content["text"])
        self.assertIn('protocolVersion: "2026-01-26"', content["text"])
        self.assertIn("appInfo:", content["text"])
        self.assertIn("tools/call", content["text"])
        self.assertIn("sendMessage: function sendMessage", content["text"])
        self.assertIn("spreadsheet_table_builder_save_intent", content["text"])

    def test_stdio_entrypoint_smoke_lists_resources_and_builds_excel_preview(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "source.xlsx"
            package_root = Path(tmpdir) / "packages"
            _write_table_builder_workbook(workbook_path)
            messages = [
                {
                    "jsonrpc": "2.0",
                    "id": 1,
                    "method": "initialize",
                    "params": {
                        "protocolVersion": "2025-06-18",
                        "capabilities": {},
                        "clientInfo": {"name": "stdio-smoke", "version": "1.0"},
                    },
                },
                {"jsonrpc": "2.0", "method": "notifications/initialized", "params": {}},
                {"jsonrpc": "2.0", "id": 2, "method": "tools/list", "params": {}},
                {"jsonrpc": "2.0", "id": 3, "method": "resources/list", "params": {}},
                {
                    "jsonrpc": "2.0",
                    "id": 4,
                    "method": "resources/read",
                    "params": {"uri": "ui://sheets-bridge/table-builder"},
                },
                {
                    "jsonrpc": "2.0",
                    "id": 5,
                    "method": "tools/call",
                    "params": {
                        "name": "spreadsheet_table_builder_ui",
                        "arguments": {
                            "workbook_path": str(workbook_path),
                            "sheet_name": "Raw",
                            "source_range": "A1:C4",
                            "package_root": str(package_root),
                        },
                    },
                },
            ]
            completed = subprocess.run(
                [sys.executable, str(REPO_ROOT / "mcp" / "sheets_bridge_server.py")],
                input="\n".join(json.dumps(message, ensure_ascii=False) for message in messages) + "\n",
                cwd=REPO_ROOT,
                check=True,
                capture_output=True,
                text=True,
                timeout=30,
            )

            responses = [json.loads(line) for line in completed.stdout.splitlines() if line.strip()]
            self.assertEqual([response["id"] for response in responses], [1, 2, 3, 4, 5])
            tools = {item["name"] for item in responses[1]["result"]["tools"]}
            self.assertIn("spreadsheet_table_builder_ui", tools)
            self.assertEqual(
                responses[2]["result"]["resources"][0]["uri"],
                "ui://sheets-bridge/table-builder",
            )
            resource_text = responses[3]["result"]["contents"][0]["text"]
            self.assertIn("SheetsBridgeHostAdapters", resource_text)
            self.assertIn("새 표 만들기", resource_text)
            preview = responses[4]["result"]["structuredContent"]
            self.assertEqual(preview["artifact_type"], "excel_workbook")
            self.assertEqual(preview["summary"]["sheet_title"], "Raw")
            self.assertTrue(Path(preview["package"]["html_path"]).exists())
            self.assertTrue(Path(preview["package"]["source_path"]).exists())

    def test_table_builder_save_intent_tool_writes_intent(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            called = handle_message(
                {
                    "jsonrpc": "2.0",
                    "id": 1,
                    "method": "tools/call",
                    "params": {
                        "name": "spreadsheet_table_builder_save_intent",
                        "arguments": {
                            "package_root": tmpdir,
                            "intent": {
                                "artifact_type": "google_sheets",
                                "source": {"spreadsheet_id": "spreadsheet-1", "qualified_range": "'Raw'!A1:C4"},
                                "output_canvas": [["", "Jan"], ["Team A", ""]],
                                "llm_prompt": "팀별 월별 매출 합계를 수식으로 채워줘.",
                            },
                        },
                    },
                }
            )
            result = called["result"]["structuredContent"]
            intent_path = Path(result["package"]["intent_path"])
            self.assertTrue(intent_path.exists())
            saved = json.loads(intent_path.read_text(encoding="utf-8"))
            self.assertEqual(saved["intent_kind"], "table_build_intent_v1")
            self.assertEqual(saved["output_canvas"], [["", "Jan"], ["Team A", ""]])
            self.assertEqual(saved["output"]["creation_mode"], "sheet")
            self.assertIn("TableBuildPlan", result["next_prompt"])

    def test_table_builder_save_intent_requires_canvas_and_prompt(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            with self.assertRaisesRegex(ValueError, "intent.output_canvas is required"):
                table_builder_module.save_table_build_intent(
                    intent={
                        "artifact_type": "google_sheets",
                        "llm_prompt": "팀별 월별 매출 합계를 수식으로 채워줘.",
                    },
                    package_root=Path(tmpdir),
                )
            with self.assertRaisesRegex(ValueError, "intent.llm_prompt is required"):
                table_builder_module.save_table_build_intent(
                    intent={
                        "artifact_type": "google_sheets",
                        "output_canvas": [["", "Jan"], ["Team A", ""]],
                    },
                    package_root=Path(tmpdir),
                )

    def test_mcp_initialize_reports_single_source_version(self) -> None:
        initialized = handle_message({"jsonrpc": "2.0", "id": 1, "method": "initialize", "params": {}})
        self.assertEqual(initialized["result"]["serverInfo"]["version"], __version__)

    def test_mcp_can_configure_oauth(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            with mock.patch.dict("os.environ", {"SHEETS_BRIDGE_CONFIG_DIR": tmpdir}):
                called = handle_message(
                    {
                        "jsonrpc": "2.0",
                        "id": 3,
                        "method": "tools/call",
                        "params": {
                            "name": "sheets_bridge_configure_oauth",
                            "arguments": {"client_id": "client-id"},
                        },
                    }
                )

        self.assertTrue(called["result"]["structuredContent"]["configured"])

    def test_mcp_configure_oauth_can_request_readwrite_scope(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            with mock.patch.dict("os.environ", {"SHEETS_BRIDGE_CONFIG_DIR": tmpdir}):
                called = handle_message(
                    {
                        "jsonrpc": "2.0",
                        "id": 4,
                        "method": "tools/call",
                        "params": {
                            "name": "sheets_bridge_configure_oauth",
                            "arguments": {"client_id": "client-id", "access": "readwrite"},
                        },
                    }
                )
                config = json.loads(auth.oauth_config_path(Path(tmpdir)).read_text(encoding="utf-8"))

        self.assertIn(auth.READWRITE_SCOPE, called["result"]["structuredContent"]["scopes"])
        self.assertEqual(config["scopes"], [auth.READWRITE_SCOPE])

    def test_mcp_configure_oauth_can_request_copy_scope(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            with mock.patch.dict("os.environ", {"SHEETS_BRIDGE_CONFIG_DIR": tmpdir}):
                called = handle_message(
                    {
                        "jsonrpc": "2.0",
                        "id": 5,
                        "method": "tools/call",
                        "params": {
                            "name": "sheets_bridge_configure_oauth",
                            "arguments": {"client_id": "client-id", "access": "copy"},
                        },
                    }
                )
                config = json.loads(auth.oauth_config_path(Path(tmpdir)).read_text(encoding="utf-8"))

        self.assertIn(auth.READWRITE_SCOPE, called["result"]["structuredContent"]["scopes"])
        self.assertIn(auth.DRIVE_FILE_SCOPE, called["result"]["structuredContent"]["scopes"])
        self.assertEqual(config["scopes"], [auth.READWRITE_SCOPE, auth.DRIVE_FILE_SCOPE])

    def test_mcp_configure_oauth_can_request_copy_full_scope(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            with mock.patch.dict("os.environ", {"SHEETS_BRIDGE_CONFIG_DIR": tmpdir}):
                called = handle_message(
                    {
                        "jsonrpc": "2.0",
                        "id": 6,
                        "method": "tools/call",
                        "params": {
                            "name": "sheets_bridge_configure_oauth",
                            "arguments": {"client_id": "client-id", "access": "copy_full"},
                        },
                    }
                )
                config = json.loads(auth.oauth_config_path(Path(tmpdir)).read_text(encoding="utf-8"))

        self.assertIn(auth.READWRITE_SCOPE, called["result"]["structuredContent"]["scopes"])
        self.assertIn(auth.DRIVE_SCOPE, called["result"]["structuredContent"]["scopes"])
        self.assertEqual(config["scopes"], [auth.READWRITE_SCOPE, auth.DRIVE_SCOPE])

    def test_chrome_resolver_reads_name_box_range_from_cdp(self) -> None:
        class FakeWebSocket:
            def __init__(self, _url: str) -> None:
                pass

            def call(self, _message: dict) -> dict:
                return {
                    "result": {
                        "result": {
                            "value": json.dumps(
                                {
                                    "url": "https://docs.google.com/spreadsheets/d/sheet-1/edit#gid=10",
                                    "title": "Sheet",
                                    "range": "O796:W802",
                                }
                            )
                        }
                    }
                }

            def close(self) -> None:
                pass

        context = resolve_current_sheet(
            transport=lambda _url: [
                {
                    "url": "https://docs.google.com/spreadsheets/d/sheet-1/edit#gid=10",
                    "title": "Sheet",
                    "webSocketDebuggerUrl": "ws://127.0.0.1/devtools/page/1",
                }
            ],
            websocket_factory=FakeWebSocket,
        )

        self.assertEqual(context["spreadsheet_id"], "sheet-1")
        self.assertEqual(context["gid"], "10")
        self.assertEqual(context["range"], "O796:W802")

    def test_auth_refresh_merges_refresh_token_without_printing_token(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            config_dir = Path(tmpdir)
            auth.save_oauth_client_config(client_id="client-id", config_dir=config_dir)
            auth.save_token(
                {
                    "access_token": "old",
                    "refresh_token": "refresh",
                    "expires_at": (datetime.now(UTC) - timedelta(seconds=5)).isoformat(),
                },
                config_dir=config_dir,
            )

            token = auth.get_access_token(
                config_dir=config_dir,
                token_transport=lambda _url, _body: {
                    "access_token": "new",
                    "expires_in": 3600,
                    "scope": "https://www.googleapis.com/auth/spreadsheets.readonly",
                },
            )

            self.assertEqual(token, "new")
            stored = json.loads(auth.token_path(config_dir).read_text(encoding="utf-8"))
            self.assertEqual(stored["refresh_token"], "refresh")

    def test_auth_rejects_write_operation_without_write_scope(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            config_dir = Path(tmpdir)
            auth.save_oauth_client_config(client_id="client-id", config_dir=config_dir)
            auth.save_token(
                {
                    "access_token": "readonly",
                    "refresh_token": "refresh",
                    "expires_at": (datetime.now(UTC) + timedelta(seconds=3600)).isoformat(),
                    "scope": auth.READONLY_SCOPE,
                },
                config_dir=config_dir,
            )

            with self.assertRaisesRegex(auth.OAuthFlowError, "write scope"):
                auth.get_access_token(config_dir=config_dir, required_scope=auth.READWRITE_SCOPE)

    def test_auth_accepts_full_drive_scope_for_copy_operation(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            config_dir = Path(tmpdir)
            auth.save_oauth_client_config(client_id="client-id", config_dir=config_dir)
            auth.save_token(
                {
                    "access_token": "copy-full",
                    "refresh_token": "refresh",
                    "expires_at": (datetime.now(UTC) + timedelta(seconds=3600)).isoformat(),
                    "scope": f"{auth.READWRITE_SCOPE} {auth.DRIVE_SCOPE}",
                },
                config_dir=config_dir,
            )

            token = auth.get_access_token(config_dir=config_dir, required_scope=auth.DRIVE_COPY_SCOPES)

            self.assertEqual(token, "copy-full")

    def test_visualize_table_io_writes_html_svg_package(self) -> None:
        metadata = _monthly_metadata()

        def transport(url: str, _token: str) -> dict:
            if "values:batchGet" in url and "valueRenderOption=FORMULA" in url:
                return {
                    "spreadsheetId": "spreadsheet-1",
                    "valueRanges": [
                        {
                            "range": "'1. MLL_ 상품별 성과 현황-2025.08'!A1:AQ129",
                            "values": _monthly_formula_window(),
                        }
                    ],
                }
            if "values:batchGet" in url:
                return {
                    "spreadsheetId": "spreadsheet-1",
                    "valueRanges": [
                        {
                            "range": "'1. MLL_ 상품별 성과 현황-2025.08'!A1:AQ129",
                            "values": _monthly_value_window(),
                        }
                    ],
                }
            return metadata

        with tempfile.TemporaryDirectory() as tmpdir:
            result = visualize_table_io(
                spreadsheet_id="spreadsheet-1",
                access_token="token",
                gid="157460759",
                target_range="A1:AQ129",
                package_root=Path(tmpdir),
                now=datetime(2026, 6, 6, tzinfo=UTC),
                transport=transport,
            )

            self.assertEqual(result["detected_pattern"], "monthly_product_performance_v1")
            self.assertTrue(Path(result["package"]["html_path"]).exists())
            self.assertTrue(Path(result["package"]["svg_path"]).exists())

    def test_minimal_formula_refactor_dry_run_builds_plan_without_write(self) -> None:
        metadata = _monthly_metadata()

        def transport(_url: str, _token: str) -> dict:
            return metadata

        def write_transport(_url: str, _token: str, _body: dict) -> dict:
            raise AssertionError("dry_run must not call write transport")

        with tempfile.TemporaryDirectory() as tmpdir:
            result = refactor_minimal_formula_sheet(
                spreadsheet_id="spreadsheet-1",
                access_token="token",
                gid="157460759",
                dry_run=True,
                package_root=Path(tmpdir),
                now=datetime(2026, 6, 6, tzinfo=UTC),
                transport=transport,
                write_transport=write_transport,
            )

            self.assertTrue(result["dry_run"])
            self.assertEqual(result["formula_anchor_count"], 15)
            self.assertEqual(result["new_sheet_title"], "MCP_ALT_MLL_2025.08_min_formula")
            self.assertTrue(Path(result["package"]["result_path"]).exists())

    def test_minimal_formula_write_requests_reference_detected_sources(self) -> None:
        writes = build_minimal_formula_write_requests(
            {
                "source_output": {"title": "1. MLL_ 상품별 성과 현황-2025.08"},
                "source_sku": {"title": "2. MLL_SKU별 성과 현황-2025.08"},
                "source_ad": {"title": "3. MLL_광고비 현황-2025.08"},
                "new_sheet_title": "MCP_ALT_MLL_2025.08_min_formula",
                "product_start_row": 6,
                "product_end_row": 129,
                "sku_end_row": 1002,
                "ad_end_row": 242,
            }
        )
        formula_count = sum(
            1
            for item in writes
            for row in item["values"]
            for value in row
            if str(value).startswith("=")
        )
        formulas = "\n".join(str(value) for item in writes for row in item["values"] for value in row)

        self.assertEqual(formula_count, 15)
        self.assertIn("ARRAYFORMULA(MMULT", formulas)
        self.assertIn("'2. MLL_SKU별 성과 현황-2025.08'", formulas)
        self.assertIn("'3. MLL_광고비 현황-2025.08'", formulas)

    def test_compare_value_matrices_reports_mismatches_and_errors(self) -> None:
        comparison = compare_value_matrices(
            [["1", "2"]],
            [["1", "#REF!"]],
            dimensions=(1, 2),
        )

        self.assertEqual(comparison["checked_cells"], 2)
        self.assertEqual(comparison["mismatch_count"], 1)
        self.assertEqual(comparison["error_count"], 1)

    def test_table_builder_ui_writes_interactive_package(self) -> None:
        metadata = _table_builder_metadata()

        def transport(url: str, _token: str) -> dict:
            if "values:batchGet" in url:
                return {
                    "spreadsheetId": "spreadsheet-1",
                    "valueRanges": [
                        {
                            "range": "'Raw'!A1:C4",
                            "values": _table_builder_values(formulas="valueRenderOption=FORMULA" in url),
                        }
                    ],
                }
            return metadata

        with tempfile.TemporaryDirectory() as tmpdir:
            result = build_table_builder_ui(
                spreadsheet_id="spreadsheet-1",
                access_token="token",
                gid="10",
                source_range="A1:C4",
                package_root=Path(tmpdir),
                now=datetime(2026, 6, 6, tzinfo=UTC),
                transport=transport,
            )
            html = Path(result["package"]["html_path"]).read_text(encoding="utf-8")
            source = json.loads(Path(result["package"]["source_path"]).read_text(encoding="utf-8"))

            self.assertIn("새 표 만들기", html)
            self.assertIn("원하는 결과표 직접 입력", html)
            self.assertIn("id=\"output-canvas\"", html)
            self.assertIn("contenteditable=\"true\"", html)
            self.assertIn("AI에게 요청할 내용", html)
            self.assertIn("id=\"llm-prompt\"", html)
            self.assertIn("llm_prompt", html)
            self.assertNotIn("id=\"calculation-prompt\"", html)
            self.assertNotIn("calculation_prompt", html)
            self.assertIn("원본에서 열 찾아 넣기(선택)", html)
            self.assertIn("AI에게 이해한 내용 확인하기", html)
            self.assertIn("이 모양으로 만들까요?", html)
            self.assertIn("제가 이해한 요청입니다. 이 모양 맞나요?", html)
            self.assertIn("buildPreview", html)
            self.assertNotIn("data-preview-mode", html)
            self.assertEqual(source["default_spec"]["output"]["creation_mode"], "sheet")

    def test_create_formula_table_from_spec_dry_run_builds_formula_grid(self) -> None:
        metadata = _table_builder_metadata()
        spec = _table_builder_spec()

        def transport(url: str, _token: str) -> dict:
            if "values:batchGet" in url:
                return {
                    "spreadsheetId": "spreadsheet-1",
                    "valueRanges": [{"range": "'Raw'!A1:C4", "values": _table_builder_values()}],
                }
            return metadata

        def write_transport(_url: str, _token: str, _body: dict) -> dict:
            raise AssertionError("dry_run must not call write transport")

        with tempfile.TemporaryDirectory() as tmpdir:
            result = create_formula_table_from_spec(
                spec=spec,
                access_token="token",
                dry_run=True,
                package_root=Path(tmpdir),
                now=datetime(2026, 6, 6, tzinfo=UTC),
                transport=transport,
                write_transport=write_transport,
            )

            self.assertEqual(result["row_label_count"], 2)
            self.assertEqual(result["column_label_count"], 2)
            self.assertEqual(result["formula_cell_count"], 4)
            self.assertEqual(result["creation_mode"], "sheet")
            self.assertTrue(Path(result["package"]["result_path"]).exists())

    def test_create_formula_table_from_spec_uses_output_canvas_labels(self) -> None:
        metadata = _table_builder_metadata()
        spec = _table_builder_spec()
        spec["output_canvas"] = [
            ["", "Feb"],
            ["A", ""],
            ["C", ""],
        ]
        spec["llm_prompt"] = "이 표의 빈 칸을 원본 데이터 기반 수식으로 채워줘."

        def transport(url: str, _token: str) -> dict:
            if "values:batchGet" in url:
                return {
                    "spreadsheetId": "spreadsheet-1",
                    "valueRanges": [{"range": "'Raw'!A1:C4", "values": _table_builder_values()}],
                }
            return metadata

        def write_transport(_url: str, _token: str, _body: dict) -> dict:
            raise AssertionError("dry_run must not call write transport")

        with tempfile.TemporaryDirectory() as tmpdir:
            result = create_formula_table_from_spec(
                spec=spec,
                access_token="token",
                dry_run=True,
                package_root=Path(tmpdir),
                now=datetime(2026, 6, 6, tzinfo=UTC),
                transport=transport,
                write_transport=write_transport,
            )

        self.assertEqual(result["label_source"], "output_canvas")
        self.assertEqual(result["row_label_count"], 2)
        self.assertEqual(result["column_label_count"], 1)
        self.assertEqual(result["formula_cell_count"], 2)
        self.assertEqual(result["llm_prompt"], "이 표의 빈 칸을 원본 데이터 기반 수식으로 채워줘.")

    def test_formula_table_spec_rejects_out_of_range_column(self) -> None:
        spec = _table_builder_spec()
        spec["fields"]["measure"]["column"] = "Z"

        with self.assertRaisesRegex(ValueError, "outside the source range"):
            normalize_formula_table_spec(spec)

    def test_formula_table_spec_allows_arbitrary_sheet_formula_template(self) -> None:
        metadata = _table_builder_metadata()
        spec = _table_builder_spec()
        spec["formula"] = {
            "template": "=IFERROR(INDEX(FILTER({measure_range},{row_label_range}={row_label_cell},{column_label_range}={column_label_cell}),1),0)"
        }

        def transport(url: str, _token: str) -> dict:
            if "values:batchGet" in url:
                return {
                    "spreadsheetId": "spreadsheet-1",
                    "valueRanges": [{"range": "'Raw'!A1:C4", "values": _table_builder_values()}],
                }
            return metadata

        def write_transport(_url: str, _token: str, _body: dict) -> dict:
            raise AssertionError("dry_run must not call write transport")

        with tempfile.TemporaryDirectory() as tmpdir:
            result = create_formula_table_from_spec(
                spec=spec,
                access_token="token",
                dry_run=True,
                package_root=Path(tmpdir),
                now=datetime(2026, 6, 6, tzinfo=UTC),
                transport=transport,
                write_transport=write_transport,
            )

        self.assertIn("INDEX(FILTER", result["formula_template"])
        self.assertEqual(result["formula_cell_count"], 4)

    def test_formula_table_spec_rejects_non_formula_template(self) -> None:
        spec = _table_builder_spec()
        spec["formula"] = {"template": "SUM(A:A)"}

        with self.assertRaisesRegex(ValueError, "must start with '='"):
            normalize_formula_table_spec(spec)

    def test_table_builder_ui_supports_excel_workbook_source(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "source.xlsx"
            _write_table_builder_workbook(workbook_path)
            result = build_table_builder_ui(
                workbook_path=str(workbook_path),
                sheet_name="Raw",
                source_range="A1:C4",
                package_root=Path(tmpdir) / "packages",
                now=datetime(2026, 6, 6, tzinfo=UTC),
            )
            html = Path(result["package"]["html_path"]).read_text(encoding="utf-8")
            source = json.loads(Path(result["package"]["source_path"]).read_text(encoding="utf-8"))

            self.assertEqual(result["artifact_type"], "excel_workbook")
            self.assertIn("새 표 만들기", html)
            self.assertIn("원하는 결과표 직접 입력", html)
            self.assertIn("id=\"output-canvas\"", html)
            self.assertIn("원본을 복사해서 만들기", html)
            self.assertIn("AI에게 이해한 내용 확인하기", html)
            self.assertIn("이 모양으로 만들까요?", html)
            self.assertNotIn("미리보기 전환", html)
            self.assertIn("source.xlsx", source["workbook_title"])
            self.assertEqual(source["default_spec"]["output"]["creation_mode"], "copy")
            self.assertTrue(source["default_spec"]["output"]["workbook_path"].endswith("source-formula-table.xlsx"))

    def test_create_formula_table_from_spec_defaults_excel_to_workbook_copy(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "source.xlsx"
            output_path = Path(tmpdir) / "output.xlsx"
            _write_table_builder_workbook(workbook_path)
            spec = _table_builder_spec(
                artifact_type="excel_workbook",
                workbook_path=str(workbook_path),
                output_workbook_path=str(output_path),
            )

            result = create_formula_table_from_spec(
                spec=spec,
                access_token="",
                dry_run=False,
                package_root=Path(tmpdir) / "packages",
                now=datetime(2026, 6, 6, tzinfo=UTC),
            )

            wb = load_workbook(output_path, data_only=False)
            source_wb = load_workbook(workbook_path, read_only=True)
            try:
                self.assertEqual(result["creation_mode"], "copy")
                self.assertEqual(result["workbook_path"], str(output_path.resolve()))
                self.assertEqual(result["rollback"]["type"], "delete_output_workbook_copy")
                self.assertIn("MCP_TABLE_TEST", wb.sheetnames)
                self.assertTrue(str(wb["MCP_TABLE_TEST"]["B3"].value).startswith("="))
                self.assertNotIn("MCP_TABLE_TEST", source_wb.sheetnames)
            finally:
                wb.close()
                source_wb.close()

    def test_create_formula_table_from_spec_can_add_excel_worksheet_in_source_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "source.xlsx"
            _write_table_builder_workbook(workbook_path)
            spec = _table_builder_spec(
                artifact_type="excel_workbook",
                workbook_path=str(workbook_path),
                creation_mode="sheet",
            )

            result = create_formula_table_from_spec(
                spec=spec,
                access_token="",
                dry_run=False,
                package_root=Path(tmpdir) / "packages",
                now=datetime(2026, 6, 6, tzinfo=UTC),
            )

            wb = load_workbook(workbook_path, data_only=False)
            try:
                self.assertEqual(result["creation_mode"], "sheet")
                self.assertEqual(result["workbook_path"], str(workbook_path.resolve()))
                self.assertEqual(result["rollback"]["type"], "delete_created_worksheet")
                self.assertIn("MCP_TABLE_TEST", wb.sheetnames)
                self.assertTrue(str(wb["MCP_TABLE_TEST"]["B3"].value).startswith("="))
            finally:
                wb.close()

    def test_excel_formula_table_spec_rejects_output_workbook_path_for_sheet_mode(self) -> None:
        spec = _table_builder_spec(
            artifact_type="excel_workbook",
            workbook_path="/tmp/source.xlsx",
            creation_mode="sheet",
        )
        spec["output"]["workbook_path"] = "/tmp/output.xlsx"

        with self.assertRaisesRegex(ValueError, "only valid when output.creation_mode='copy'"):
            normalize_formula_table_spec(spec)

    def test_desktop_excel_formula_rewrite_map_records_formula_cells(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            map_path = Path(tmpdir) / "formulas.tsv"

            table_builder_module._write_excel_formula_rewrite_map(
                map_path,
                [
                    ["Team", "Jan", "Feb"],
                    ["A", "=SUM('Raw'!$C$2:$C$3)", "=SUM('Raw'!$D$2:$D$3)"],
                ],
            )

            lines = map_path.read_text(encoding="utf-8").splitlines()
            self.assertEqual(
                lines,
                [
                    "B2\t=SUM('Raw'!$C$2:$C$3)",
                    "C2\t=SUM('Raw'!$D$2:$D$3)",
                ],
            )

    def test_google_sheets_formula_table_copy_mode_dry_run_plans_drive_copy(self) -> None:
        metadata = _table_builder_metadata()
        spec = _table_builder_spec()
        spec["output"]["creation_mode"] = "copy"
        spec["output"]["copy_title"] = "Copied formula table"

        def transport(url: str, _token: str) -> dict:
            if "values:batchGet" in url:
                return {
                    "spreadsheetId": "spreadsheet-1",
                    "valueRanges": [{"range": "'Raw'!A1:C4", "values": _table_builder_values()}],
                }
            return metadata

        def write_transport(_url: str, _token: str, _body: dict) -> dict:
            raise AssertionError("dry_run must not call write transport")

        with tempfile.TemporaryDirectory() as tmpdir:
            result = create_formula_table_from_spec(
                spec=spec,
                access_token="token",
                dry_run=True,
                package_root=Path(tmpdir),
                now=datetime(2026, 6, 6, tzinfo=UTC),
                transport=transport,
                write_transport=write_transport,
            )

        self.assertEqual(result["creation_mode"], "copy")
        self.assertEqual(result["planned_copy"]["copy_title"], "Copied formula table")
        self.assertEqual(result["rollback"]["type"], "delete_copied_spreadsheet")

    def test_google_sheets_formula_table_copy_mode_writes_to_copied_spreadsheet(self) -> None:
        source_metadata = _table_builder_metadata()
        copied_metadata = {
            **_table_builder_metadata(),
            "spreadsheetId": "spreadsheet-copy",
            "properties": {"title": "Copied formula table", "locale": "ko_KR", "timeZone": "Asia/Seoul"},
        }
        spec = _table_builder_spec()
        spec["output"]["creation_mode"] = "copy"
        spec["output"]["copy_title"] = "Copied formula table"
        write_calls = []

        def transport(url: str, _token: str) -> dict:
            if "values:batchGet" in url:
                spreadsheet_id = "spreadsheet-copy" if "/spreadsheet-copy/" in url else "spreadsheet-1"
                return {
                    "spreadsheetId": spreadsheet_id,
                    "valueRanges": [{"range": "'Raw'!A1:C4", "values": _table_builder_values()}],
                }
            if "/spreadsheet-copy" in url:
                return copied_metadata
            return source_metadata

        def write_transport(url: str, _token: str, body: dict) -> dict:
            write_calls.append((url, body))
            if "drive/v3/files" in url and url.endswith("/copy?fields=id%2Cname%2CmimeType%2CwebViewLink"):
                return {"id": "spreadsheet-copy", "name": body["name"], "webViewLink": "https://docs.google.com/spreadsheets/d/spreadsheet-copy/edit"}
            if ":batchUpdate" in url:
                return {"replies": [{}]}
            if "values:batchUpdate" in url:
                return {"totalUpdatedCells": 9, "totalUpdatedRows": 3, "totalUpdatedColumns": 3}
            raise AssertionError(f"unexpected write url: {url}")

        with tempfile.TemporaryDirectory() as tmpdir:
            result = create_formula_table_from_spec(
                spec=spec,
                access_token="token",
                dry_run=False,
                package_root=Path(tmpdir),
                now=datetime(2026, 6, 6, tzinfo=UTC),
                transport=transport,
                write_transport=write_transport,
            )

        self.assertEqual(result["spreadsheet_id"], "spreadsheet-copy")
        self.assertEqual(result["source_spreadsheet_id"], "spreadsheet-1")
        self.assertEqual(result["rollback"]["type"], "delete_copied_spreadsheet")
        self.assertTrue(any("drive/v3/files/spreadsheet-1/copy" in call[0] for call in write_calls))
        self.assertTrue(any("spreadsheets/spreadsheet-copy:batchUpdate" in call[0] for call in write_calls))

    def test_rollback_created_sheet_uses_delete_sheet_request(self) -> None:
        write_calls = []

        def write_transport(url: str, token: str, body: dict) -> dict:
            write_calls.append((url, token, body))
            return {"replies": [{}]}

        result = rollback_created_artifact(
            rollback={"type": "delete_created_sheet", "spreadsheet_id": "spreadsheet-1", "sheet_id": 123},
            access_token="token",
            write_transport=write_transport,
        )

        self.assertEqual(result["status"], "rolled_back")
        self.assertEqual(write_calls[0][2], {"requests": [{"deleteSheet": {"sheetId": 123}}]})

    def test_rollback_copied_spreadsheet_uses_drive_delete(self) -> None:
        delete_calls = []

        def delete_transport(url: str, token: str) -> dict:
            delete_calls.append((url, token))
            return {}

        result = rollback_created_artifact(
            rollback={"type": "delete_copied_spreadsheet", "file_id": "spreadsheet-copy"},
            access_token="token",
            delete_transport=delete_transport,
        )

        self.assertEqual(result["status"], "rolled_back")
        self.assertIn("drive/v3/files/spreadsheet-copy", delete_calls[0][0])

    def test_rollback_excel_workbook_copy_deletes_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "copy.xlsx"
            _write_table_builder_workbook(workbook_path)

            result = rollback_created_artifact(
                rollback={"type": "delete_output_workbook_copy", "workbook_path": str(workbook_path)},
            )

            self.assertEqual(result["status"], "rolled_back")
            self.assertFalse(workbook_path.exists())

    def test_rollback_excel_created_worksheet_removes_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "source.xlsx"
            _write_table_builder_workbook(workbook_path)
            wb = load_workbook(workbook_path)
            try:
                wb.create_sheet("MCP_TABLE_TEST")
                wb.save(workbook_path)
            finally:
                wb.close()

            result = rollback_created_artifact(
                rollback={"type": "delete_created_worksheet", "workbook_path": str(workbook_path), "sheet_title": "MCP_TABLE_TEST"},
            )

            wb = load_workbook(workbook_path, read_only=True)
            try:
                self.assertEqual(result["status"], "rolled_back")
                self.assertNotIn("MCP_TABLE_TEST", wb.sheetnames)
            finally:
                wb.close()

    def test_validate_excel_formula_results_static_path(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "source.xlsx"
            _write_table_builder_workbook(workbook_path)

            result = validate_excel_formula_results(
                workbook_path=str(workbook_path),
                worksheet="Raw",
                cells=["A1"],
                run_excel_engine=False,
            )

            self.assertIn(result["status"], {"static_passed", "review_required"})
            self.assertEqual(result["excel_engine"]["status"], "not_run")

    def test_validate_excel_formula_results_handles_sparse_empty_cells(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "sparse.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Sparse"
            ws["A1"] = "Header"
            ws["Z20"] = "=SUM(1,2)"
            wb.save(workbook_path)
            wb.close()

            result = validate_excel_formula_results(
                workbook_path=str(workbook_path),
                worksheet="Sparse",
                cells=["Z20"],
                run_excel_engine=False,
            )

            self.assertIn(result["status"], {"static_passed", "review_required"})
            self.assertGreaterEqual(result["static_scan"]["summary"]["formula_count"], 1)

    def test_validate_excel_formula_results_can_sample_without_subprocess_runner(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "sample.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Raw"
            ws["A1"] = "=SUM(1,2)"
            wb.save(workbook_path)
            wb.close()

            with mock.patch.object(table_builder_module, "sample_workbook_cells", return_value={"A1": "3"}):
                result = validate_excel_formula_results(
                    workbook_path=str(workbook_path),
                    worksheet="Raw",
                    cells=["A1"],
                    run_excel_engine=True,
                )

            self.assertEqual(result["excel_engine"]["status"], "passed")
            self.assertEqual(result["excel_engine"]["values"], {"A1": "3"})


def _monthly_metadata() -> dict:
    return {
        "spreadsheetId": "spreadsheet-1",
        "properties": {
            "title": "Monthly performance workbook",
            "locale": "ko_KR",
            "timeZone": "Asia/Seoul",
        },
        "sheets": [
            {
                "properties": {
                    "sheetId": 157460759,
                    "title": "1. MLL_ 상품별 성과 현황-2025.08",
                    "index": 0,
                    "gridProperties": {"rowCount": 245, "columnCount": 43},
                }
            },
            {
                "properties": {
                    "sheetId": 200,
                    "title": "2. MLL_SKU별 성과 현황-2025.08",
                    "index": 1,
                    "gridProperties": {"rowCount": 1002, "columnCount": 38},
                }
            },
            {
                "properties": {
                    "sheetId": 300,
                    "title": "3. MLL_광고비 현황-2025.08",
                    "index": 2,
                    "gridProperties": {"rowCount": 242, "columnCount": 5},
                }
            },
        ],
    }


def _monthly_value_window() -> list[list[str]]:
    rows = [[""] * 43 for _ in range(129)]
    rows[0][0] = "MLL 상품별 성과"
    rows[4][0:13] = [
        "구분",
        "카테고리",
        "상품",
        "런칭일",
        "기획자",
        "마케터",
        "",
        "결제액 Total",
        "광고비",
        "",
        "",
        "CAC",
        "2025-08-01",
    ]
    rows[5][0:13] = ["신규", "부업", "상품 A", "2025-08-01", "PM", "MKT", "", "100", "10", "", "", "10%", "100"]
    return rows


def _monthly_formula_window() -> list[list[str]]:
    rows = [[""] * 43 for _ in range(129)]
    rows[4] = _monthly_value_window()[4]
    rows[5][7] = "=SUM(M6:AQ6)"
    rows[5][8] = "=SUMIFS('3. MLL_광고비 현황-2025.08'!$E:$E,'3. MLL_광고비 현황-2025.08'!$D:$D,$C6)"
    rows[5][11] = "=IFERROR(I6/H6,0)"
    rows[5][12] = "=SUMIF('2. MLL_SKU별 성과 현황-2025.08'!$C$4:$C,$C6,'2. MLL_SKU별 성과 현황-2025.08'!H$4:H)"
    return rows


def _table_builder_metadata() -> dict:
    return {
        "spreadsheetId": "spreadsheet-1",
        "properties": {"title": "Builder workbook", "locale": "ko_KR", "timeZone": "Asia/Seoul"},
        "sheets": [
            {
                "properties": {
                    "sheetId": 10,
                    "title": "Raw",
                    "index": 0,
                    "gridProperties": {"rowCount": 20, "columnCount": 8},
                }
            }
        ],
    }


def _table_builder_values(*, formulas: bool = False) -> list[list[str]]:
    if formulas:
        return [
            ["Team", "Month", "Revenue"],
            ["A", "Jan", "=100"],
            ["A", "Feb", "=120"],
            ["B", "Jan", "=80"],
        ]
    return [
        ["Team", "Month", "Revenue"],
        ["A", "Jan", "100"],
        ["A", "Feb", "120"],
        ["B", "Jan", "80"],
    ]


def _table_builder_spec(
    *,
    artifact_type: str = "google_sheets",
    workbook_path: str = "",
    creation_mode: str = "",
    output_workbook_path: str = "",
) -> dict:
    source = {
        "artifact_type": artifact_type,
        "spreadsheet_id": "spreadsheet-1" if artifact_type == "google_sheets" else "",
        "workbook_path": workbook_path,
        "sheet_title": "Raw",
        "qualified_range": "'Raw'!A1:C4",
        "header_row": 1,
    }
    output = {"sheet_title": "MCP_TABLE_TEST", "title": "Team by Month"}
    if creation_mode:
        output["creation_mode"] = creation_mode
    if output_workbook_path:
        output["workbook_path"] = output_workbook_path
    return {
        "schema_version": "1.0",
        "spec_kind": "formula_table_apply_v1",
        "artifact_type": artifact_type,
        "spreadsheet_id": "spreadsheet-1" if artifact_type == "google_sheets" else "",
        "source": source,
        "fields": {
            "row_label": {"column": "A", "header": "Team", "selected_cell": "A1"},
            "column_label": {"column": "B", "header": "Month", "selected_cell": "B1"},
            "measure": {"column": "C", "header": "Revenue", "selected_cell": "C1"},
        },
        "formula": {
            "template": "=IFERROR(SUMIFS({measure_range},{row_label_range},{row_label_cell},{column_label_range},{column_label_cell}),0)"
        },
        "output": output,
    }


def _write_table_builder_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw"
    rows = _table_builder_values()
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


if __name__ == "__main__":
    unittest.main()
