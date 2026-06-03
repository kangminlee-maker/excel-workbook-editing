from __future__ import annotations

import argparse
import json
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from zipfile import ZipFile

from openpyxl.utils import get_column_letter, range_boundaries

SCHEMA_VERSION = "0.1"
MAIN_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"


def build_view_state_profile(
    manifest_path: Path,
    capture_quality_paths: list[Path] | None = None,
    *,
    max_sheet_xml_bytes: int = 100_000_000,
) -> dict[str, Any]:
    manifest_path = manifest_path.expanduser().resolve()
    quality_paths = [
        path.expanduser().resolve()
        for path in (capture_quality_paths or [])
    ]
    manifest = _read_json(manifest_path)
    workbook_path = Path(manifest["source"]["path"]).expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"missing workbook: {workbook_path}")
    quality_packages = [
        (path, _read_json(path))
        for path in quality_paths
    ]
    quality_results = _quality_results(quality_packages)
    selected_sheet_names = _selected_sheet_names(manifest, quality_results)
    manifest_sheets = {sheet["name"]: sheet for sheet in manifest["workbook"]["sheets"]}

    with ZipFile(workbook_path) as zf:
        sheet_profiles = [
            _sheet_profile(
                zf,
                manifest_sheets[sheet_name],
                max_sheet_xml_bytes=max_sheet_xml_bytes,
            )
            for sheet_name in selected_sheet_names
        ]

    sheet_by_name = {sheet["name"]: sheet for sheet in sheet_profiles}
    analyses = [
        _capture_window_analysis(result, sheet_by_name.get(result.get("sheet")))
        for result in quality_results
        if result.get("sheet") in sheet_by_name
    ]
    gate_results = [_gate_result(analysis) for analysis in analyses]

    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at": _utc_now(),
        "source_artifacts": {
            "manifest": str(manifest_path),
            "capture_quality_files": [str(path) for path in quality_paths],
        },
        "source": manifest["source"],
        "limits": {"max_sheet_xml_bytes": max_sheet_xml_bytes},
        "method": {
            "name": "deterministic_workbook_view_state_profile",
            "authority": "workbook_xml_view_state_evidence_not_semantic_truth",
            "decision_boundary": (
                "Preserve the original visible state; treat hidden/revealed variants as "
                "diagnostic projections, not replacements for the source workbook view."
            ),
        },
        "sheets": sheet_profiles,
        "capture_window_analyses": analyses,
        "gate_results": gate_results,
        "summary": _summary(sheet_profiles, analyses, gate_results),
        "parser_observations": _parser_observations(sheet_profiles, analyses),
    }


def _quality_results(quality_packages: list[tuple[Path, dict[str, Any]]]) -> list[dict[str, Any]]:
    results = []
    for path, package in quality_packages:
        for result in package.get("quality_results", []):
            results.append(
                {
                    **result,
                    "source_capture_quality_file": str(path),
                    "source_capture_quality_name": path.name,
                }
            )
    return results


def _selected_sheet_names(
    manifest: dict[str, Any],
    quality_results: list[dict[str, Any]],
) -> list[str]:
    manifest_order = [sheet["name"] for sheet in manifest["workbook"]["sheets"]]
    result_sheets = {
        result.get("sheet")
        for result in quality_results
        if result.get("sheet")
    }
    if not result_sheets:
        return manifest_order
    return [sheet_name for sheet_name in manifest_order if sheet_name in result_sheets]


def _sheet_profile(
    zf: ZipFile,
    manifest_sheet: dict[str, Any],
    *,
    max_sheet_xml_bytes: int,
) -> dict[str, Any]:
    entry = manifest_sheet.get("entry")
    base = {
        "name": manifest_sheet["name"],
        "entry": entry,
        "sheet_state": manifest_sheet.get("state", "visible"),
        "dimension": manifest_sheet.get("dimension"),
        "dimension_bounds": manifest_sheet.get("dimension_bounds"),
    }
    if not entry or entry not in zf.namelist():
        return {
            **base,
            "detail_status": "missing_entry",
            "entry_size_bytes": 0,
            "sheet_pr": {},
            "outline_pr": {},
            "sheet_format_pr": {},
            "sheet_views": [],
            "panes": [],
            "selections": [],
            "auto_filters": [],
            "sort_states": [],
            "hidden_row_spans": [],
            "zero_height_row_spans": [],
            "outline_row_spans": [],
            "collapsed_row_spans": [],
            "hidden_column_spans": [],
            "zero_width_column_spans": [],
            "outline_column_spans": [],
            "collapsed_column_spans": [],
            "summary": _sheet_summary({}, {}, [], []),
            "parser_observations": [
                {"level": "warning", "message": "Worksheet XML entry is missing."}
            ],
        }
    info = zf.getinfo(entry)
    if info.file_size > max_sheet_xml_bytes:
        return {
            **base,
            "detail_status": "skipped_large_xml",
            "entry_size_bytes": info.file_size,
            "sheet_pr": {},
            "outline_pr": {},
            "sheet_format_pr": {},
            "sheet_views": [],
            "panes": [],
            "selections": [],
            "auto_filters": [],
            "sort_states": [],
            "hidden_row_spans": [],
            "zero_height_row_spans": [],
            "outline_row_spans": [],
            "collapsed_row_spans": [],
            "hidden_column_spans": [],
            "zero_width_column_spans": [],
            "outline_column_spans": [],
            "collapsed_column_spans": [],
            "summary": _sheet_summary({}, {}, [], []),
            "parser_observations": [
                {
                    "level": "warning",
                    "message": "Worksheet XML was too large for view-state profiling.",
                }
            ],
        }
    scanned = _scan_sheet(zf, entry)
    public_scanned = {
        key: value
        for key, value in scanned.items()
        if key not in {"_row_states", "_column_states"}
    }
    return {
        **base,
        "detail_status": "scanned",
        "entry_size_bytes": info.file_size,
        **public_scanned,
        "summary": _sheet_summary(
            scanned["_row_states"],
            scanned["_column_states"],
            scanned["auto_filters"],
            scanned["panes"],
        ),
        "parser_observations": _sheet_observations(scanned),
    }


def _scan_sheet(zf: ZipFile, entry: str) -> dict[str, Any]:
    sheet_pr: dict[str, Any] = {}
    outline_pr: dict[str, Any] = {}
    sheet_format_pr: dict[str, Any] = {}
    sheet_views: list[dict[str, Any]] = []
    panes: list[dict[str, Any]] = []
    selections: list[dict[str, Any]] = []
    auto_filters: list[dict[str, Any]] = []
    sort_states: list[dict[str, Any]] = []
    row_states: dict[int, dict[str, Any]] = {}
    column_states: list[dict[str, Any]] = []

    with zf.open(entry) as handle:
        for event, elem in ET.iterparse(handle, events=("start", "end")):
            if event == "start":
                tag = _local_name(elem.tag)
                if tag == "sheetPr":
                    sheet_pr = dict(elem.attrib)
                elif tag == "outlinePr":
                    outline_pr = dict(elem.attrib)
                elif tag == "sheetFormatPr":
                    sheet_format_pr = dict(elem.attrib)
                elif tag == "sheetView":
                    sheet_views.append(dict(elem.attrib))
                elif tag == "pane":
                    panes.append(dict(elem.attrib))
                elif tag == "selection":
                    selections.append(dict(elem.attrib))
                elif tag == "autoFilter":
                    auto_filter = dict(elem.attrib)
                    auto_filter["bounds"] = _bounds_from_range(auto_filter.get("ref"))
                    auto_filters.append(auto_filter)
                elif tag == "sortState":
                    sort_state = dict(elem.attrib)
                    sort_state["bounds"] = _bounds_from_range(sort_state.get("ref"))
                    sort_states.append(sort_state)
                elif tag == "row":
                    state = _row_state(elem)
                    if _is_material_row_state(state):
                        row_states[state["row"]] = state
                elif tag == "col":
                    state = _column_state(elem)
                    if _is_material_column_state(state):
                        column_states.append(state)
            elif event == "end":
                elem.clear()

    return {
        "sheet_pr": sheet_pr,
        "outline_pr": outline_pr,
        "sheet_format_pr": sheet_format_pr,
        "sheet_views": sheet_views,
        "panes": panes,
        "selections": selections,
        "auto_filters": auto_filters,
        "sort_states": sort_states,
        "hidden_row_spans": _row_spans(row_states, lambda item: item["hidden"]),
        "zero_height_row_spans": _row_spans(row_states, lambda item: item["zero_height"]),
        "outline_row_spans": _row_spans(row_states, lambda item: item["outline_level"] > 0),
        "collapsed_row_spans": _row_spans(row_states, lambda item: item["collapsed"]),
        "hidden_column_spans": _column_spans(column_states, lambda item: item["hidden"]),
        "zero_width_column_spans": _column_spans(column_states, lambda item: item["zero_width"]),
        "outline_column_spans": _column_spans(
            column_states,
            lambda item: item["outline_level"] > 0,
        ),
        "collapsed_column_spans": _column_spans(column_states, lambda item: item["collapsed"]),
        "_row_states": row_states,
        "_column_states": column_states,
    }


def _row_state(elem: ET.Element) -> dict[str, Any]:
    height = _float_or_none(elem.attrib.get("ht"))
    return {
        "row": int(elem.attrib.get("r", "0")),
        "height": height,
        "hidden": elem.attrib.get("hidden") == "1",
        "zero_height": height is not None and height <= 1,
        "custom_height": elem.attrib.get("customHeight") == "1",
        "outline_level": _int_or_zero(elem.attrib.get("outlineLevel")),
        "collapsed": elem.attrib.get("collapsed") == "1",
    }


def _column_state(elem: ET.Element) -> dict[str, Any]:
    width = _float_or_none(elem.attrib.get("width"))
    min_column = int(elem.attrib.get("min", "0"))
    max_column = int(elem.attrib.get("max", str(min_column)))
    return {
        "min_column": min_column,
        "max_column": max_column,
        "min_column_letter": get_column_letter(min_column),
        "max_column_letter": get_column_letter(max_column),
        "width": width,
        "hidden": elem.attrib.get("hidden") == "1",
        "zero_width": width is not None and width <= 0.1,
        "custom_width": elem.attrib.get("customWidth") == "1",
        "outline_level": _int_or_zero(elem.attrib.get("outlineLevel")),
        "collapsed": elem.attrib.get("collapsed") == "1",
    }


def _is_material_row_state(state: dict[str, Any]) -> bool:
    return (
        state["hidden"]
        or state["zero_height"]
        or state["outline_level"] > 0
        or state["collapsed"]
    )


def _is_material_column_state(state: dict[str, Any]) -> bool:
    return (
        state["hidden"]
        or state["zero_width"]
        or state["outline_level"] > 0
        or state["collapsed"]
    )


def _row_spans(
    row_states: dict[int, dict[str, Any]],
    predicate: Any,
) -> list[dict[str, Any]]:
    rows = [row for row, state in row_states.items() if predicate(state)]
    rows.sort()
    spans = []
    start = None
    previous = None
    for row in rows:
        if start is None:
            start = row
            previous = row
            continue
        if previous is not None and row == previous + 1:
            previous = row
            continue
        spans.append(_row_span(start, previous, row_states))
        start = row
        previous = row
    if start is not None and previous is not None:
        spans.append(_row_span(start, previous, row_states))
    return spans


def _row_span(
    start_row: int,
    end_row: int,
    row_states: dict[int, dict[str, Any]],
) -> dict[str, Any]:
    states = [row_states[row] for row in range(start_row, end_row + 1) if row in row_states]
    return {
        "start_row": start_row,
        "end_row": end_row,
        "row_count": end_row - start_row + 1,
        "hidden_count": sum(1 for state in states if state["hidden"]),
        "zero_height_count": sum(1 for state in states if state["zero_height"]),
        "collapsed_count": sum(1 for state in states if state["collapsed"]),
        "max_outline_level": max([state["outline_level"] for state in states] or [0]),
        "min_height": min(
            [state["height"] for state in states if state["height"] is not None],
            default=None,
        ),
        "max_height": max(
            [state["height"] for state in states if state["height"] is not None],
            default=None,
        ),
    }


def _column_spans(
    column_states: list[dict[str, Any]],
    predicate: Any,
) -> list[dict[str, Any]]:
    spans = []
    for state in column_states:
        if not predicate(state):
            continue
        spans.append(
            {
                "start_column": state["min_column"],
                "end_column": state["max_column"],
                "start_column_letter": state["min_column_letter"],
                "end_column_letter": state["max_column_letter"],
                "column_count": state["max_column"] - state["min_column"] + 1,
                "hidden": state["hidden"],
                "zero_width": state["zero_width"],
                "collapsed": state["collapsed"],
                "outline_level": state["outline_level"],
                "width": state["width"],
            }
        )
    return spans


def _sheet_summary(
    row_states: dict[int, dict[str, Any]],
    column_states: list[dict[str, Any]],
    auto_filters: list[dict[str, Any]],
    panes: list[dict[str, Any]],
) -> dict[str, int]:
    return {
        "hidden_row_count": sum(1 for state in row_states.values() if state["hidden"]),
        "zero_height_row_count": sum(
            1 for state in row_states.values() if state["zero_height"]
        ),
        "outline_row_count": sum(
            1 for state in row_states.values() if state["outline_level"] > 0
        ),
        "collapsed_row_count": sum(1 for state in row_states.values() if state["collapsed"]),
        "hidden_column_count": sum(
            state["max_column"] - state["min_column"] + 1
            for state in column_states
            if state["hidden"]
        ),
        "outline_column_count": sum(
            state["max_column"] - state["min_column"] + 1
            for state in column_states
            if state["outline_level"] > 0
        ),
        "auto_filter_count": len(auto_filters),
        "frozen_pane_count": sum(1 for pane in panes if pane.get("state") == "frozen"),
    }


def _sheet_observations(scanned: dict[str, Any]) -> list[dict[str, str]]:
    observations = []
    if scanned["sheet_pr"].get("filterMode") == "1" and scanned["auto_filters"]:
        observations.append(
            {
                "level": "info",
                "message": "Sheet has filterMode=1 and autoFilter metadata; hidden rows may represent current filtered view state.",
            }
        )
    if scanned["hidden_row_spans"]:
        observations.append(
            {
                "level": "info",
                "message": "Sheet contains hidden row spans that affect visible Excel render captures.",
            }
        )
    if scanned["hidden_column_spans"]:
        observations.append(
            {
                "level": "info",
                "message": "Sheet contains hidden column spans that affect visible Excel render captures.",
            }
        )
    return observations


def _capture_window_analysis(
    quality_result: dict[str, Any],
    sheet_profile: dict[str, Any] | None,
) -> dict[str, Any]:
    range_text = quality_result.get("capture_window_range") or quality_result.get("requested_range")
    bounds = _bounds_from_range(range_text)
    if sheet_profile is None or bounds is None:
        row_summary = _empty_row_summary()
        column_summary = _empty_column_summary()
        filter_overlap = []
        classification = "range_or_sheet_unavailable"
        authority_decision = "review_required"
        next_action = "inspect_source_artifacts"
    else:
        row_summary = _row_state_summary(bounds, sheet_profile)
        column_summary = _column_state_summary(bounds, sheet_profile)
        filter_overlap = _filter_overlaps(bounds, sheet_profile.get("auto_filters", []))
        classification = _classification(quality_result, row_summary, column_summary, filter_overlap)
        authority_decision = _authority_decision(classification)
        next_action = _next_action(classification, quality_result)
    return {
        "id": f"view_state_{quality_result.get('id')}",
        "type": "view_state_capture_window_analysis",
        "capture_quality_file": quality_result.get("source_capture_quality_name"),
        "quality_result_id": quality_result.get("id"),
        "capture_id": quality_result.get("capture_id"),
        "target_id": quality_result.get("target_id"),
        "quality_status": quality_result.get("status"),
        "sheet": quality_result.get("sheet"),
        "range": range_text,
        "bounds": bounds,
        "row_state_summary": row_summary,
        "column_state_summary": column_summary,
        "filter_overlap": filter_overlap,
        "sheet_view_signals": _sheet_view_signals(sheet_profile),
        "classification": classification,
        "authority_decision": authority_decision,
        "recommended_next_action": next_action,
        "evidence_refs": [
            item
            for item in [
                quality_result.get("id"),
                quality_result.get("capture_id"),
                quality_result.get("source_capture_quality_name"),
            ]
            if item
        ],
    }


def _row_state_summary(
    bounds: dict[str, int],
    sheet_profile: dict[str, Any],
) -> dict[str, Any]:
    hidden_spans = _intersect_row_spans(bounds, sheet_profile.get("hidden_row_spans", []))
    zero_height_spans = _intersect_row_spans(bounds, sheet_profile.get("zero_height_row_spans", []))
    outline_spans = _intersect_row_spans(bounds, sheet_profile.get("outline_row_spans", []))
    collapsed_spans = _intersect_row_spans(bounds, sheet_profile.get("collapsed_row_spans", []))
    row_count = bounds["max_row"] - bounds["min_row"] + 1
    hidden_count = sum(span["row_count"] for span in hidden_spans)
    zero_height_count = sum(span["row_count"] for span in zero_height_spans)
    outline_count = sum(span["row_count"] for span in outline_spans)
    collapsed_count = sum(span["row_count"] for span in collapsed_spans)
    visible_count = max(0, row_count - hidden_count - zero_height_count)
    return {
        "row_count": row_count,
        "hidden_row_count": hidden_count,
        "zero_height_row_count": zero_height_count,
        "outline_row_count": outline_count,
        "collapsed_row_count": collapsed_count,
        "visible_row_count": visible_count,
        "hidden_row_ratio": _ratio(hidden_count + zero_height_count, row_count),
        "hidden_spans": hidden_spans,
        "zero_height_spans": zero_height_spans,
        "outline_spans": outline_spans,
        "collapsed_spans": collapsed_spans,
    }


def _column_state_summary(
    bounds: dict[str, int],
    sheet_profile: dict[str, Any],
) -> dict[str, Any]:
    hidden_spans = _intersect_column_spans(bounds, sheet_profile.get("hidden_column_spans", []))
    zero_width_spans = _intersect_column_spans(bounds, sheet_profile.get("zero_width_column_spans", []))
    outline_spans = _intersect_column_spans(bounds, sheet_profile.get("outline_column_spans", []))
    collapsed_spans = _intersect_column_spans(bounds, sheet_profile.get("collapsed_column_spans", []))
    column_count = bounds["max_column"] - bounds["min_column"] + 1
    hidden_count = sum(span["column_count"] for span in hidden_spans)
    zero_width_count = sum(span["column_count"] for span in zero_width_spans)
    outline_count = sum(span["column_count"] for span in outline_spans)
    collapsed_count = sum(span["column_count"] for span in collapsed_spans)
    visible_count = max(0, column_count - hidden_count - zero_width_count)
    return {
        "column_count": column_count,
        "hidden_column_count": hidden_count,
        "zero_width_column_count": zero_width_count,
        "outline_column_count": outline_count,
        "collapsed_column_count": collapsed_count,
        "visible_column_count": visible_count,
        "hidden_column_ratio": _ratio(hidden_count + zero_width_count, column_count),
        "hidden_spans": hidden_spans,
        "zero_width_spans": zero_width_spans,
        "outline_spans": outline_spans,
        "collapsed_spans": collapsed_spans,
    }


def _intersect_row_spans(
    bounds: dict[str, int],
    spans: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    intersections = []
    for span in spans:
        start = max(bounds["min_row"], span["start_row"])
        end = min(bounds["max_row"], span["end_row"])
        if start > end:
            continue
        intersections.append(
            {
                **span,
                "start_row": start,
                "end_row": end,
                "row_count": end - start + 1,
            }
        )
    return intersections


def _intersect_column_spans(
    bounds: dict[str, int],
    spans: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    intersections = []
    for span in spans:
        start = max(bounds["min_column"], span["start_column"])
        end = min(bounds["max_column"], span["end_column"])
        if start > end:
            continue
        intersections.append(
            {
                **span,
                "start_column": start,
                "end_column": end,
                "start_column_letter": get_column_letter(start),
                "end_column_letter": get_column_letter(end),
                "column_count": end - start + 1,
            }
        )
    return intersections


def _filter_overlaps(
    bounds: dict[str, int],
    auto_filters: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    overlaps = []
    for auto_filter in auto_filters:
        filter_bounds = auto_filter.get("bounds")
        if not filter_bounds:
            continue
        intersection = _bounds_intersection(bounds, filter_bounds)
        if intersection:
            overlaps.append(
                {
                    "ref": auto_filter.get("ref"),
                    "bounds": filter_bounds,
                    "intersection": intersection,
                }
            )
    return overlaps


def _sheet_view_signals(sheet_profile: dict[str, Any] | None) -> dict[str, Any]:
    if sheet_profile is None:
        return {
            "filter_mode": False,
            "auto_filter_count": 0,
            "frozen_pane_count": 0,
            "sheet_view_top_left_cells": [],
        }
    return {
        "filter_mode": sheet_profile.get("sheet_pr", {}).get("filterMode") == "1",
        "auto_filter_count": len(sheet_profile.get("auto_filters", [])),
        "frozen_pane_count": sheet_profile.get("summary", {}).get("frozen_pane_count", 0),
        "sheet_view_top_left_cells": [
            item.get("topLeftCell")
            for item in sheet_profile.get("sheet_views", [])
            if item.get("topLeftCell")
        ] + [
            item.get("topLeftCell")
            for item in sheet_profile.get("panes", [])
            if item.get("topLeftCell")
        ],
    }


def _classification(
    quality_result: dict[str, Any],
    row_summary: dict[str, Any],
    column_summary: dict[str, Any],
    filter_overlap: list[dict[str, Any]],
) -> str:
    quality_status = quality_result.get("status")
    row_count = row_summary["row_count"] or 0
    visible_rows = row_summary["visible_row_count"] or 0
    hidden_rows = row_summary["hidden_row_count"] + row_summary["zero_height_row_count"]
    hidden_columns = (
        column_summary["hidden_column_count"]
        + column_summary["zero_width_column_count"]
    )
    if row_count and hidden_rows >= row_count:
        return "all_rows_hidden_or_zero_height"
    if row_count and hidden_rows and visible_rows <= max(4, row_count * 0.2):
        if filter_overlap:
            return "filtered_or_hidden_rows_explain_capture_failure"
        return "hidden_rows_dominate_capture_window"
    if hidden_rows and filter_overlap:
        return "filtered_rows_affect_capture_window"
    if hidden_rows:
        return "mixed_visible_hidden_rows"
    if hidden_columns:
        return "hidden_columns_affect_capture_window"
    if quality_status in {"recapture_required", "capture_failed"}:
        return "capture_issue_without_view_state_explanation"
    return "no_material_view_state_signal"


def _authority_decision(classification: str) -> str:
    if classification in {
        "all_rows_hidden_or_zero_height",
        "filtered_or_hidden_rows_explain_capture_failure",
        "hidden_rows_dominate_capture_window",
        "filtered_rows_affect_capture_window",
        "mixed_visible_hidden_rows",
    }:
        return "separate_visible_render_authority_from_structural_data_authority"
    if classification == "hidden_columns_affect_capture_window":
        return "preserve_visible_column_state_and_use_structural_columns_for_dataflow"
    if classification == "capture_issue_without_view_state_explanation":
        return "recapture_or_capture_method_review_required"
    return "continue_with_visible_render_authority"


def _next_action(classification: str, quality_result: dict[str, Any]) -> str:
    if classification in {
        "all_rows_hidden_or_zero_height",
        "filtered_or_hidden_rows_explain_capture_failure",
        "hidden_rows_dominate_capture_window",
    }:
        return "add_non_authoritative_unhide_or_clear_filter_diagnostic_before_coordinate_normalization"
    if classification in {"filtered_rows_affect_capture_window", "mixed_visible_hidden_rows"}:
        return "record_visible_state_and_keep_hidden_rows_available_for_structural_extraction"
    if classification == "hidden_columns_affect_capture_window":
        return "record_visible_column_state_and_avoid_treating_hidden_columns_as_absent_data"
    if quality_result.get("status") in {"recapture_required", "capture_failed"}:
        return "review_capture_method_or_expand_target_selection"
    return "eligible_for_coordinate_normalization"


def _gate_result(analysis: dict[str, Any]) -> dict[str, Any]:
    classification = analysis["classification"]
    if classification in {
        "all_rows_hidden_or_zero_height",
        "filtered_or_hidden_rows_explain_capture_failure",
        "hidden_rows_dominate_capture_window",
    }:
        status = "explained_by_view_state"
    elif classification in {
        "filtered_rows_affect_capture_window",
        "mixed_visible_hidden_rows",
        "hidden_columns_affect_capture_window",
    }:
        status = "view_state_affects_window"
    elif classification == "capture_issue_without_view_state_explanation":
        status = "unexplained_by_view_state"
    else:
        status = "no_material_view_state"
    return {
        "id": f"gate_{analysis['id']}",
        "type": "view_state_gate_result",
        "analysis_id": analysis["id"],
        "target_id": analysis.get("target_id"),
        "gate_type": "capture_view_state_explanation",
        "status": status,
        "classification": classification,
        "authority_decision": analysis["authority_decision"],
        "recommended_next_action": analysis["recommended_next_action"],
        "evidence_refs": analysis["evidence_refs"],
    }


def _summary(
    sheets: list[dict[str, Any]],
    analyses: list[dict[str, Any]],
    gate_results: list[dict[str, Any]],
) -> dict[str, int]:
    return {
        "sheet_count": len(sheets),
        "scanned_sheet_count": sum(1 for sheet in sheets if sheet["detail_status"] == "scanned"),
        "sheet_with_filter_mode_count": sum(
            1 for sheet in sheets if sheet.get("sheet_pr", {}).get("filterMode") == "1"
        ),
        "sheet_with_hidden_rows_count": sum(
            1 for sheet in sheets if sheet["summary"]["hidden_row_count"] > 0
        ),
        "hidden_row_count": sum(sheet["summary"]["hidden_row_count"] for sheet in sheets),
        "hidden_column_count": sum(
            sheet["summary"]["hidden_column_count"] for sheet in sheets
        ),
        "capture_window_analysis_count": len(analyses),
        "view_state_explained_count": sum(
            1 for gate in gate_results if gate["status"] == "explained_by_view_state"
        ),
        "view_state_affects_count": sum(
            1 for gate in gate_results if gate["status"] == "view_state_affects_window"
        ),
        "unexplained_by_view_state_count": sum(
            1 for gate in gate_results if gate["status"] == "unexplained_by_view_state"
        ),
        "non_authoritative_unhide_diagnostic_count": sum(
            1
            for analysis in analyses
            if analysis["recommended_next_action"]
            == "add_non_authoritative_unhide_or_clear_filter_diagnostic_before_coordinate_normalization"
        ),
    }


def _parser_observations(
    sheets: list[dict[str, Any]],
    analyses: list[dict[str, Any]],
) -> list[dict[str, str]]:
    observations = [
        {
            "level": "info",
            "message": "View-state profile is deterministic workbook XML evidence. It explains render behavior but is not semantic truth.",
        }
    ]
    diagnostic_count = sum(
        1
        for analysis in analyses
        if analysis["recommended_next_action"]
        == "add_non_authoritative_unhide_or_clear_filter_diagnostic_before_coordinate_normalization"
    )
    if diagnostic_count:
        observations.append(
            {
                "level": "warning",
                "message": f"{diagnostic_count} capture windows need a non-authoritative unhide or clear-filter diagnostic before coordinate normalization.",
            }
        )
    skipped = [sheet["name"] for sheet in sheets if sheet["detail_status"] == "skipped_large_xml"]
    if skipped:
        observations.append(
            {
                "level": "warning",
                "message": f"Skipped large worksheet XML for view-state profiling: {', '.join(skipped)}.",
            }
        )
    return observations


def _empty_row_summary() -> dict[str, Any]:
    return {
        "row_count": None,
        "hidden_row_count": 0,
        "zero_height_row_count": 0,
        "outline_row_count": 0,
        "collapsed_row_count": 0,
        "visible_row_count": None,
        "hidden_row_ratio": None,
        "hidden_spans": [],
        "zero_height_spans": [],
        "outline_spans": [],
        "collapsed_spans": [],
    }


def _empty_column_summary() -> dict[str, Any]:
    return {
        "column_count": None,
        "hidden_column_count": 0,
        "zero_width_column_count": 0,
        "outline_column_count": 0,
        "collapsed_column_count": 0,
        "visible_column_count": None,
        "hidden_column_ratio": None,
        "hidden_spans": [],
        "zero_width_spans": [],
        "outline_spans": [],
        "collapsed_spans": [],
    }


def _bounds_from_range(value: str | None) -> dict[str, int] | None:
    if not value:
        return None
    try:
        min_col, min_row, max_col, max_row = range_boundaries(value)
    except ValueError:
        return None
    return {
        "min_row": min_row,
        "min_column": min_col,
        "max_row": max_row,
        "max_column": max_col,
    }


def _bounds_intersection(
    left: dict[str, int],
    right: dict[str, int],
) -> dict[str, int] | None:
    min_row = max(left["min_row"], right["min_row"])
    max_row = min(left["max_row"], right["max_row"])
    min_column = max(left["min_column"], right["min_column"])
    max_column = min(left["max_column"], right["max_column"])
    if min_row > max_row or min_column > max_column:
        return None
    return {
        "min_row": min_row,
        "min_column": min_column,
        "max_row": max_row,
        "max_column": max_column,
    }


def _ratio(numerator: int, denominator: int) -> float | None:
    if denominator <= 0:
        return None
    return round(numerator / denominator, 6)


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _int_or_zero(value: str | None) -> int:
    if value is None:
        return 0
    try:
        return int(value)
    except ValueError:
        return 0


def _float_or_none(value: str | None) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except ValueError:
        return None


def _read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Profile workbook hidden rows, filters, panes, and other view-state evidence."
    )
    parser.add_argument("manifest", type=Path)
    parser.add_argument("capture_quality", type=Path, nargs="*")
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--max-sheet-xml-bytes", type=int, default=100_000_000)
    args = parser.parse_args()
    package = build_view_state_profile(
        args.manifest,
        args.capture_quality,
        max_sheet_xml_bytes=args.max_sheet_xml_bytes,
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(package, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(package["summary"], ensure_ascii=False))


if __name__ == "__main__":
    main()
