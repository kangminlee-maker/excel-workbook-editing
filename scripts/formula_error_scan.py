from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

EXCEL_ERROR_VALUES = {
    "#NULL!",
    "#DIV/0!",
    "#VALUE!",
    "#REF!",
    "#NAME?",
    "#NUM!",
    "#N/A",
    "#GETTING_DATA",
}


def scan_formula_errors(workbook_path: Path, *, sample_limit: int = 50) -> dict[str, Any]:
    workbook_path = workbook_path.expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"missing workbook: {workbook_path}")

    wb_formula = load_workbook(workbook_path, read_only=False, data_only=False)
    wb_values = load_workbook(workbook_path, read_only=False, data_only=True)
    try:
        sheets = []
        for ws_formula in wb_formula.worksheets:
            ws_values = wb_values[ws_formula.title]
            sheets.append(
                _scan_sheet(ws_formula, ws_values, sample_limit=sample_limit)
            )

        summary = {
            "formula_cell_count": sum(s["formula_cell_count"] for s in sheets),
            "literal_error_count": sum(s["literal_error_count"] for s in sheets),
            "formula_error_count": sum(s["formula_error_count"] for s in sheets),
            "blank_cached_formula_count": sum(
                s["blank_cached_formula_count"] for s in sheets
            ),
        }
        summary["error_count"] = (
            summary["literal_error_count"] + summary["formula_error_count"]
        )
        summary["status"] = "pass" if summary["error_count"] == 0 else "fail"

        return {
            "workbook": str(workbook_path),
            "summary": summary,
            "sheets": sheets,
            "notes": [
                "Run this after real Excel recalculation when possible.",
                "blank_cached_formula_count is a warning signal, not an error by itself.",
            ],
        }
    finally:
        wb_formula.close()
        wb_values.close()


def _scan_sheet(
    ws_formula: Any,
    ws_values: Any,
    *,
    sample_limit: int,
) -> dict[str, Any]:
    formula_cell_count = 0
    literal_error_samples: list[dict[str, Any]] = []
    formula_error_samples: list[dict[str, Any]] = []
    blank_cached_formula_samples: list[dict[str, Any]] = []
    literal_error_count = 0
    formula_error_count = 0
    blank_cached_formula_count = 0

    for cell in ws_formula._cells.values():
        value = cell.value
        is_formula = cell.data_type == "f" or (
            isinstance(value, str) and value.startswith("=")
        )
        is_literal_error = cell.data_type == "e" or value in EXCEL_ERROR_VALUES

        if is_literal_error and not is_formula:
            literal_error_count += 1
            if len(literal_error_samples) < sample_limit:
                literal_error_samples.append(
                    {"cell": cell.coordinate, "value": value}
                )

        if not is_formula:
            continue

        formula_cell_count += 1
        cached_cell = ws_values[cell.coordinate]
        cached_value = cached_cell.value
        cached_is_error = (
            cached_cell.data_type == "e" or cached_value in EXCEL_ERROR_VALUES
        )

        if cached_is_error:
            formula_error_count += 1
            if len(formula_error_samples) < sample_limit:
                formula_error_samples.append(
                    {
                        "cell": cell.coordinate,
                        "formula": value,
                        "cached_value": cached_value,
                    }
                )
        elif cached_value is None:
            blank_cached_formula_count += 1
            if len(blank_cached_formula_samples) < sample_limit:
                blank_cached_formula_samples.append(
                    {"cell": cell.coordinate, "formula": value}
                )

    return {
        "name": ws_formula.title,
        "formula_cell_count": formula_cell_count,
        "literal_error_count": literal_error_count,
        "formula_error_count": formula_error_count,
        "blank_cached_formula_count": blank_cached_formula_count,
        "literal_error_samples": literal_error_samples,
        "formula_error_samples": formula_error_samples,
        "blank_cached_formula_samples": blank_cached_formula_samples,
    }


def _to_markdown(report: dict[str, Any]) -> str:
    summary = report["summary"]
    lines = [
        f"# Formula Error Scan: {Path(report['workbook']).name}",
        "",
        "## Summary",
        "",
        f"- Status: {summary['status']}",
        f"- Formula cells: {summary['formula_cell_count']}",
        f"- Formula errors: {summary['formula_error_count']}",
        f"- Literal error cells: {summary['literal_error_count']}",
        f"- Blank cached formulas: {summary['blank_cached_formula_count']}",
        "",
        "## Sheets",
        "",
        "| Sheet | Formulas | Formula errors | Literal errors | Blank cached formulas |",
        "|---|---:|---:|---:|---:|",
    ]
    for sheet in report["sheets"]:
        lines.append(
            "| {name} | {formula_cell_count} | {formula_error_count} | "
            "{literal_error_count} | {blank_cached_formula_count} |".format(**sheet)
        )
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Scan workbook formula and literal error cells."
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--sample-limit", type=int, default=50)
    parser.add_argument(
        "--format",
        choices=["json", "markdown"],
        default="json",
        help="Output format. Defaults to json for machine consumption.",
    )
    args = parser.parse_args()

    report = scan_formula_errors(args.workbook, sample_limit=args.sample_limit)
    if args.format == "markdown":
        print(_to_markdown(report))
    else:
        print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
