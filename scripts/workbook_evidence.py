from __future__ import annotations

import argparse
import hashlib
import json
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

SCHEMA_VERSION = "0.1"


def build_evidence_package(
    workbook_path: Path,
    *,
    max_cell_observations: int = 1000,
    preview_chars: int = 160,
) -> dict[str, Any]:
    workbook_path = workbook_path.expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"missing workbook: {workbook_path}")

    wb = load_workbook(workbook_path, read_only=False, data_only=False)
    try:
        sheets = [
            _sheet_evidence(
                ws,
                sheet_index=index,
                max_cell_observations=max_cell_observations,
                preview_chars=preview_chars,
            )
            for index, ws in enumerate(wb.worksheets)
        ]
        coordinate_maps = {
            "status": "grid_only",
            "detail": "Phase 1 maps workbook ranges to grid bounds only; canvas and capture coordinates are added by render capture.",
            "sheets": [_sheet_coordinate_seed(sheet) for sheet in sheets],
        }
        return {
            "schema_version": SCHEMA_VERSION,
            "package_kind": "direct_workbook_observation",
            "generated_at": _utc_now(),
            "source": {
                "path": str(workbook_path),
                "file_name": workbook_path.name,
                "size_bytes": workbook_path.stat().st_size,
                "sha256": _sha256(workbook_path),
            },
            "evidence_layers": {
                "workbook_structure": {
                    "status": "available",
                    "detail": "Extracted with openpyxl from workbook XML.",
                },
                "formula_dataflow": {
                    "status": "partial",
                    "detail": "Formula expressions are collected; dependency graph extraction is not implemented in Phase 1.",
                },
                "rendered_visual": {
                    "status": "not_captured",
                    "detail": "Excel render capture is planned for the next implementation stage.",
                },
            },
            "workbook": {
                "active_sheet": wb.active.title if wb.worksheets else None,
                "sheet_count": len(wb.worksheets),
                "named_ranges": _defined_names(wb),
                "external_link_count": len(getattr(wb, "_external_links", []) or []),
            },
            "sheets": sheets,
            "coordinate_maps": coordinate_maps,
            "summary": _summary(sheets),
            "limits": {
                "max_cell_observations_per_sheet": max_cell_observations,
                "preview_chars": preview_chars,
            },
            "parser_observations": _package_observations(sheets),
        }
    finally:
        wb.close()


def build_evidence_package_from_artifacts(
    *,
    manifest_path: Path,
    readonly_sample_path: Path,
    formula_patterns_path: Path,
    structural_style_profile_path: Path,
    block_candidates_path: Path,
    table_io_pipelines_path: Path,
    cross_validation_plan_path: Path,
    render_captures_path: Path,
    capture_quality_path: Path,
    recapture_candidate_plan_path: Path,
    recapture_candidate_captures_path: Path,
    recapture_candidate_quality_path: Path,
    view_state_preflight_path: Path,
    view_state_profile_path: Path,
    coordinate_normalization_path: Path,
    visual_features_path: Path,
    gate_execution_path: Path,
    boundary_decisions_path: Path,
    pipeline_role_validation_path: Path,
    general_domain_root: Path | None = None,
) -> dict[str, Any]:
    artifact_paths = {
        "manifest": manifest_path,
        "readonly_sample": readonly_sample_path,
        "formula_patterns": formula_patterns_path,
        "structural_style_profile": structural_style_profile_path,
        "block_candidates": block_candidates_path,
        "table_io_pipelines": table_io_pipelines_path,
        "cross_validation_plan": cross_validation_plan_path,
        "render_captures": render_captures_path,
        "capture_quality": capture_quality_path,
        "recapture_candidate_plan": recapture_candidate_plan_path,
        "recapture_candidate_captures": recapture_candidate_captures_path,
        "recapture_candidate_quality": recapture_candidate_quality_path,
        "view_state_preflight": view_state_preflight_path,
        "view_state_profile": view_state_profile_path,
        "coordinate_normalization": coordinate_normalization_path,
        "visual_features": visual_features_path,
        "gate_execution": gate_execution_path,
        "boundary_decisions": boundary_decisions_path,
        "pipeline_role_validation": pipeline_role_validation_path,
    }
    resolved_paths = {
        name: path.expanduser().resolve()
        for name, path in artifact_paths.items()
    }
    artifacts = {
        name: _read_json(path)
        for name, path in resolved_paths.items()
    }
    manifest = artifacts["manifest"]
    sheets = _artifact_sheets(
        manifest=manifest,
        readonly_sample=artifacts["readonly_sample"],
        structural_style_profile=artifacts["structural_style_profile"],
        view_state_preflight=artifacts["view_state_preflight"],
    )
    domain_refs = _domain_knowledge_refs(general_domain_root)
    return {
        "schema_version": SCHEMA_VERSION,
        "package_kind": "artifact_assembled_workbook_understanding",
        "generated_at": _utc_now(),
        "source": manifest["source"],
        "source_artifacts": {
            name: str(path)
            for name, path in resolved_paths.items()
        },
        "artifact_inventory": _artifact_inventory(artifacts, resolved_paths),
        "evidence_layers": _artifact_evidence_layers(artifacts, domain_refs),
        "workbook": _artifact_workbook(manifest),
        "sheets": sheets,
        "coordinate_maps": _artifact_coordinate_maps(
            artifacts["coordinate_normalization"],
            sheets,
        ),
        "decision_indexes": _decision_indexes(
            artifacts["gate_execution"],
            artifacts["boundary_decisions"],
            artifacts["pipeline_role_validation"],
        ),
        "domain_knowledge_refs": domain_refs,
        "lineage_refs": _lineage_refs(artifacts),
        "review_queue": _review_queue(
            artifacts["gate_execution"],
            artifacts["boundary_decisions"],
            artifacts["pipeline_role_validation"],
        ),
        "summary": _artifact_summary(sheets, artifacts, domain_refs),
        "limits": {
            "max_cell_observations_per_sheet": max(
                int(artifacts["readonly_sample"].get("limits", {}).get("default_max_rows") or 1),
                1,
            ),
            "preview_chars": max(
                int(artifacts["readonly_sample"].get("limits", {}).get("preview_chars") or 1),
                1,
            ),
        },
        "parser_observations": _artifact_package_observations(artifacts, domain_refs),
    }


def _artifact_sheets(
    *,
    manifest: dict[str, Any],
    readonly_sample: dict[str, Any],
    structural_style_profile: dict[str, Any],
    view_state_preflight: dict[str, Any],
) -> list[dict[str, Any]]:
    sample_by_name = {
        sheet["name"]: sheet
        for sheet in readonly_sample.get("sheets", [])
    }
    style_by_name = {
        sheet["name"]: sheet
        for sheet in structural_style_profile.get("sheets", [])
    }
    view_state_by_name = {
        sheet["name"]: sheet
        for sheet in view_state_preflight.get("sheets", [])
    }
    sheets = []
    for index, manifest_sheet in enumerate(manifest.get("workbook", {}).get("sheets", [])):
        name = manifest_sheet["name"]
        style_sheet = style_by_name.get(name, {})
        sample_sheet = sample_by_name.get(name, {})
        view_state_sheet = view_state_by_name.get(name, {})
        bounds = manifest_sheet.get("dimension_bounds") or {
            "min_row": 1,
            "min_column": 1,
            "max_row": 1,
            "max_column": 1,
        }
        cells, formula_observations, cell_summary = _artifact_cells(sample_sheet)
        merged_ranges = [
            _range_payload(item["range"])
            for item in style_sheet.get("merge_ranges", [])
            if item.get("range")
        ]
        objects = {
            "images": _artifact_objects(manifest_sheet.get("drawing_objects", []), "image"),
            "charts": [],
        }
        return_sheet = {
            "name": name,
            "index": index,
            "sheet_state": manifest_sheet.get("state") or view_state_sheet.get("sheet_state") or "visible",
            "dimensions": manifest_sheet.get("dimension") or "A1:A1",
            "grid_bounds": _normalize_bounds(bounds),
            "max_row": int(bounds.get("max_row") or sample_sheet.get("max_row") or 1),
            "max_column": int(bounds.get("max_column") or sample_sheet.get("max_column") or 1),
            "freeze_panes": _freeze_panes(view_state_sheet),
            "auto_filter_ref": _auto_filter_ref(view_state_sheet),
            "print_area": None,
            "default_row_height": (
                view_state_sheet.get("sheet_format_pr", {}).get("defaultRowHeight")
            ),
            "default_column_width": None,
            "row_dimensions": _artifact_row_dimensions(style_sheet, view_state_sheet),
            "column_dimensions": _artifact_column_dimensions(style_sheet, view_state_sheet),
            "merged_ranges": merged_ranges,
            "tables": _artifact_tables(manifest_sheet),
            "cells": cells,
            "cell_observation_summary": cell_summary,
            "formula_observations": formula_observations,
            "objects": objects,
            "visual_feature_seeds": _visual_feature_seeds(
                merged_ranges=merged_ranges,
                tables=_artifact_tables(manifest_sheet),
                objects=objects,
            ),
            "parser_observations": _artifact_sheet_observations(
                manifest_sheet,
                sample_sheet,
                view_state_sheet,
            ),
        }
        sheets.append(return_sheet)
    return sheets


def _artifact_cells(
    sample_sheet: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    cells = []
    formulas = []
    non_empty_count = 0
    formula_count = 0
    for window in sample_sheet.get("windows", []):
        for row in window.get("rows", []):
            for sample_cell in row.get("cells", []):
                value_type = sample_cell.get("value_type") or "blank"
                if value_type != "blank":
                    non_empty_count += 1
                formula = sample_cell.get("formula")
                if formula:
                    formula_count += 1
                    formulas.append(
                        {
                            "cell": sample_cell["cell"],
                            "formula": formula,
                            "value_preview": sample_cell.get("value_preview"),
                            "number_format": "",
                        }
                    )
                cells.append(
                    {
                        "cell": sample_cell["cell"],
                        "row": sample_cell["row"],
                        "column": sample_cell["column"],
                        "value_type": value_type,
                        "value_preview": sample_cell.get("value_preview"),
                        "formula": formula,
                        "number_format": "",
                        "style": _empty_style_summary(),
                        "alignment": _empty_alignment_summary(),
                        "border_sides": [],
                    }
                )
    return (
        cells,
        formulas,
        {
            "non_empty_cell_count": non_empty_count,
            "formula_cell_count": formula_count,
            "styled_cell_count": 0,
            "observed_cell_count": len(cells),
            "omitted_cell_count": 0,
            "truncated": False,
        },
    )


def _artifact_objects(objects: list[dict[str, Any]], object_type: str) -> list[dict[str, Any]]:
    return [
        {
            "id": item.get("id") or f"{object_type}_{index + 1}",
            "type": object_type,
            "anchor": {
                "kind": item.get("source", {}).get("kind") or "drawing_object",
                "from": item.get("from"),
                "to": item.get("to"),
            },
        }
        for index, item in enumerate(objects)
    ]


def _artifact_tables(manifest_sheet: dict[str, Any]) -> list[dict[str, Any]]:
    tables = []
    relationships = manifest_sheet.get("relationships", {})
    if isinstance(relationships, dict):
        table_refs = relationships.get("table_parts", [])
    elif isinstance(relationships, list):
        table_refs = [
            item
            for item in relationships
            if str(item.get("type", "")).lower().endswith("/table")
        ]
    else:
        table_refs = []
    for index, table_ref in enumerate(table_refs):
        tables.append({"name": f"table_{index + 1}", "range": str(table_ref)})
    return tables


def _artifact_row_dimensions(
    style_sheet: dict[str, Any],
    view_state_sheet: dict[str, Any],
) -> list[dict[str, Any]]:
    hidden_rows = {
        row
        for span in view_state_sheet.get("hidden_row_spans", [])
        for row in range(int(span["start_row"]), int(span["end_row"]) + 1)
        if int(span["end_row"]) - int(span["start_row"]) < 1000
    }
    rows = []
    for dim in style_sheet.get("row_dimensions", []):
        row_number = int(dim["row"])
        rows.append(
            {
                "row": row_number,
                "height": dim.get("height"),
                "hidden": bool(dim.get("hidden")) or row_number in hidden_rows,
                "outline_level": 0,
            }
        )
    return rows


def _artifact_column_dimensions(
    style_sheet: dict[str, Any],
    view_state_sheet: dict[str, Any],
) -> list[dict[str, Any]]:
    hidden_columns = {
        column
        for span in view_state_sheet.get("hidden_column_spans", [])
        for column in range(int(span["start_column"]), int(span["end_column"]) + 1)
    }
    columns = []
    for dim in style_sheet.get("column_dimensions", []):
        min_column = int(dim["min_column"])
        max_column = int(dim["max_column"])
        for column in range(min_column, max_column + 1):
            columns.append(
                {
                    "column": get_column_letter(column),
                    "width": dim.get("width"),
                    "hidden": bool(dim.get("hidden")) or column in hidden_columns,
                    "outline_level": 0,
                }
            )
    return columns


def _freeze_panes(view_state_sheet: dict[str, Any]) -> str | None:
    panes = view_state_sheet.get("panes", [])
    if not panes:
        return None
    top_left = panes[0].get("topLeftCell")
    if top_left:
        return str(top_left)
    return json.dumps(panes[0], ensure_ascii=False, sort_keys=True)


def _auto_filter_ref(view_state_sheet: dict[str, Any]) -> str | None:
    filters = view_state_sheet.get("auto_filters", [])
    if not filters:
        return None
    first = filters[0]
    if isinstance(first, dict):
        return first.get("ref")
    return str(first)


def _normalize_bounds(bounds: dict[str, Any]) -> dict[str, int]:
    return {
        "min_row": int(bounds.get("min_row") or 1),
        "min_column": int(bounds.get("min_column") or 1),
        "max_row": int(bounds.get("max_row") or 1),
        "max_column": int(bounds.get("max_column") or 1),
    }


def _empty_style_summary() -> dict[str, Any]:
    return {
        "has_style": False,
        "bold": False,
        "italic": False,
        "font_size": None,
        "fill_type": None,
        "fill_color": None,
    }


def _empty_alignment_summary() -> dict[str, Any]:
    return {"horizontal": None, "vertical": None, "wrap_text": False}


def _artifact_sheet_observations(
    manifest_sheet: dict[str, Any],
    sample_sheet: dict[str, Any],
    view_state_sheet: dict[str, Any],
) -> list[dict[str, str]]:
    observations = []
    if manifest_sheet.get("detail_status") != "scanned":
        observations.append(
            {
                "level": "warning",
                "message": f"Manifest detail status is {manifest_sheet.get('detail_status')}; sheet evidence is partial.",
            }
        )
    if not sample_sheet:
        observations.append(
            {"level": "warning", "message": "Read-only sample is missing for this sheet."}
        )
    view_summary = view_state_sheet.get("summary", {})
    if view_summary.get("hidden_row_count") or view_summary.get("hidden_column_count"):
        observations.append(
            {
                "level": "warning",
                "message": "Sheet has hidden rows or columns; visible-state and structural authority must remain separate.",
            }
        )
    return observations


def _artifact_workbook(manifest: dict[str, Any]) -> dict[str, Any]:
    sheets = manifest.get("workbook", {}).get("sheets", [])
    active_sheet = next(
        (sheet["name"] for sheet in sheets if sheet.get("state") == "visible"),
        sheets[0]["name"] if sheets else None,
    )
    return {
        "active_sheet": active_sheet,
        "sheet_count": int(manifest.get("workbook", {}).get("sheet_count") or len(sheets)),
        "named_ranges": [],
        "external_link_count": len(manifest.get("workbook", {}).get("external_links", [])),
    }


def _artifact_coordinate_maps(
    coordinate_normalization: dict[str, Any],
    sheets: list[dict[str, Any]],
) -> dict[str, Any]:
    mappings = coordinate_normalization.get("coordinate_mappings", [])
    return {
        "status": "normalized_captures_available" if mappings else "grid_only",
        "detail": "Grid coordinates are available for all sheets; capture mappings are included for normalized render captures.",
        "sheets": [_sheet_coordinate_seed(sheet) for sheet in sheets],
        "capture_mappings": [
            {
                "id": mapping.get("id"),
                "status": mapping.get("status"),
                "sheet": mapping.get("sheet"),
                "cell_range": mapping.get("cell_range"),
                "capture_id": mapping.get("capture_id"),
                "target_id": mapping.get("target_id"),
                "quality_status": mapping.get("quality_status"),
                "view_state_classification": mapping.get("view_state_classification"),
            }
            for mapping in mappings
        ],
    }


def _artifact_evidence_layers(
    artifacts: dict[str, dict[str, Any]],
    domain_refs: list[dict[str, Any]],
) -> dict[str, Any]:
    visual_summary = artifacts["visual_features"].get("summary", {})
    return {
        "workbook_structure": {
            "status": "available",
            "detail": "Fast ZIP/XML manifest, read-only samples, structural style profile, and block candidates are available.",
        },
        "formula_dataflow": {
            "status": "available",
            "detail": "Formula pattern profile, table I/O pipelines, gate execution, and pipeline role validation are available.",
        },
        "rendered_visual": {
            "status": "partial",
            "detail": (
                f"{visual_summary.get('detected_count', 0)} captures have detected visual features; "
                "uncaptured and view-state-blocked regions remain review evidence."
            ),
        },
        "view_state": {
            "status": "available",
            "detail": "View-state preflight and capture reconciliation artifacts are available.",
        },
        "boundary_decisions": {
            "status": "available",
            "detail": "Accepted and review-required boundary decisions are available.",
        },
        "pipeline_roles": {
            "status": "available",
            "detail": "Pipeline role validations are available.",
        },
        "domain_knowledge": {
            "status": "available" if domain_refs else "unavailable",
            "detail": "General-domain evidence refs are available." if domain_refs else "No domain refs were supplied.",
        },
    }


def _artifact_inventory(
    artifacts: dict[str, dict[str, Any]],
    paths: dict[str, Path],
) -> list[dict[str, Any]]:
    inventory = []
    for name, artifact in artifacts.items():
        inventory.append(
            {
                "id": name,
                "path": str(paths[name]),
                "schema_version": artifact.get("schema_version"),
                "generated_at": artifact.get("generated_at"),
                "summary": artifact.get("summary", {}),
            }
        )
    return inventory


def _decision_indexes(
    gate_execution: dict[str, Any],
    boundary_decisions: dict[str, Any],
    pipeline_role_validation: dict[str, Any],
) -> dict[str, list[str]]:
    gate_results = gate_execution.get("gate_results", [])
    boundaries = boundary_decisions.get("boundary_decisions", [])
    roles = pipeline_role_validation.get("role_validations", [])
    return {
        "accepted_gate_result_ids": _ids_by_status(gate_results, "accepted"),
        "rejected_gate_result_ids": _ids_by_status(gate_results, "rejected"),
        "review_required_gate_result_ids": _ids_by_status(gate_results, "review_required"),
        "accepted_boundary_decision_ids": _ids_by_status(boundaries, "accepted"),
        "rejected_boundary_decision_ids": _ids_by_status(boundaries, "rejected"),
        "review_required_boundary_decision_ids": _ids_by_status(boundaries, "review_required"),
        "accepted_pipeline_role_validation_ids": _ids_by_status(roles, "accepted"),
        "rejected_pipeline_role_validation_ids": _ids_by_status(roles, "rejected"),
        "review_required_pipeline_role_validation_ids": _ids_by_status(roles, "review_required"),
    }


def _ids_by_status(items: list[dict[str, Any]], status: str) -> list[str]:
    return sorted(str(item["id"]) for item in items if item.get("status") == status)


def _domain_knowledge_refs(general_domain_root: Path | None) -> list[dict[str, Any]]:
    if general_domain_root is None:
        return []
    root = general_domain_root.expanduser().resolve()
    if not root.exists():
        return []
    refs = []
    for path in sorted(root.glob("*.md")):
        refs.append(
            {
                "id": f"general_domain:accounting-kr/{path.name}",
                "layer": "general_domain",
                "path": str(path),
                "scope": "current_sample_workbook",
                "status": "available",
            }
        )
    return refs


def _lineage_refs(artifacts: dict[str, dict[str, Any]]) -> list[dict[str, str]]:
    refs = []
    for target, artifact in artifacts.items():
        for source in (artifact.get("source_artifacts") or {}):
            refs.append({"from": source, "to": target, "relation": "source_artifact"})
    refs.extend(
        [
            {"from": "manifest", "to": "readonly_sample", "relation": "same_source_workbook"},
            {"from": "block_candidates", "to": "table_io_pipelines", "relation": "projects"},
            {"from": "table_io_pipelines", "to": "pipeline_role_validation", "relation": "validates"},
            {"from": "gate_execution", "to": "boundary_decisions", "relation": "validates"},
        ]
    )
    return refs


def _review_queue(
    gate_execution: dict[str, Any],
    boundary_decisions: dict[str, Any],
    pipeline_role_validation: dict[str, Any],
) -> list[dict[str, Any]]:
    queue = []
    for item in gate_execution.get("gate_results", []):
        if item.get("status") == "review_required":
            queue.append(_review_item("gate_result", item))
    for item in boundary_decisions.get("boundary_decisions", []):
        if item.get("status") == "review_required":
            queue.append(_review_item("boundary_decision", item))
    for item in pipeline_role_validation.get("role_validations", []):
        if item.get("status") == "review_required":
            queue.append(_review_item("pipeline_role_validation", item))
    return queue


def _review_item(kind: str, item: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": item["id"],
        "kind": kind,
        "status": item.get("status"),
        "reason": item.get("reason"),
        "sheet": item.get("sheet") or item.get("output_ref", {}).get("sheet"),
        "range": item.get("range") or item.get("output_ref", {}).get("range"),
        "evidence_refs": item.get("evidence_refs", []),
    }


def _artifact_summary(
    sheets: list[dict[str, Any]],
    artifacts: dict[str, dict[str, Any]],
    domain_refs: list[dict[str, Any]],
) -> dict[str, int]:
    direct_summary = _summary(sheets)
    gate_summary = artifacts["gate_execution"].get("summary", {})
    boundary_summary = artifacts["boundary_decisions"].get("summary", {})
    role_summary = artifacts["pipeline_role_validation"].get("summary", {})
    visual_summary = artifacts["visual_features"].get("summary", {})
    capture_summary = artifacts["render_captures"].get("summary", {})
    direct_summary.update(
        {
            "artifact_count": len(artifacts),
            "capture_result_count": (
                capture_summary.get("capture_count")
                or capture_summary.get("captured_count")
                or capture_summary.get("selected_target_count")
                or 0
            ),
            "visual_feature_result_count": visual_summary.get("feature_result_count", 0),
            "accepted_gate_count": gate_summary.get("accepted_count", 0),
            "review_required_gate_count": gate_summary.get("review_required_count", 0),
            "accepted_boundary_count": boundary_summary.get("accepted_count", 0),
            "review_required_boundary_count": boundary_summary.get("review_required_count", 0),
            "accepted_pipeline_role_count": role_summary.get("accepted_count", 0),
            "review_required_pipeline_role_count": role_summary.get("review_required_count", 0),
            "review_queue_count": (
                gate_summary.get("review_required_count", 0)
                + boundary_summary.get("review_required_count", 0)
                + role_summary.get("review_required_count", 0)
            ),
            "domain_knowledge_ref_count": len(domain_refs),
        }
    )
    return direct_summary


def _artifact_package_observations(
    artifacts: dict[str, dict[str, Any]],
    domain_refs: list[dict[str, Any]],
) -> list[dict[str, str]]:
    observations = [
        {
            "level": "info",
            "message": "Evidence package assembled from prior deterministic artifacts; it does not reopen or mutate the source workbook.",
        },
        {
            "level": "warning",
            "message": "Formula results are not Excel-engine recalculation authority in this package.",
        },
    ]
    review_queue_count = (
        artifacts["gate_execution"].get("summary", {}).get("review_required_count", 0)
        + artifacts["boundary_decisions"].get("summary", {}).get("review_required_count", 0)
        + artifacts["pipeline_role_validation"].get("summary", {}).get("review_required_count", 0)
    )
    if review_queue_count:
        observations.append(
            {
                "level": "warning",
                "message": f"{review_queue_count} gate, boundary, or pipeline-role items remain review-required.",
            }
        )
    if domain_refs:
        observations.append(
            {
                "level": "info",
                "message": "General-domain knowledge refs are attached separately from workbook evidence.",
            }
        )
    return observations


def _sheet_evidence(
    ws: Any,
    *,
    sheet_index: int,
    max_cell_observations: int,
    preview_chars: int,
) -> dict[str, Any]:
    dimensions = ws.calculate_dimension()
    min_col, min_row, max_col, max_row = range_boundaries(dimensions)
    cells, formula_observations, cell_summary = _cell_observations(
        ws,
        max_cell_observations=max_cell_observations,
        preview_chars=preview_chars,
    )
    merged_ranges = [_range_payload(str(rng)) for rng in ws.merged_cells.ranges]
    tables = _table_observations(ws)
    objects = {
        "images": _object_observations(getattr(ws, "_images", []) or [], "image"),
        "charts": _object_observations(getattr(ws, "_charts", []) or [], "chart"),
    }

    return {
        "name": ws.title,
        "index": sheet_index,
        "sheet_state": ws.sheet_state,
        "dimensions": dimensions,
        "grid_bounds": {
            "min_row": min_row,
            "min_column": min_col,
            "max_row": max_row,
            "max_column": max_col,
        },
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
        "auto_filter_ref": ws.auto_filter.ref,
        "print_area": _json_value(getattr(ws, "print_area", None)),
        "default_row_height": _json_value(ws.sheet_format.defaultRowHeight),
        "default_column_width": _json_value(ws.sheet_format.defaultColWidth),
        "row_dimensions": _row_dimensions(ws),
        "column_dimensions": _column_dimensions(ws),
        "merged_ranges": merged_ranges,
        "tables": tables,
        "cells": cells,
        "cell_observation_summary": cell_summary,
        "formula_observations": formula_observations,
        "objects": objects,
        "visual_feature_seeds": _visual_feature_seeds(
            merged_ranges=merged_ranges,
            tables=tables,
            objects=objects,
        ),
        "parser_observations": _sheet_observations(cell_summary),
    }


def _cell_observations(
    ws: Any,
    *,
    max_cell_observations: int,
    preview_chars: int,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    cells: list[dict[str, Any]] = []
    formulas: list[dict[str, Any]] = []
    non_empty_count = 0
    formula_count = 0
    styled_cell_count = 0
    omitted_count = 0

    sorted_cells = sorted(ws._cells.values(), key=lambda cell: (cell.row, cell.column))
    for cell in sorted_cells:
        if cell.has_style:
            styled_cell_count += 1

        value = cell.value
        is_formula = _is_formula(cell)
        if value is None and not is_formula:
            continue

        non_empty_count += 1
        if is_formula:
            formula_count += 1
            formulas.append(
                {
                    "cell": cell.coordinate,
                    "formula": str(value),
                    "value_preview": _preview(value, preview_chars),
                    "number_format": cell.number_format,
                }
            )

        if len(cells) >= max_cell_observations:
            omitted_count += 1
            continue

        cells.append(
            {
                "cell": cell.coordinate,
                "row": cell.row,
                "column": cell.column,
                "value_type": _value_type(value, is_formula=is_formula),
                "value_preview": _preview(value, preview_chars),
                "formula": str(value) if is_formula else None,
                "number_format": cell.number_format,
                "style": _style_summary(cell),
                "alignment": _alignment_summary(cell),
                "border_sides": _border_sides(cell),
            }
        )

    return (
        cells,
        formulas,
        {
            "non_empty_cell_count": non_empty_count,
            "formula_cell_count": formula_count,
            "styled_cell_count": styled_cell_count,
            "observed_cell_count": len(cells),
            "omitted_cell_count": omitted_count,
            "truncated": omitted_count > 0,
        },
    )


def _defined_names(wb: Any) -> list[dict[str, Any]]:
    names: list[dict[str, Any]] = []
    for name, defined_name in wb.defined_names.items():
        destinations = []
        try:
            destinations = [
                {"sheet": sheet, "range": coord}
                for sheet, coord in defined_name.destinations
            ]
        except Exception:
            destinations = []

        names.append(
            {
                "name": str(name),
                "attr_text": _json_value(getattr(defined_name, "attr_text", None)),
                "hidden": bool(getattr(defined_name, "hidden", False)),
                "destinations": destinations,
            }
        )
    return names


def _table_observations(ws: Any) -> list[dict[str, Any]]:
    tables: list[dict[str, Any]] = []
    for name, table_obj in ws.tables.items():
        ref = table_obj if isinstance(table_obj, str) else getattr(table_obj, "ref", None)
        tables.append({"name": str(name), "range": str(ref or "")})
    return tables


def _object_observations(objects: list[Any], object_type: str) -> list[dict[str, Any]]:
    return [
        {
            "id": f"{object_type}_{index + 1}",
            "type": object_type,
            "anchor": _anchor_summary(getattr(obj, "anchor", None)),
        }
        for index, obj in enumerate(objects)
    ]


def _anchor_summary(anchor: Any) -> dict[str, Any]:
    if anchor is None:
        return {"kind": "unknown"}
    if isinstance(anchor, str):
        return {"kind": "cell_ref", "ref": anchor}

    start = getattr(anchor, "_from", None)
    end = getattr(anchor, "to", None)
    payload: dict[str, Any] = {"kind": type(anchor).__name__}
    if start is not None:
        payload["from"] = _anchor_marker(start)
    if end is not None:
        payload["to"] = _anchor_marker(end)
    return payload


def _anchor_marker(marker: Any) -> dict[str, Any]:
    column = int(getattr(marker, "col", 0)) + 1
    row = int(getattr(marker, "row", 0)) + 1
    return {
        "cell": f"{get_column_letter(column)}{row}",
        "row": row,
        "column": column,
        "row_offset": int(getattr(marker, "rowOff", 0) or 0),
        "column_offset": int(getattr(marker, "colOff", 0) or 0),
    }


def _row_dimensions(ws: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for index, dim in sorted(ws.row_dimensions.items()):
        rows.append(
            {
                "row": int(index),
                "height": _json_value(dim.height),
                "hidden": bool(dim.hidden),
                "outline_level": int(dim.outlineLevel or 0),
            }
        )
    return rows


def _column_dimensions(ws: Any) -> list[dict[str, Any]]:
    columns: list[dict[str, Any]] = []
    for key, dim in sorted(ws.column_dimensions.items()):
        columns.append(
            {
                "column": str(key),
                "width": _json_value(dim.width),
                "hidden": bool(dim.hidden),
                "outline_level": int(dim.outlineLevel or 0),
            }
        )
    return columns


def _range_payload(range_ref: str) -> dict[str, Any]:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    return {
        "range": range_ref,
        "bounds": {
            "min_row": min_row,
            "min_column": min_col,
            "max_row": max_row,
            "max_column": max_col,
        },
    }


def _visual_feature_seeds(
    *,
    merged_ranges: list[dict[str, Any]],
    tables: list[dict[str, Any]],
    objects: dict[str, list[dict[str, Any]]],
) -> list[dict[str, Any]]:
    seeds: list[dict[str, Any]] = []
    for merged_range in merged_ranges:
        seeds.append(
            {
                "type": "merged_range",
                "source": merged_range["range"],
                "reason": "Merged cells often indicate titles, grouped labels, or visual section headers.",
            }
        )
    for table in tables:
        seeds.append(
            {
                "type": "declared_table",
                "source": table["range"],
                "reason": "Workbook XML declares this range as a table.",
            }
        )
    for image in objects["images"]:
        seeds.append(
            {
                "type": "image_anchor",
                "source": image["id"],
                "reason": "Image anchors are candidates for visual document blocks.",
            }
        )
    for chart in objects["charts"]:
        seeds.append(
            {
                "type": "chart_anchor",
                "source": chart["id"],
                "reason": "Chart anchors are candidates for visual document blocks.",
            }
        )
    return seeds


def _sheet_coordinate_seed(sheet: dict[str, Any]) -> dict[str, Any]:
    return {
        "sheet": sheet["name"],
        "status": "grid_only",
        "dimensions": sheet["dimensions"],
        "grid_bounds": sheet["grid_bounds"],
    }


def _summary(sheets: list[dict[str, Any]]) -> dict[str, int]:
    return {
        "sheet_count": len(sheets),
        "hidden_sheet_count": sum(1 for sheet in sheets if sheet["sheet_state"] != "visible"),
        "non_empty_cell_count": sum(
            sheet["cell_observation_summary"]["non_empty_cell_count"]
            for sheet in sheets
        ),
        "formula_cell_count": sum(
            sheet["cell_observation_summary"]["formula_cell_count"]
            for sheet in sheets
        ),
        "styled_cell_count": sum(
            sheet["cell_observation_summary"]["styled_cell_count"]
            for sheet in sheets
        ),
        "merged_range_count": sum(len(sheet["merged_ranges"]) for sheet in sheets),
        "table_count": sum(len(sheet["tables"]) for sheet in sheets),
        "image_count": sum(len(sheet["objects"]["images"]) for sheet in sheets),
        "chart_count": sum(len(sheet["objects"]["charts"]) for sheet in sheets),
    }


def _package_observations(sheets: list[dict[str, Any]]) -> list[dict[str, str]]:
    observations: list[dict[str, str]] = []
    if any(
        sheet["cell_observation_summary"]["truncated"]
        for sheet in sheets
    ):
        observations.append(
            {
                "level": "warning",
                "message": "One or more sheets exceeded the cell observation limit; increase max_cell_observations for fuller evidence.",
            }
        )
    observations.append(
        {
            "level": "info",
            "message": "Rendered visual capture is not included in Phase 1 evidence.",
        }
    )
    return observations


def _sheet_observations(cell_summary: dict[str, Any]) -> list[dict[str, str]]:
    if not cell_summary["truncated"]:
        return []
    return [
        {
            "level": "warning",
            "message": "Cell observations were truncated for this sheet.",
        }
    ]


def _style_summary(cell: Any) -> dict[str, Any]:
    return {
        "has_style": bool(cell.has_style),
        "bold": bool(cell.font.bold),
        "italic": bool(cell.font.italic),
        "font_size": _json_value(cell.font.sz),
        "fill_type": _json_value(cell.fill.fill_type),
        "fill_color": _color(cell.fill.fgColor),
    }


def _alignment_summary(cell: Any) -> dict[str, Any]:
    return {
        "horizontal": _json_value(cell.alignment.horizontal),
        "vertical": _json_value(cell.alignment.vertical),
        "wrap_text": bool(cell.alignment.wrap_text),
    }


def _border_sides(cell: Any) -> list[str]:
    sides: list[str] = []
    for name in ("left", "right", "top", "bottom"):
        side = getattr(cell.border, name)
        if side and side.style:
            sides.append(name)
    return sides


def _color(color_obj: Any) -> str | None:
    if color_obj is None:
        return None
    color_type = getattr(color_obj, "type", None)
    if color_type == "rgb":
        return _json_value(getattr(color_obj, "rgb", None))
    if color_type == "indexed":
        return f"indexed:{getattr(color_obj, 'indexed', None)}"
    if color_type == "theme":
        tint = getattr(color_obj, "tint", 0)
        return f"theme:{getattr(color_obj, 'theme', None)}:{tint}"
    return None


def _is_formula(cell: Any) -> bool:
    value = cell.value
    return cell.data_type == "f" or (isinstance(value, str) and value.startswith("="))


def _value_type(value: Any, *, is_formula: bool) -> str:
    if is_formula:
        return "formula"
    if value is None:
        return "blank"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, (datetime, date, time)):
        return "datetime"
    return "string"


def _preview(value: Any, preview_chars: int) -> str | None:
    if value is None:
        return None
    if isinstance(value, (datetime, date, time)):
        text = value.isoformat()
    else:
        text = str(value)
    if len(text) <= preview_chars:
        return text
    return text[: max(preview_chars - 3, 0)] + "..."


def _json_value(value: Any) -> Any:
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, (list, tuple)):
        return [_json_value(item) for item in value]
    return str(value)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build a workbook evidence package for document-shaped Excel understanding."
    )
    parser.add_argument("workbook", type=Path, nargs="?")
    parser.add_argument("--output", type=Path)
    parser.add_argument("--max-cell-observations", type=int, default=1000)
    parser.add_argument("--preview-chars", type=int, default=160)
    parser.add_argument("--manifest", type=Path)
    parser.add_argument("--readonly-sample", type=Path)
    parser.add_argument("--formula-patterns", type=Path)
    parser.add_argument("--structural-style-profile", type=Path)
    parser.add_argument("--block-candidates", type=Path)
    parser.add_argument("--table-io-pipelines", type=Path)
    parser.add_argument("--cross-validation-plan", type=Path)
    parser.add_argument("--render-captures", type=Path)
    parser.add_argument("--capture-quality", type=Path)
    parser.add_argument("--recapture-candidate-plan", type=Path)
    parser.add_argument("--recapture-candidate-captures", type=Path)
    parser.add_argument("--recapture-candidate-quality", type=Path)
    parser.add_argument("--view-state-preflight", type=Path)
    parser.add_argument("--view-state-profile", type=Path)
    parser.add_argument("--coordinate-normalization", type=Path)
    parser.add_argument("--visual-features", type=Path)
    parser.add_argument("--gate-execution", type=Path)
    parser.add_argument("--boundary-decisions", type=Path)
    parser.add_argument("--pipeline-role-validation", type=Path)
    parser.add_argument("--general-domain-root", type=Path)
    args = parser.parse_args()

    if args.manifest:
        required = [
            "readonly_sample",
            "formula_patterns",
            "structural_style_profile",
            "block_candidates",
            "table_io_pipelines",
            "cross_validation_plan",
            "render_captures",
            "capture_quality",
            "recapture_candidate_plan",
            "recapture_candidate_captures",
            "recapture_candidate_quality",
            "view_state_preflight",
            "view_state_profile",
            "coordinate_normalization",
            "visual_features",
            "gate_execution",
            "boundary_decisions",
            "pipeline_role_validation",
        ]
        missing = [name for name in required if getattr(args, name) is None]
        if missing:
            parser.error(f"--manifest artifact mode is missing: {', '.join(missing)}")
        package = build_evidence_package_from_artifacts(
            manifest_path=args.manifest,
            readonly_sample_path=args.readonly_sample,
            formula_patterns_path=args.formula_patterns,
            structural_style_profile_path=args.structural_style_profile,
            block_candidates_path=args.block_candidates,
            table_io_pipelines_path=args.table_io_pipelines,
            cross_validation_plan_path=args.cross_validation_plan,
            render_captures_path=args.render_captures,
            capture_quality_path=args.capture_quality,
            recapture_candidate_plan_path=args.recapture_candidate_plan,
            recapture_candidate_captures_path=args.recapture_candidate_captures,
            recapture_candidate_quality_path=args.recapture_candidate_quality,
            view_state_preflight_path=args.view_state_preflight,
            view_state_profile_path=args.view_state_profile,
            coordinate_normalization_path=args.coordinate_normalization,
            visual_features_path=args.visual_features,
            gate_execution_path=args.gate_execution,
            boundary_decisions_path=args.boundary_decisions,
            pipeline_role_validation_path=args.pipeline_role_validation,
            general_domain_root=args.general_domain_root,
        )
    else:
        if args.workbook is None:
            parser.error("workbook path is required unless --manifest artifact mode is used")
        package = build_evidence_package(
            args.workbook,
            max_cell_observations=args.max_cell_observations,
            preview_chars=args.preview_chars,
        )
    payload = json.dumps(package, ensure_ascii=False, indent=2)
    if args.output:
        args.output.write_text(payload + "\n", encoding="utf-8")
    else:
        print(payload)


if __name__ == "__main__":
    main()
