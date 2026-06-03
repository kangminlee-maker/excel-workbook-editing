from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl.utils import get_column_letter

SCHEMA_VERSION = "0.1"


def build_cross_validation_plan(
    block_candidates_path: Path,
    table_io_pipelines_path: Path,
) -> dict[str, Any]:
    block_candidates_path = block_candidates_path.expanduser().resolve()
    table_io_pipelines_path = table_io_pipelines_path.expanduser().resolve()
    candidates = _read_json(block_candidates_path)
    pipelines_package = _read_json(table_io_pipelines_path)
    sheet_index = _sheet_index(candidates)
    targets: dict[str, dict[str, Any]] = {}

    for pipeline in pipelines_package.get("pipelines", []):
        _add_pipeline_target(targets, pipeline, sheet_index)

    for sheet in candidates.get("sheets", []):
        for result in sheet.get("boundary_gate_results", []):
            _add_boundary_target(targets, sheet, result, sheet_index)
        for relation in sheet.get("relations", []):
            _add_image_hierarchy_target(targets, relation, sheet_index)

    finalized = [_finalize_target(target) for target in targets.values()]
    finalized.sort(key=lambda target: (-target["score"], target["id"]))
    recommended_first_batch_target_ids = _recommended_first_batch_target_ids(finalized)
    gate_checks = [
        check for target in finalized for check in target.get("gate_checks", [])
    ]
    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at": _utc_now(),
        "source_artifacts": {
            "block_candidates": str(block_candidates_path),
            "table_io_pipelines": str(table_io_pipelines_path),
        },
        "capture_targets": finalized,
        "recommended_first_batch_target_ids": recommended_first_batch_target_ids,
        "gate_checks": gate_checks,
        "summary": _summary(finalized, gate_checks, recommended_first_batch_target_ids),
        "parser_observations": [
            {
                "level": "info",
                "message": "Cross-validation targets prioritize render capture and deterministic visual/formula gates. They are a review plan, not evidence that capture has already happened.",
            }
        ],
    }


def _sheet_index(candidates: dict[str, Any]) -> dict[str, Any]:
    sheets = {sheet["name"]: sheet for sheet in candidates.get("sheets", [])}
    blocks_by_id = {}
    regions_by_id = {}
    for sheet in sheets.values():
        for block in sheet.get("blocks", []):
            blocks_by_id[block["id"]] = block
        for region in sheet.get("cell_regions", []):
            regions_by_id[region["id"]] = region
    return {
        "sheets": sheets,
        "blocks_by_id": blocks_by_id,
        "regions_by_id": regions_by_id,
    }


def _add_pipeline_target(
    targets: dict[str, dict[str, Any]],
    pipeline: dict[str, Any],
    sheet_index: dict[str, Any],
) -> None:
    output_ref = pipeline.get("output_ref") or {}
    sheet = output_ref.get("sheet")
    bounds = _normalize_bounds(output_ref.get("bounds"))
    target_id = _target_id(output_ref.get("kind"), output_ref.get("id"), sheet, bounds)
    target = targets.setdefault(
        target_id,
        _target_seed(target_id, "pipeline_output", sheet, bounds, output_ref),
    )
    _add_unique(target["related_pipeline_ids"], pipeline["id"])
    _add_unique(target["evidence_refs"], pipeline["id"])

    _score(target, 30, "pipeline_output_needs_capture_context")
    role = pipeline.get("role")
    if role == "report":
        _score(target, 25, "report_pipeline")
        _add_gate(
            target,
            "pivot_cache_visual_alignment",
            [pipeline["id"], *pipeline.get("evidence_refs", [])],
            [
                "Captured pivot display bbox exists at the declared output range.",
                "Pivot output remains linked to a pivot cache source rather than being accepted as raw table data.",
            ],
            [
                "Captured bbox is missing or displaced from the workbook XML location.",
                "Displayed pivot range overlaps unrelated blocks without a clear visual boundary.",
            ],
        )
    elif role == "summary":
        _score(target, 20, "summary_formula_pipeline")
        _add_gate(
            target,
            "formula_summary_visual_alignment",
            [pipeline["id"], *pipeline.get("evidence_refs", [])],
            [
                "Summary output visually reads as a summary/report region.",
                "Formula inputs remain traceable to the pipeline input refs.",
            ],
            [
                "Summary formulas visually belong to a different table or section.",
                "Formula refs point outside the visually associated region without a documented bridge.",
            ],
        )
    elif role == "transform":
        _score(target, 8, "transform_pipeline")

    kind = output_ref.get("kind")
    if kind == "pivot_table":
        _score(target, 20, "pivot_table_output")
        block_id = output_ref.get("block_id")
        if block_id:
            _add_unique(target["related_block_ids"], block_id)
    elif kind == "cell_region":
        _score(target, 8, "cell_region_output")
        region_id = output_ref.get("region_id")
        block_id = output_ref.get("block_id")
        if region_id:
            _add_unique(target["related_region_ids"], region_id)
        if block_id:
            _add_unique(target["related_block_ids"], block_id)

    flags = set(pipeline.get("review_flags", []))
    if "unresolved_input_region" in flags:
        _score(target, 35, "unresolved_input_region")
        _add_gate(
            target,
            "unresolved_input_region_resolution",
            [pipeline["id"]],
            [
                "Each workbook_range input is mapped to an existing cell region or explicitly remains a scalar/reference cell.",
                "If the input should be part of a nearby region, the region boundary is expanded or split deterministically.",
            ],
            [
                "A formula input has no owning cell region after capture and boundary review.",
                "The captured visual block contradicts the current input-region mapping.",
            ],
        )
    if "pivot_cache_dependency" in flags:
        _score(target, 25, "pivot_cache_dependency")
    if "external_workbook_dependency" in flags:
        _score(target, 35, "external_workbook_dependency")
    if "repeated_formula_family" in flags:
        _score(target, 15, "repeated_formula_family")
        _add_gate(
            target,
            "formula_region_coherence",
            [pipeline["id"]],
            [
                "Repeated formula family stays inside one visually coherent region or a deliberate repeated section.",
                "Formula signature grouping agrees with the captured table boundary.",
            ],
            [
                "Same formula family spans visually unrelated tables.",
                "Captured boundary cuts through a repeated formula family without a documented exception.",
            ],
        )
    if "same_sheet_dataflow" in flags:
        _score(target, 8, "same_sheet_dataflow")

    for ref in pipeline.get("input_refs", []):
        if ref.get("kind") in {"workbook_range", "pivot_cache_source_range"}:
            _add_gate(
                target,
                "pipeline_input_output_alignment",
                [pipeline["id"], ref.get("id")],
                [
                    "The input range has an owning region, scalar role, or explicit external/source role.",
                    "The output visually belongs to the same calculation story as its formula or pivot inputs.",
                ],
                [
                    "Input/output relation crosses unrelated visual sections without formula or hierarchy evidence.",
                    "Input range points to a hidden or uncaptured region that cannot be reviewed.",
                ],
            )

    _link_overlapping_boundaries(target, sheet_index)


def _add_boundary_target(
    targets: dict[str, dict[str, Any]],
    sheet: dict[str, Any],
    result: dict[str, Any],
    sheet_index: dict[str, Any],
) -> None:
    related_region_ids = result.get("related_region_ids", [])
    region = None
    for region_id in related_region_ids:
        region = sheet_index["regions_by_id"].get(region_id)
        if region:
            break
    bounds = _normalize_bounds(region.get("bounds") if region else None)
    target_ref = _region_target_ref(sheet["name"], region) if region else {
        "id": result["candidate_id"],
        "kind": "boundary_region",
        "workbook": None,
        "sheet": sheet["name"],
        "range": None,
        "block_id": None,
        "region_id": None,
        "bounds": bounds,
        "label": result["candidate_id"],
    }
    target_id = _target_id(target_ref["kind"], target_ref["id"], sheet["name"], bounds)
    target = targets.setdefault(
        target_id,
        _target_seed(target_id, "boundary_review", sheet["name"], bounds, target_ref),
    )
    _add_unique(target["related_boundary_gate_ids"], result["id"])
    _add_unique(target["evidence_refs"], result["id"])
    for region_id in related_region_ids:
        _add_unique(target["related_region_ids"], region_id)

    status = result.get("status")
    if status == "strong_candidate":
        _score(target, 35, "strong_boundary_candidate")
    elif status == "review_candidate":
        _score(target, 20, "review_boundary_candidate")
    else:
        _score(target, 8, "weak_boundary_signal")
    _add_gate(
        target,
        "boundary_confirmation",
        [result["id"], result.get("candidate_id")],
        [
            "Captured visual boundary supports the proposed split or the boundary is demoted.",
            "Formula, header, style, or whitespace evidence agrees with the final split/merge decision.",
        ],
        [
            "Style-only evidence is not visible in capture.",
            "Boundary separates a coherent formula family or pivot/report block without stronger evidence.",
        ],
    )


def _add_image_hierarchy_target(
    targets: dict[str, dict[str, Any]],
    relation: dict[str, Any],
    sheet_index: dict[str, Any],
) -> None:
    from_block = sheet_index["blocks_by_id"].get(relation.get("from"))
    to_block = sheet_index["blocks_by_id"].get(relation.get("to"))
    if not from_block or not to_block:
        return
    if from_block.get("type") != "image" and to_block.get("type") != "image":
        return
    image_block = from_block if from_block.get("type") == "image" else to_block
    other_block = to_block if image_block is from_block else from_block
    bounds = _union_bounds(
        _normalize_bounds(image_block.get("bounds")),
        _normalize_bounds(other_block.get("bounds")),
    )
    target_ref = _block_target_ref(image_block, kind="image_block")
    target_id = f"target_image_hierarchy_{_slug(relation['id'])}"
    target = targets.setdefault(
        target_id,
        _target_seed(
            target_id,
            "image_hierarchy",
            image_block["source"]["sheet"],
            bounds,
            target_ref,
        ),
    )
    _score(target, 65, "image_table_hierarchy_candidate")
    _add_unique(target["related_block_ids"], image_block["id"])
    _add_unique(target["related_block_ids"], other_block["id"])
    _add_unique(target["evidence_refs"], relation["id"])
    _add_gate(
        target,
        "image_table_hierarchy_confirmation",
        [relation["id"]],
        [
            "Captured image and table/region are visually grouped, adjacent, or captioned.",
            "No closer competing block has stronger hierarchy evidence.",
        ],
        [
            "Image is only decorative or unrelated to the neighboring data block.",
            "Capture shows a different reading order than the XML anchor/proximity relation.",
        ],
    )


def _target_seed(
    target_id: str,
    target_type: str,
    sheet: str | None,
    bounds: dict[str, int] | None,
    target_ref: dict[str, Any],
) -> dict[str, Any]:
    return {
        "id": target_id,
        "type": "visual_formula_validation_target",
        "target_type": target_type,
        "status": "candidate",
        "priority": "low",
        "score": 0,
        "sheet": sheet,
        "range": _range_label(bounds),
        "bounds": bounds,
        "capture_window": _capture_window(sheet, bounds),
        "target_ref": _normalize_target_ref(target_ref),
        "related_pipeline_ids": [],
        "related_block_ids": [],
        "related_region_ids": [],
        "related_boundary_gate_ids": [],
        "reasons": [],
        "gate_checks": [],
        "review_questions": [],
        "evidence_refs": [],
    }


def _finalize_target(target: dict[str, Any]) -> dict[str, Any]:
    target["score"] = min(target["score"], 100)
    if target["score"] >= 85:
        target["priority"] = "high"
    elif target["score"] >= 55:
        target["priority"] = "medium"
    else:
        target["priority"] = "low"
    target["reasons"] = sorted(set(target["reasons"]))
    for key in [
        "related_pipeline_ids",
        "related_block_ids",
        "related_region_ids",
        "related_boundary_gate_ids",
        "evidence_refs",
    ]:
        target[key] = sorted(set(item for item in target[key] if item))
    target["gate_checks"] = sorted(
        target["gate_checks"],
        key=lambda check: (check["gate_type"], check["id"]),
    )
    target["review_questions"] = _review_questions(target)
    return target


def _score(target: dict[str, Any], points: int, reason: str) -> None:
    target["score"] += points
    _add_unique(target["reasons"], reason)


def _add_gate(
    target: dict[str, Any],
    gate_type: str,
    deterministic_inputs: list[Any],
    pass_conditions: list[str],
    failure_signals: list[str],
) -> None:
    gate_id = f"gate_{_slug(target['id'])}_{_slug(gate_type)}"
    if any(check["id"] == gate_id for check in target["gate_checks"]):
        return
    target["gate_checks"].append(
        {
            "id": gate_id,
            "type": "visual_formula_gate_check",
            "target_id": target["id"],
            "gate_type": gate_type,
            "status": "pending_capture",
            "deterministic_inputs": [
                str(item) for item in deterministic_inputs if item is not None
            ],
            "pass_conditions": pass_conditions,
            "failure_signals": failure_signals,
        }
    )


def _link_overlapping_boundaries(
    target: dict[str, Any],
    sheet_index: dict[str, Any],
) -> None:
    sheet_name = target.get("sheet")
    bounds = target.get("bounds")
    if not sheet_name or not bounds:
        return
    sheet = sheet_index["sheets"].get(sheet_name)
    if not sheet:
        return
    for result in sheet.get("boundary_gate_results", []):
        related_region_ids = result.get("related_region_ids", [])
        if target["target_ref"].get("region_id") not in related_region_ids:
            continue
        _add_unique(target["related_boundary_gate_ids"], result["id"])
        _add_unique(target["evidence_refs"], result["id"])
        _add_gate(
            target,
            "boundary_confirmation",
            [result["id"], result.get("candidate_id")],
            [
                "Captured visual boundary supports the proposed split or the boundary is demoted.",
                "Formula, header, style, or whitespace evidence agrees with the final split/merge decision.",
            ],
            [
                "Style-only evidence is not visible in capture.",
                "Boundary separates a coherent formula family or pivot/report block without stronger evidence.",
            ],
        )
        if result.get("status") == "strong_candidate":
            _score(target, 12, "overlaps_strong_boundary_candidate")
        elif result.get("status") == "review_candidate":
            _score(target, 6, "overlaps_review_boundary_candidate")


def _review_questions(target: dict[str, Any]) -> list[str]:
    questions = []
    reasons = set(target.get("reasons", []))
    if "unresolved_input_region" in reasons:
        questions.append("이 pipeline input은 독립 scalar/reference인가요, 아니면 주변 표 영역에 포함되어야 하나요?")
    if "pivot_cache_dependency" in reasons or "pivot_table_output" in reasons:
        questions.append("표시된 pivot range가 raw table이 아니라 pivot cache 기반 report로 보이나요?")
    if "review_boundary_candidate" in reasons or "strong_boundary_candidate" in reasons:
        questions.append("캡처 기준으로 이 region 내부/주변 경계가 실제 표 분리로 보이나요?")
    if "image_table_hierarchy_candidate" in reasons:
        questions.append("이미지와 인접 표/영역 사이에 설명/캡션/소속 관계가 있나요?")
    if not questions:
        questions.append("캡처 결과가 현재 셀 범위와 계산 관계를 그대로 지지하나요?")
    return questions


def _region_target_ref(sheet_name: str, region: dict[str, Any]) -> dict[str, Any]:
    bounds = _normalize_bounds(region["bounds"])
    return {
        "id": region["id"],
        "kind": "cell_region",
        "workbook": None,
        "sheet": sheet_name,
        "range": _range_label(bounds),
        "block_id": region.get("parent_seed_block_id"),
        "region_id": region["id"],
        "bounds": bounds,
        "label": region.get("label") or region["id"],
    }


def _block_target_ref(block: dict[str, Any], *, kind: str | None = None) -> dict[str, Any]:
    bounds = _normalize_bounds(block["bounds"])
    return {
        "id": block["id"],
        "kind": kind or block["type"],
        "workbook": None,
        "sheet": block["source"]["sheet"],
        "range": _range_label(bounds),
        "block_id": block["id"],
        "region_id": None,
        "bounds": bounds,
        "label": block.get("label") or block["id"],
    }


def _normalize_target_ref(ref: dict[str, Any]) -> dict[str, Any]:
    normalized = {
        "id": ref.get("id") or "unknown",
        "kind": ref.get("kind") or "unknown",
        "workbook": ref.get("workbook"),
        "sheet": ref.get("sheet"),
        "range": ref.get("range"),
        "block_id": ref.get("block_id"),
        "region_id": ref.get("region_id"),
        "bounds": _normalize_bounds(ref.get("bounds")),
        "label": ref.get("label"),
    }
    if normalized["range"] is None:
        normalized["range"] = _range_label(normalized["bounds"])
    return normalized


def _capture_window(sheet: str | None, bounds: dict[str, int] | None) -> dict[str, Any]:
    expanded = _expand_bounds(bounds, row_margin=2, column_margin=1)
    return {
        "sheet": sheet,
        "range": _range_label(expanded),
        "bounds": expanded,
        "authority": "excel_render_capture",
        "coordinate_systems": ["cell_range", "grid_coordinate", "capture_bbox"],
    }


def _normalize_bounds(bounds: dict[str, Any] | None) -> dict[str, int] | None:
    if not bounds:
        return None
    min_row = bounds.get("min_row", bounds.get("start_row"))
    min_column = bounds.get("min_column", bounds.get("start_column"))
    max_row = bounds.get("max_row", bounds.get("end_row"))
    max_column = bounds.get("max_column", bounds.get("end_column"))
    if None in (min_row, min_column, max_row, max_column):
        return None
    return {
        "min_row": int(min_row),
        "min_column": int(min_column),
        "max_row": int(max_row),
        "max_column": int(max_column),
    }


def _expand_bounds(
    bounds: dict[str, int] | None,
    *,
    row_margin: int,
    column_margin: int,
) -> dict[str, int] | None:
    if bounds is None:
        return None
    return {
        "min_row": max(1, bounds["min_row"] - row_margin),
        "min_column": max(1, bounds["min_column"] - column_margin),
        "max_row": bounds["max_row"] + row_margin,
        "max_column": bounds["max_column"] + column_margin,
    }


def _union_bounds(
    left: dict[str, int] | None,
    right: dict[str, int] | None,
) -> dict[str, int] | None:
    if left is None:
        return right
    if right is None:
        return left
    return {
        "min_row": min(left["min_row"], right["min_row"]),
        "min_column": min(left["min_column"], right["min_column"]),
        "max_row": max(left["max_row"], right["max_row"]),
        "max_column": max(left["max_column"], right["max_column"]),
    }


def _range_label(bounds: dict[str, int] | None) -> str | None:
    if bounds is None:
        return None
    return (
        f"{get_column_letter(bounds['min_column'])}{bounds['min_row']}:"
        f"{get_column_letter(bounds['max_column'])}{bounds['max_row']}"
    )


def _target_id(
    kind: str | None,
    ref_id: str | None,
    sheet: str | None,
    bounds: dict[str, int] | None,
) -> str:
    if ref_id:
        return f"target_{_slug(ref_id)}"
    if sheet and bounds:
        return f"target_{_slug(sheet)}_{_slug(_range_label(bounds) or 'range')}"
    return "target_unknown"


def _summary(
    targets: list[dict[str, Any]],
    gate_checks: list[dict[str, Any]],
    recommended_first_batch_target_ids: list[str],
) -> dict[str, int]:
    return {
        "capture_target_count": len(targets),
        "high_priority_count": sum(1 for target in targets if target["priority"] == "high"),
        "medium_priority_count": sum(1 for target in targets if target["priority"] == "medium"),
        "low_priority_count": sum(1 for target in targets if target["priority"] == "low"),
        "sheet_count": len({target["sheet"] for target in targets if target.get("sheet")}),
        "pipeline_target_count": sum(1 for target in targets if target["related_pipeline_ids"]),
        "pivot_report_target_count": sum(
            1 for target in targets if "pivot_table_output" in target["reasons"]
        ),
        "unresolved_input_target_count": sum(
            1 for target in targets if "unresolved_input_region" in target["reasons"]
        ),
        "boundary_target_count": sum(
            1 for target in targets if target["related_boundary_gate_ids"]
        ),
        "image_hierarchy_target_count": sum(
            1 for target in targets if target["target_type"] == "image_hierarchy"
        ),
        "gate_check_count": len(gate_checks),
        "recommended_first_batch_count": len(recommended_first_batch_target_ids),
    }


def _recommended_first_batch_target_ids(
    targets: list[dict[str, Any]],
    *,
    limit: int = 12,
) -> list[str]:
    eligible = [
        target for target in targets if target["priority"] in {"high", "medium"}
    ]
    selected: list[str] = []
    selected_sheets: set[str] = set()
    for target in eligible:
        sheet = target.get("sheet")
        if not sheet or sheet in selected_sheets:
            continue
        selected.append(target["id"])
        selected_sheets.add(sheet)
        if len(selected) >= limit:
            return selected
    for target in eligible:
        if target["id"] in selected:
            continue
        selected.append(target["id"])
        if len(selected) >= limit:
            break
    return selected


def _add_unique(items: list[str], value: str | None) -> None:
    if value and value not in items:
        items.append(value)


def _slug(value: str) -> str:
    return "".join(ch if ch.isalnum() else "_" for ch in value).strip("_").lower()


def _read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Plan render-capture and visual/formula cross-validation targets."
    )
    parser.add_argument("block_candidates", type=Path)
    parser.add_argument("table_io_pipelines", type=Path)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()

    package = build_cross_validation_plan(
        args.block_candidates,
        args.table_io_pipelines,
    )
    text = json.dumps(package, ensure_ascii=False, indent=2)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(text + "\n", encoding="utf-8")
    else:
        print(text)


if __name__ == "__main__":
    main()
