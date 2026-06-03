from __future__ import annotations

import argparse
import html
import json
import re
from collections import Counter, defaultdict
from datetime import UTC, datetime
from pathlib import Path
from statistics import mean
from typing import Any

from openpyxl.utils import get_column_letter, range_boundaries

from google_sheets_live_manifest import render_live_manifest_html


SCHEMA_VERSION = "0.1"
SECTION_HEADING_RE = re.compile(r"^[A-Z]\.\s+")


def build_live_block_candidates(
    *,
    live_manifest_path: Path,
    top_left_sample_path: Path,
    live_view_formula_profile_path: Path,
    parser_window_smoke_path: Path | None = None,
    max_blank_gap: int = 0,
    sample_limit: int = 8,
) -> dict[str, Any]:
    live_manifest_path = live_manifest_path.expanduser().resolve()
    top_left_sample_path = top_left_sample_path.expanduser().resolve()
    live_view_formula_profile_path = live_view_formula_profile_path.expanduser().resolve()
    parser_window_smoke_path = (
        parser_window_smoke_path.expanduser().resolve()
        if parser_window_smoke_path
        else None
    )
    manifest = _read_json(live_manifest_path)
    top_left_sample = _read_json(top_left_sample_path)
    formula_profile = _read_json(live_view_formula_profile_path)
    parser_window_smoke = (
        _read_json(parser_window_smoke_path)
        if parser_window_smoke_path and parser_window_smoke_path.exists()
        else None
    )
    sample_tabs = {tab["title"]: tab for tab in top_left_sample.get("tabs", [])}
    formula_by_sheet = _formula_observations_by_sheet(formula_profile)
    dependency_edges_by_sheet = _dependency_edges_by_sheet(formula_profile)
    view_state_by_sheet = {
        item["sheet"]: item for item in formula_profile.get("view_state_surfaces", [])
    }

    sheets = []
    for sheet in manifest["workbook"]["sheets"]:
        sheets.append(
            _sheet_candidates(
                manifest_sheet=sheet,
                sample_tab=sample_tabs.get(sheet["name"], {}),
                view_state=view_state_by_sheet.get(sheet["name"], {}),
                formula_observations=formula_by_sheet.get(sheet["name"], []),
                dependency_edges=dependency_edges_by_sheet.get(sheet["name"], []),
                parser_window_smoke=parser_window_smoke,
                max_blank_gap=max_blank_gap,
                sample_limit=sample_limit,
            )
        )
    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at": _utc_now(),
        "source": {
            "spreadsheet_id": manifest["source"]["spreadsheet_id"],
            "spreadsheet_url": manifest["source"].get("spreadsheet_url"),
            "title": manifest["source"]["title"],
            "source_artifacts": {
                "live_manifest": str(live_manifest_path),
                "top_left_sample": str(top_left_sample_path),
                "live_view_formula_profile": str(live_view_formula_profile_path),
                **(
                    {"parser_window_smoke": str(parser_window_smoke_path)}
                    if parser_window_smoke_path
                    else {}
                ),
            },
        },
        "authority": {
            "source_document": "live_google_sheet",
            "input_authority": "existing_profile_window_artifacts",
            "expanded_live_reads_performed": False,
            "profile_window": manifest["limits"]["profile_range"],
            "parser_window_contract_status": (
                "verified_for_current_policy_limits"
                if _parser_window_ops_verified(parser_window_smoke)
                else "not_verified"
            ),
            "formula_result_authority": "not_established",
            "candidate_status": "candidate_not_accepted_graph_claim",
        },
        "sheets": sheets,
        "summary": _summary(sheets),
        "parser_observations": _parser_observations(sheets, parser_window_smoke),
    }


def write_live_block_candidates_package(
    *,
    out_dir: Path,
    access_preflight_path: Path,
    live_manifest_path: Path,
    live_view_formula_profile_path: Path,
    live_block_candidates: dict[str, Any],
) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    block_candidates_path = out_dir / "live-block-candidates.json"
    block_candidates_path.write_text(
        json.dumps(live_block_candidates, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    access_preflight = _read_json(access_preflight_path)
    manifest = _read_json(live_manifest_path)
    view_formula_profile = _read_json(live_view_formula_profile_path)
    (out_dir / "index.html").write_text(
        render_live_manifest_html(
            access_preflight=access_preflight,
            manifest=manifest,
            live_view_formula_profile=view_formula_profile,
            live_block_candidates=live_block_candidates,
        ),
        encoding="utf-8",
    )


def _sheet_candidates(
    *,
    manifest_sheet: dict[str, Any],
    sample_tab: dict[str, Any],
    view_state: dict[str, Any],
    formula_observations: list[dict[str, Any]],
    dependency_edges: list[dict[str, Any]],
    parser_window_smoke: dict[str, Any] | None,
    max_blank_gap: int,
    sample_limit: int,
) -> dict[str, Any]:
    display_rows = sample_tab.get("display_rows", [])
    row_groups = _row_groups(display_rows, max_blank_gap=max_blank_gap)
    blocks = []
    regions = []
    relations = []
    for index, rows in enumerate(row_groups, start=1):
        block = _block_from_rows(
            manifest_sheet["name"],
            index,
            rows,
            sample_limit=sample_limit,
        )
        blocks.append(block)
        regions.append(_region_from_block(block))
    if not row_groups and (formula_observations or manifest_sheet["state"] == "hidden"):
        block = _support_surface_block(
            manifest_sheet,
            formula_observations,
            view_state,
        )
        blocks.append(block)
        regions.append(_region_from_block(block))
    formula_block = _formula_block(manifest_sheet, formula_observations)
    if formula_block:
        blocks.append(formula_block)
        regions.append(_region_from_block(formula_block))
    object_block = _object_block(manifest_sheet)
    if object_block:
        blocks.append(object_block)
        regions.append(_region_from_block(object_block))

    relations.extend(_section_relations(blocks, manifest_sheet["name"]))
    relations.extend(_dependency_relations(manifest_sheet["name"], dependency_edges))
    read_candidates = _read_candidates(
        manifest_sheet,
        display_rows,
        formula_observations,
        parser_window_smoke,
    )
    return {
        "name": manifest_sheet["name"],
        "sheet_id": manifest_sheet["sheet_id"],
        "index": manifest_sheet["index"],
        "state": manifest_sheet["state"],
        "dimensions": manifest_sheet["dimensions"],
        "profile_window": manifest_sheet["profile_window"],
        "role_hints": manifest_sheet["role_hints"],
        "view_state_status": view_state.get(
            "diagnostic_status",
            "no_profile_window_view_state_risk",
        ),
        "blocks": blocks,
        "cell_regions": regions,
        "relations": relations,
        "read_candidates": read_candidates,
        "summary": _sheet_summary(blocks, regions, relations, read_candidates),
        "parser_observations": _sheet_observations(
            manifest_sheet,
            blocks,
            formula_observations,
            read_candidates,
        ),
    }


def _row_groups(
    display_rows: list[list[Any]],
    *,
    max_blank_gap: int,
) -> list[list[dict[str, Any]]]:
    groups: list[list[dict[str, Any]]] = []
    current: list[dict[str, Any]] = []
    blank_gap = 0
    for row_index, row in enumerate(display_rows, start=1):
        cells = [
            {
                "row": row_index,
                "column": column_index,
                "value": value,
            }
            for column_index, value in enumerate(row, start=1)
            if value not in ("", None)
        ]
        if cells:
            if blank_gap > max_blank_gap and current:
                groups.append(current)
                current = []
            current.append({"row": row_index, "cells": cells})
            blank_gap = 0
        else:
            blank_gap += 1
    if current:
        groups.append(current)
    return groups


def _block_from_rows(
    sheet_name: str,
    index: int,
    rows: list[dict[str, Any]],
    *,
    sample_limit: int,
) -> dict[str, Any]:
    bounds = _bounds_for_rows(rows)
    metrics = _row_metrics(rows)
    subtype = _block_subtype(rows, metrics)
    label = _block_label(rows, subtype)
    confidence = _block_confidence(subtype, metrics)
    return {
        "id": f"block_{_slug(sheet_name)}_{index:03d}",
        "type": _block_type(subtype),
        "subtype": subtype,
        "sheet": sheet_name,
        "label": label,
        "bounds": bounds,
        "metrics": metrics,
        "preview": _preview(rows, sample_limit=sample_limit),
        "evidence": ["top_left_sample.display_rows"],
        "confidence": confidence,
        "status": "candidate",
    }


def _support_surface_block(
    manifest_sheet: dict[str, Any],
    formula_observations: list[dict[str, Any]],
    view_state: dict[str, Any],
) -> dict[str, Any]:
    sheet_name = manifest_sheet["name"]
    return {
        "id": f"block_{_slug(sheet_name)}_support_surface",
        "type": "support_surface",
        "subtype": (
            "hidden_formula_support_surface"
            if formula_observations
            else "hidden_sheet_surface"
        ),
        "sheet": sheet_name,
        "label": sheet_name,
        "bounds": _profile_bounds(manifest_sheet),
        "metrics": {
            "formula_observation_count": len(formula_observations),
            "hidden_sheet": manifest_sheet["state"] == "hidden",
        },
        "preview": [
            f"{sheet_name}: hidden/support surface",
            f"view_state={view_state.get('diagnostic_status', 'unknown')}",
        ],
        "evidence": ["live_manifest.workbook.sheets", "live_view_formula_profile.view_state_surfaces"],
        "confidence": 0.68 if formula_observations else 0.5,
        "status": "candidate",
    }


def _formula_block(
    manifest_sheet: dict[str, Any],
    formula_observations: list[dict[str, Any]],
) -> dict[str, Any] | None:
    if not formula_observations:
        return None
    rows = [item["row"] for item in formula_observations if item.get("row")]
    columns = [item["column"] for item in formula_observations if item.get("column")]
    if not rows or not columns:
        bounds = _profile_bounds(manifest_sheet)
    else:
        bounds = _bounds(min(rows), max(rows), min(columns), max(columns))
    classifications = sorted(
        {
            classification
            for item in formula_observations
            for classification in item.get("classifications", [])
        }
    )
    return {
        "id": f"block_{_slug(manifest_sheet['name'])}_formula_surface",
        "type": "formula_region_candidate",
        "subtype": "formula_dependency_surface",
        "sheet": manifest_sheet["name"],
        "label": f"{manifest_sheet['name']} formula surface",
        "bounds": bounds,
        "metrics": {
            "formula_observation_count": len(formula_observations),
            "classification_count": len(classifications),
        },
        "preview": [
            f"{item['cell']}: {item['formula'][:120]}"
            for item in formula_observations[:6]
        ],
        "evidence": ["live_view_formula_profile.formula_observations"],
        "confidence": 0.78,
        "status": "candidate",
    }


def _object_block(manifest_sheet: dict[str, Any]) -> dict[str, Any] | None:
    objects = manifest_sheet["object_counts"]
    if not any(objects.get(key, 0) for key in ("charts", "banded_ranges", "merges_in_profile_window")):
        return None
    return {
        "id": f"block_{_slug(manifest_sheet['name'])}_object_surface",
        "type": "object_surface",
        "subtype": "sheet_object_summary_surface",
        "sheet": manifest_sheet["name"],
        "label": f"{manifest_sheet['name']} object surface",
        "bounds": _profile_bounds(manifest_sheet),
        "metrics": objects,
        "preview": [
            f"charts={objects.get('charts', 0)}",
            f"banded_ranges={objects.get('banded_ranges', 0)}",
            f"merges={objects.get('merges_in_profile_window', 0)}",
        ],
        "evidence": ["live_manifest.workbook.sheets.object_counts"],
        "confidence": 0.58,
        "status": "candidate",
    }


def _region_from_block(block: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": f"region_{block['id'].removeprefix('block_')}",
        "type": "cell_region",
        "subtype": block["subtype"],
        "sheet": block["sheet"],
        "parent_block_id": block["id"],
        "label": block["label"],
        "bounds": block["bounds"],
        "metrics": block["metrics"],
        "preview": block["preview"][:4],
        "evidence": block["evidence"],
        "confidence": block["confidence"],
        "status": "candidate",
    }


def _section_relations(blocks: list[dict[str, Any]], sheet_name: str) -> list[dict[str, Any]]:
    relations = []
    current_section: dict[str, Any] | None = None
    for block in blocks:
        if block["type"] == "section_heading":
            current_section = block
            continue
        if current_section and block["bounds"]["start_row"] > current_section["bounds"]["start_row"]:
            relations.append(
                {
                    "id": f"rel_{current_section['id']}_contains_{block['id']}",
                    "type": "section_contains_block_candidate",
                    "sheet": sheet_name,
                    "from": current_section["id"],
                    "to": block["id"],
                    "evidence": ["top_left_sample.display_rows.ordering"],
                    "confidence": 0.62,
                    "status": "candidate",
                }
            )
    return relations


def _dependency_relations(sheet_name: str, dependency_edges: list[dict[str, Any]]) -> list[dict[str, Any]]:
    relations = []
    for edge in dependency_edges:
        relations.append(
            {
                "id": f"rel_{_slug(edge['id'])}",
                "type": "formula_dependency_candidate",
                "sheet": sheet_name,
                "from": f"block_{_slug(sheet_name)}_formula_surface",
                "to": (
                    f"sheet:{edge['target_sheet']}"
                    if edge.get("target_sheet")
                    else edge["target_kind"]
                ),
                "evidence": [edge["id"]],
                "confidence": 0.75 if edge["target_status"] == "known_sheet" else 0.45,
                "status": edge["target_status"],
            }
        )
    return relations


def _read_candidates(
    manifest_sheet: dict[str, Any],
    display_rows: list[list[Any]],
    formula_observations: list[dict[str, Any]],
    parser_window_smoke: dict[str, Any] | None,
) -> list[dict[str, Any]]:
    candidates = []
    profile_end = _profile_end_row(manifest_sheet)
    row_count = manifest_sheet["dimensions"].get("row_count", 0)
    contract_status = (
        "verified_for_current_policy_limits"
        if _parser_window_ops_verified(parser_window_smoke)
        else "blocked_missing_broker_contract"
    )
    if row_count > profile_end and (display_rows or formula_observations):
        end_row = min(row_count, profile_end + 80)
        candidates.append(
            {
                "id": f"read_{_slug(manifest_sheet['name'])}_next_window",
                "operation": "inspect.values_window",
                "range": f"'{manifest_sheet['name']}'!A{profile_end + 1}:Z{end_row}",
                "reason": "sheet extends beyond current profile window",
                "status": contract_status,
            }
        )
        if formula_observations:
            candidates.append(
                {
                    "id": f"read_{_slug(manifest_sheet['name'])}_formula_next_window",
                    "operation": "inspect.formula_window",
                    "range": f"'{manifest_sheet['name']}'!A{profile_end + 1}:Z{end_row}",
                    "reason": "formula surface extends beyond current profile window",
                    "status": contract_status,
                }
            )
    if not display_rows and formula_observations:
        candidates.append(
            {
                "id": f"read_{_slug(manifest_sheet['name'])}_profile_values",
                "operation": "inspect.values_window",
                "range": f"'{manifest_sheet['name']}'!A1:Z80",
                "reason": "sheet has formula text but no display rows in current sample",
                "status": contract_status,
            }
        )
        candidates.append(
            {
                "id": f"read_{_slug(manifest_sheet['name'])}_profile_formulas",
                "operation": "inspect.formula_window",
                "range": f"'{manifest_sheet['name']}'!A1:Z80",
                "reason": "confirm formula-bearing profile window through broker-backed formula read",
                "status": contract_status,
            }
        )
    return candidates


def _bounds_for_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    row_numbers = [row["row"] for row in rows]
    columns = [cell["column"] for row in rows for cell in row["cells"]]
    return _bounds(min(row_numbers), max(row_numbers), min(columns), max(columns))


def _profile_bounds(manifest_sheet: dict[str, Any]) -> dict[str, Any]:
    profile_range = manifest_sheet["profile_window"]["range"]
    min_col, min_row, max_col, max_row = range_boundaries(profile_range)
    return _bounds(min_row, max_row, min_col, max_col)


def _bounds(start_row: int, end_row: int, start_column: int, end_column: int) -> dict[str, Any]:
    start_cell = f"{get_column_letter(start_column)}{start_row}"
    end_cell = f"{get_column_letter(end_column)}{end_row}"
    return {
        "start_row": start_row,
        "end_row": end_row,
        "start_column": start_column,
        "end_column": end_column,
        "a1_range": f"{start_cell}:{end_cell}" if start_cell != end_cell else start_cell,
    }


def _row_metrics(rows: list[dict[str, Any]]) -> dict[str, Any]:
    non_empty_by_row = [len(row["cells"]) for row in rows]
    text_count = 0
    numeric_count = 0
    date_like_count = 0
    for row in rows:
        for cell in row["cells"]:
            value = str(cell["value"])
            if _is_number_like(value):
                numeric_count += 1
            elif _is_date_like(value):
                date_like_count += 1
            else:
                text_count += 1
    return {
        "row_count": len(rows),
        "non_empty_cell_count": sum(non_empty_by_row),
        "max_non_empty_cells_per_row": max(non_empty_by_row),
        "avg_non_empty_cells_per_row": round(mean(non_empty_by_row), 2),
        "text_cell_count": text_count,
        "number_like_cell_count": numeric_count,
        "date_like_cell_count": date_like_count,
    }


def _block_subtype(rows: list[dict[str, Any]], metrics: dict[str, Any]) -> str:
    preview_text = " ".join(str(cell["value"]) for row in rows for cell in row["cells"][:3])
    first_text = str(rows[0]["cells"][0]["value"]) if rows and rows[0]["cells"] else ""
    if SECTION_HEADING_RE.match(first_text):
        return "section_heading"
    if "Comment" in preview_text or preview_text.strip().startswith("-"):
        return "commentary_text_block"
    if "표" in preview_text or "Data from" in preview_text:
        return "labeled_table_candidate"
    if metrics["number_like_cell_count"] >= 2 or metrics["max_non_empty_cells_per_row"] >= 4:
        return "table_like_region"
    return "text_block"


def _block_type(subtype: str) -> str:
    if subtype == "section_heading":
        return "section_heading"
    if "table" in subtype or subtype == "table_like_region":
        return "table_candidate"
    return "text_block"


def _block_label(rows: list[dict[str, Any]], subtype: str) -> str | None:
    if not rows or not rows[0]["cells"]:
        return None
    if subtype in {"section_heading", "labeled_table_candidate"}:
        return str(rows[0]["cells"][0]["value"])
    return str(rows[0]["cells"][0]["value"])[:80]


def _block_confidence(subtype: str, metrics: dict[str, Any]) -> float:
    if subtype == "section_heading":
        return 0.82
    if subtype == "labeled_table_candidate":
        return 0.72
    if subtype == "table_like_region" and metrics["number_like_cell_count"] >= 4:
        return 0.7
    if subtype == "commentary_text_block":
        return 0.68
    return 0.55


def _preview(rows: list[dict[str, Any]], *, sample_limit: int) -> list[str]:
    preview = []
    for row in rows[:sample_limit]:
        values = [str(cell["value"]).replace("\n", " ").strip() for cell in row["cells"]]
        preview.append(f"R{row['row']}: " + " | ".join(values[:8]))
    return preview


def _formula_observations_by_sheet(profile: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    result: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in profile.get("formula_observations", []):
        result[item["sheet"]].append(item)
    return result


def _dependency_edges_by_sheet(profile: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    result: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in profile.get("dependency_edges", []):
        result[item["source_sheet"]].append(item)
    return result


def _sheet_summary(
    blocks: list[dict[str, Any]],
    regions: list[dict[str, Any]],
    relations: list[dict[str, Any]],
    read_candidates: list[dict[str, Any]],
) -> dict[str, int]:
    type_counts = Counter(block["type"] for block in blocks)
    return {
        "block_count": len(blocks),
        "cell_region_count": len(regions),
        "relation_count": len(relations),
        "read_candidate_count": len(read_candidates),
        "table_candidate_count": type_counts["table_candidate"],
        "text_block_count": type_counts["text_block"],
        "section_heading_count": type_counts["section_heading"],
        "formula_region_candidate_count": type_counts["formula_region_candidate"],
        "object_surface_count": type_counts["object_surface"],
        "support_surface_count": type_counts["support_surface"],
    }


def _sheet_observations(
    manifest_sheet: dict[str, Any],
    blocks: list[dict[str, Any]],
    formula_observations: list[dict[str, Any]],
    read_candidates: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    observations = []
    if manifest_sheet["state"] == "hidden":
        observations.append(
            {
                "level": "warning",
                "message": "Hidden sheet is treated as structural evidence, not visible review output.",
            }
        )
    if not blocks:
        observations.append(
            {
                "level": "warning",
                "message": "No display-derived block candidates were available in the current profile window.",
            }
        )
    if formula_observations:
        observations.append(
            {
                "level": "info",
                "message": "Formula observations were attached as candidate formula surfaces; formula results are not established.",
            }
        )
    if read_candidates:
        observations.append(
            {
                "level": "info",
                "message": "Additional bounded broker reads are proposed for later tuning.",
            }
        )
    return observations


def _summary(sheets: list[dict[str, Any]]) -> dict[str, int | str]:
    return {
        "sheet_count": len(sheets),
        "block_count": sum(sheet["summary"]["block_count"] for sheet in sheets),
        "cell_region_count": sum(sheet["summary"]["cell_region_count"] for sheet in sheets),
        "relation_count": sum(sheet["summary"]["relation_count"] for sheet in sheets),
        "read_candidate_count": sum(sheet["summary"]["read_candidate_count"] for sheet in sheets),
        "table_candidate_count": sum(sheet["summary"]["table_candidate_count"] for sheet in sheets),
        "text_block_count": sum(sheet["summary"]["text_block_count"] for sheet in sheets),
        "section_heading_count": sum(sheet["summary"]["section_heading_count"] for sheet in sheets),
        "formula_region_candidate_count": sum(sheet["summary"]["formula_region_candidate_count"] for sheet in sheets),
        "object_surface_count": sum(sheet["summary"]["object_surface_count"] for sheet in sheets),
        "support_surface_count": sum(sheet["summary"]["support_surface_count"] for sheet in sheets),
        "candidate_status": "generated_from_profile_windows",
    }


def _parser_observations(
    sheets: list[dict[str, Any]],
    parser_window_smoke: dict[str, Any] | None,
) -> list[dict[str, Any]]:
    observations = [
        {
            "level": "info",
            "message": "Block and region candidates are deterministic parser seeds, not accepted document graph claims.",
        }
    ]
    if _parser_window_ops_verified(parser_window_smoke):
        observations.append(
            {
                "level": "info",
                "message": "Bounded parser-window broker operations are verified for current policy limits.",
            }
        )
    else:
        observations.append(
            {
                "level": "warning",
                "message": "Bounded parser-window broker operations are not verified; read candidates are stop conditions.",
            }
        )
    hidden_count = sum(1 for sheet in sheets if sheet["state"] == "hidden")
    if hidden_count:
        observations.append(
            {
                "level": "warning",
                "message": f"{hidden_count} hidden sheets are structural evidence but not visible review surfaces.",
            }
        )
    return observations


def _profile_end_row(manifest_sheet: dict[str, Any]) -> int:
    profile_range = manifest_sheet["profile_window"]["range"]
    return range_boundaries(profile_range)[3]


def _parser_window_ops_verified(parser_window_smoke: dict[str, Any] | None) -> bool:
    if not parser_window_smoke:
        return False
    required = {
        "inspect.grid_window",
        "inspect.values_window",
        "inspect.formula_window",
    }
    passed = {
        item.get("operation")
        for item in parser_window_smoke.get("smoke_results", [])
        if item.get("result") == "passed"
    }
    return required <= passed


def _is_number_like(value: str) -> bool:
    text = value.strip().replace(",", "").replace("%", "")
    if not text:
        return False
    try:
        float(text)
    except ValueError:
        return False
    return True


def _is_date_like(value: str) -> bool:
    return bool(re.search(r"\d{4}\.\s*\d{1,2}\.\s*\d{1,2}", value))


def _slug(value: Any) -> str:
    text = str(value or "none")
    text = re.sub(r"[^A-Za-z0-9가-힣]+", "_", text).strip("_").lower()
    return text or "none"


def _read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _utc_now() -> str:
    return datetime.now(UTC).isoformat().replace("+00:00", "Z")


def render_live_block_candidates_section(block_candidates: dict[str, Any]) -> str:
    metrics = "".join(
        f'<div class="metric"><div class="label">{_esc(key)}</div>'
        f'<div class="value">{_esc(value)}</div></div>'
        for key, value in block_candidates["summary"].items()
    )
    sheet_rows = "".join(
        "<tr>"
        f"<td>{_esc(sheet['index'])}</td>"
        f"<td>{_esc(sheet['name'])}</td>"
        f"<td>{_esc(sheet['state'])}</td>"
        f"<td>{_esc(sheet['view_state_status'])}</td>"
        f"<td>{_esc(sheet['summary']['block_count'])}</td>"
        f"<td>{_esc(sheet['summary']['table_candidate_count'])}</td>"
        f"<td>{_esc(sheet['summary']['formula_region_candidate_count'])}</td>"
        f"<td>{_esc(sheet['summary']['read_candidate_count'])}</td>"
        "</tr>"
        for sheet in block_candidates["sheets"]
    )
    block_rows = "".join(
        "<tr>"
        f"<td>{_esc(sheet['name'])}</td>"
        f"<td>{_esc(block['type'])}</td>"
        f"<td>{_esc(block['subtype'])}</td>"
        f"<td>{_esc(block.get('label'))}</td>"
        f"<td>{_esc(block['bounds']['a1_range'])}</td>"
        f"<td>{_esc(block['confidence'])}</td>"
        f"<td>{_esc(' / '.join(block['preview'][:2]))}</td>"
        "</tr>"
        for sheet in block_candidates["sheets"]
        for block in sheet["blocks"][:8]
    )
    read_rows = "".join(
        "<tr>"
        f"<td>{_esc(sheet['name'])}</td>"
        f"<td>{_esc(item['operation'])}</td>"
        f"<td>{_esc(item['range'])}</td>"
        f"<td>{_esc(item['status'])}</td>"
        f"<td>{_esc(item['reason'])}</td>"
        "</tr>"
        for sheet in block_candidates["sheets"]
        for item in sheet["read_candidates"][:4]
    )
    if not read_rows:
        read_rows = '<tr><td colspan="5">No read candidates.</td></tr>'
    observations = "".join(
        "<tr>"
        f"<td>{_esc(item['level'])}</td>"
        f"<td>{_esc(item['message'])}</td>"
        "</tr>"
        for item in block_candidates["parser_observations"]
    )
    return f"""
  <h2>Live Block / Region Candidates</h2>
  <section class="grid">{metrics}</section>
  <h2>Sheet Candidate Summary</h2>
  <section class="panel"><table><thead><tr><th>#</th><th>Sheet</th><th>State</th><th>View State</th><th>Blocks</th><th>Tables</th><th>Formula Surfaces</th><th>Read Candidates</th></tr></thead><tbody>{sheet_rows}</tbody></table></section>
  <h2>Block Candidate Samples</h2>
  <section class="panel"><table><thead><tr><th>Sheet</th><th>Type</th><th>Subtype</th><th>Label</th><th>Range</th><th>Confidence</th><th>Preview</th></tr></thead><tbody>{block_rows}</tbody></table></section>
  <h2>Bounded Read Candidates</h2>
  <section class="panel"><table><thead><tr><th>Sheet</th><th>Operation</th><th>Range</th><th>Status</th><th>Reason</th></tr></thead><tbody>{read_rows}</tbody></table></section>
  <h2>Block Candidate Observations</h2>
  <section class="panel"><table><thead><tr><th>Level</th><th>Message</th></tr></thead><tbody>{observations}</tbody></table></section>
"""


def _esc(value: Any) -> str:
    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False)
    return html.escape(str(value))


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build Google Sheets live block and region candidates from profile artifacts."
    )
    parser.add_argument("--out-dir", type=Path, required=True)
    parser.add_argument("--access-preflight", type=Path, required=True)
    parser.add_argument("--live-manifest", type=Path, required=True)
    parser.add_argument("--top-left-sample", type=Path, required=True)
    parser.add_argument("--live-view-formula-profile", type=Path, required=True)
    parser.add_argument("--parser-window-smoke", type=Path)
    parser.add_argument("--max-blank-gap", type=int, default=0)
    args = parser.parse_args()

    block_candidates = build_live_block_candidates(
        live_manifest_path=args.live_manifest,
        top_left_sample_path=args.top_left_sample,
        live_view_formula_profile_path=args.live_view_formula_profile,
        parser_window_smoke_path=args.parser_window_smoke,
        max_blank_gap=args.max_blank_gap,
    )
    write_live_block_candidates_package(
        out_dir=args.out_dir,
        access_preflight_path=args.access_preflight,
        live_manifest_path=args.live_manifest,
        live_view_formula_profile_path=args.live_view_formula_profile,
        live_block_candidates=block_candidates,
    )


if __name__ == "__main__":
    main()
