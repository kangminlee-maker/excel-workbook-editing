from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl.utils import get_column_letter, range_boundaries

SCHEMA_VERSION = "0.1"
MAX_COLUMNS_PER_TILE = 18
VISIBLE_CONTEXT_ROWS = 80
EXPANDED_CONTEXT_ROWS = 80


def build_recapture_candidate_plan(
    render_captures_path: Path,
    capture_quality_path: Path,
) -> dict[str, Any]:
    render_captures_path = render_captures_path.expanduser().resolve()
    capture_quality_path = capture_quality_path.expanduser().resolve()
    render_captures = _read_json(render_captures_path)
    capture_quality = _read_json(capture_quality_path)
    captures_by_id = {
        capture.get("id"): capture
        for capture in render_captures.get("captures", [])
    }
    groups = []
    targets = []
    for quality_result in capture_quality.get("quality_results", []):
        if quality_result.get("status") == "usable":
            continue
        capture = captures_by_id.get(quality_result.get("capture_id"))
        if not capture:
            continue
        group_targets = _candidate_targets(capture, quality_result)
        if not group_targets:
            continue
        groups.append(
            {
                "source_capture_id": capture.get("id"),
                "source_quality_result_id": quality_result.get("id"),
                "source_quality_status": quality_result.get("status"),
                "sheet": capture.get("sheet"),
                "source_range": quality_result.get("capture_window_range"),
                "recommendations": quality_result.get("recommendations", []),
                "candidate_target_ids": [target["id"] for target in group_targets],
            }
        )
        targets.extend(group_targets)
    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at": _utc_now(),
        "source_artifacts": {
            "render_captures": str(render_captures_path),
            "capture_quality": str(capture_quality_path),
        },
        "method": {
            "name": "recapture_candidate_generation",
            "max_columns_per_tile": MAX_COLUMNS_PER_TILE,
            "visible_context_rows": VISIBLE_CONTEXT_ROWS,
            "expanded_context_rows": EXPANDED_CONTEXT_ROWS,
            "authority": "candidate_generation_not_final_selection",
        },
        "candidate_groups": groups,
        "capture_targets": targets,
        "recommended_first_batch_target_ids": [target["id"] for target in targets],
        "summary": _summary(groups, targets),
        "parser_observations": [
            {
                "level": "info",
                "message": "Recapture candidates are alternatives for experiment and review. They are not accepted final recapture plans.",
            }
        ],
    }


def _candidate_targets(
    capture: dict[str, Any],
    quality_result: dict[str, Any],
) -> list[dict[str, Any]]:
    base_bounds = _bounds_from_range(quality_result.get("capture_window_range"))
    if base_bounds is None:
        base_bounds = _bounds_from_capture_window(capture.get("capture_window") or {})
    if base_bounds is None:
        return []
    recommendations = set(quality_result.get("recommendations", []))
    target_ref = capture.get("target_ref") or {}
    ref_bounds = _normalize_bounds(target_ref.get("bounds"))
    targets = []

    if quality_result.get("status") == "recapture_required":
        targets.append(
            _target(
                "same_window_control",
                capture,
                quality_result,
                base_bounds,
                "Control candidate: recapture the original window to distinguish transient Excel capture problems from range-selection problems.",
            )
        )

    if "recapture_with_expanded_window_or_zoom" in recommendations:
        expanded = _expanded_context_bounds(base_bounds)
        targets.append(
            _target(
                "expanded_row_context",
                capture,
                quality_result,
                expanded,
                "Expand rows downward from the original window to test whether more visible row context fixes a too-thin capture.",
            )
        )

    if "recapture_with_visible_row_context" in recommendations:
        visible = _visible_context_bounds(base_bounds, ref_bounds)
        if "recapture_with_tiling" in recommendations:
            for index, tile in enumerate(_column_tiles(visible), 1):
                targets.append(
                    _target(
                        "visible_row_context_tile",
                        capture,
                        quality_result,
                        tile,
                        "Shift to likely visible data rows and tile columns so the candidate tests row visibility and width separately.",
                        tile_index=index,
                        tile_count=len(_column_tiles(visible)),
                    )
                )
        else:
            targets.append(
                _target(
                    "visible_row_context",
                    capture,
                    quality_result,
                    visible,
                    "Shift to likely visible data rows below the hidden or collapsed header band.",
                )
            )

    if (
        "recapture_with_tiling" in recommendations
        and "recapture_with_visible_row_context" not in recommendations
    ):
        tiles = _column_tiles(base_bounds)
        for index, tile in enumerate(tiles, 1):
            targets.append(
                _target(
                    "column_tile",
                    capture,
                    quality_result,
                    tile,
                    "Tile a wide range into narrower column windows for visual review.",
                    tile_index=index,
                    tile_count=len(tiles),
                )
            )
    return targets


def _target(
    strategy: str,
    capture: dict[str, Any],
    quality_result: dict[str, Any],
    bounds: dict[str, int],
    rationale: str,
    *,
    tile_index: int = 1,
    tile_count: int = 1,
) -> dict[str, Any]:
    source_capture_id = capture.get("id") or "capture"
    target_id = f"candidate_{_slug(source_capture_id)}_{strategy}_{tile_index:02d}"
    quality_status = quality_result.get("status")
    priority = "high" if quality_status == "recapture_required" else "medium"
    score = 95 if priority == "high" else 70
    range_text = _range_label(bounds)
    gate_id = f"gate_{target_id}_candidate_capture_quality"
    evidence_refs = [
        item
        for item in [source_capture_id, quality_result.get("id")]
        if item
    ]
    return {
        "id": target_id,
        "type": "visual_formula_validation_target",
        "target_type": "recapture_candidate",
        "status": "candidate",
        "priority": priority,
        "score": score,
        "sheet": capture.get("sheet"),
        "range": range_text,
        "bounds": bounds,
        "capture_window": {
            "sheet": capture.get("sheet"),
            "range": range_text,
            "bounds": bounds,
            "authority": "excel_render_capture",
            "coordinate_systems": ["cell_range", "grid_coordinate", "capture_bbox"],
        },
        "target_ref": capture.get("target_ref") or {},
        "source_capture_id": source_capture_id,
        "source_quality_result_id": quality_result.get("id"),
        "source_quality_status": quality_status,
        "candidate_strategy": strategy,
        "candidate_rationale": rationale,
        "tile_index": tile_index,
        "tile_count": tile_count,
        "related_pipeline_ids": [],
        "related_block_ids": [],
        "related_region_ids": [],
        "related_boundary_gate_ids": [],
        "reasons": [
            quality_status or "quality_flagged",
            *quality_result.get("recommendations", []),
        ],
        "gate_checks": [
            {
                "id": gate_id,
                "type": "visual_formula_gate_check",
                "target_id": target_id,
                "gate_type": "recapture_candidate_quality_confirmation",
                "status": "pending_capture",
                "deterministic_inputs": evidence_refs,
                "pass_conditions": [
                    "Candidate capture quality improves over the source capture.",
                    "Candidate range still covers the visual evidence needed for the source gate or a clearly documented tile of it.",
                ],
                "failure_signals": [
                    "Candidate capture remains too thin, clipped, blank, or unreadable.",
                    "Candidate range no longer represents the source gate's intended workbook evidence.",
                ],
            }
        ],
        "review_questions": [
            "이 candidate capture가 원래 capture보다 downstream visual gate에 더 적합한가요?"
        ],
        "evidence_refs": evidence_refs,
    }


def _visible_context_bounds(
    base_bounds: dict[str, int],
    ref_bounds: dict[str, int] | None,
) -> dict[str, int]:
    if ref_bounds:
        start_row = ref_bounds["max_row"] + 1
    else:
        start_row = base_bounds["max_row"] + 1
    return {
        "min_row": start_row,
        "min_column": base_bounds["min_column"],
        "max_row": start_row + VISIBLE_CONTEXT_ROWS - 1,
        "max_column": base_bounds["max_column"],
    }


def _expanded_context_bounds(bounds: dict[str, int]) -> dict[str, int]:
    return {
        "min_row": bounds["min_row"],
        "min_column": bounds["min_column"],
        "max_row": bounds["min_row"] + EXPANDED_CONTEXT_ROWS - 1,
        "max_column": bounds["max_column"],
    }


def _column_tiles(bounds: dict[str, int]) -> list[dict[str, int]]:
    tiles = []
    start = bounds["min_column"]
    while start <= bounds["max_column"]:
        end = min(start + MAX_COLUMNS_PER_TILE - 1, bounds["max_column"])
        tiles.append(
            {
                "min_row": bounds["min_row"],
                "min_column": start,
                "max_row": bounds["max_row"],
                "max_column": end,
            }
        )
        start = end + 1
    return tiles


def _bounds_from_capture_window(capture_window: dict[str, Any]) -> dict[str, int] | None:
    bounds = _normalize_bounds(capture_window.get("bounds"))
    if bounds:
        return bounds
    return _bounds_from_range(capture_window.get("range"))


def _bounds_from_range(range_text: str | None) -> dict[str, int] | None:
    if not range_text:
        return None
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_text)
    except ValueError:
        return None
    return {
        "min_row": min_row,
        "min_column": min_col,
        "max_row": max_row,
        "max_column": max_col,
    }


def _normalize_bounds(bounds: dict[str, Any] | None) -> dict[str, int] | None:
    if not bounds:
        return None
    min_row = bounds.get("min_row", bounds.get("start_row"))
    min_column = bounds.get("min_column", bounds.get("start_column"))
    max_row = bounds.get("max_row", bounds.get("end_row"))
    max_column = bounds.get("max_column", bounds.get("end_column"))
    if None in (min_row, min_column, max_row, max_column):
        return None
    return {
        "min_row": int(min_row),
        "min_column": int(min_column),
        "max_row": int(max_row),
        "max_column": int(max_column),
    }


def _range_label(bounds: dict[str, int]) -> str:
    return (
        f"{get_column_letter(bounds['min_column'])}{bounds['min_row']}:"
        f"{get_column_letter(bounds['max_column'])}{bounds['max_row']}"
    )


def _summary(groups: list[dict[str, Any]], targets: list[dict[str, Any]]) -> dict[str, int]:
    return {
        "source_capture_count": len(groups),
        "candidate_target_count": len(targets),
        "high_priority_count": sum(1 for target in targets if target["priority"] == "high"),
        "medium_priority_count": sum(1 for target in targets if target["priority"] == "medium"),
        "same_window_control_count": _strategy_count(targets, "same_window_control"),
        "expanded_row_context_count": _strategy_count(targets, "expanded_row_context"),
        "visible_row_context_count": _strategy_count(targets, "visible_row_context"),
        "visible_row_context_tile_count": _strategy_count(targets, "visible_row_context_tile"),
        "column_tile_count": _strategy_count(targets, "column_tile"),
        "sheet_count": len({target.get("sheet") for target in targets}),
        "gate_check_count": sum(len(target.get("gate_checks", [])) for target in targets),
        "recommended_first_batch_count": len(targets),
    }


def _strategy_count(targets: list[dict[str, Any]], strategy: str) -> int:
    return sum(1 for target in targets if target.get("candidate_strategy") == strategy)


def _slug(value: str) -> str:
    return "".join(ch if ch.isalnum() else "_" for ch in value).strip("_").lower()


def _read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build recapture candidate targets from capture quality results."
    )
    parser.add_argument("render_captures", type=Path)
    parser.add_argument("capture_quality", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    package = build_recapture_candidate_plan(args.render_captures, args.capture_quality)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(package, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(package["summary"], ensure_ascii=False))


if __name__ == "__main__":
    main()
