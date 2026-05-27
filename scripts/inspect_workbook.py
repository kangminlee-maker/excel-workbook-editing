from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def inspect_workbook(workbook_path: Path, *, sample_limit: int = 20) -> dict[str, Any]:
    workbook_path = workbook_path.expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"missing workbook: {workbook_path}")

    wb = load_workbook(workbook_path, read_only=False, data_only=False)
    try:
        sheets = [_inspect_sheet(ws, sample_limit=sample_limit) for ws in wb.worksheets]
        return {
            "workbook": str(workbook_path),
            "sheet_count": len(wb.worksheets),
            "active_sheet": wb.active.title if wb.worksheets else None,
            "named_ranges": _defined_names(wb, sample_limit=sample_limit),
            "external_link_count": len(getattr(wb, "_external_links", []) or []),
            "sheets": sheets,
            "summary": {
                "formula_cell_count": sum(s["formula_cell_count"] for s in sheets),
                "merged_range_count": sum(s["merged_range_count"] for s in sheets),
                "table_count": sum(s["table_count"] for s in sheets),
                "data_validation_count": sum(s["data_validation_count"] for s in sheets),
                "conditional_formatting_count": sum(
                    s["conditional_formatting_count"] for s in sheets
                ),
                "hidden_sheet_count": sum(
                    1 for s in sheets if s["sheet_state"] != "visible"
                ),
            },
        }
    finally:
        wb.close()


def _inspect_sheet(ws: Any, *, sample_limit: int) -> dict[str, Any]:
    formula_samples: list[dict[str, str]] = []
    formula_count = 0

    for cell in ws._cells.values():
        value = cell.value
        if cell.data_type == "f" or (isinstance(value, str) and value.startswith("=")):
            formula_count += 1
            if len(formula_samples) < sample_limit:
                formula_samples.append({"cell": cell.coordinate, "formula": str(value)})

    tables = []
    for name, table_obj in ws.tables.items():
        ref = table_obj if isinstance(table_obj, str) else getattr(table_obj, "ref", None)
        tables.append({"name": str(name), "ref": str(ref or "")})

    merged_ranges = [str(rng) for rng in ws.merged_cells.ranges]

    return {
        "name": ws.title,
        "sheet_state": ws.sheet_state,
        "dimensions": ws.calculate_dimension(),
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
        "auto_filter_ref": ws.auto_filter.ref,
        "formula_cell_count": formula_count,
        "formula_samples": formula_samples,
        "merged_range_count": len(merged_ranges),
        "merged_range_samples": merged_ranges[:sample_limit],
        "table_count": len(tables),
        "tables": tables[:sample_limit],
        "data_validation_count": len(ws.data_validations.dataValidation),
        "conditional_formatting_count": len(ws.conditional_formatting),
        "chart_count": len(getattr(ws, "_charts", []) or []),
        "image_count": len(getattr(ws, "_images", []) or []),
    }


def _defined_names(wb: Any, *, sample_limit: int) -> list[dict[str, Any]]:
    names: list[dict[str, Any]] = []
    for name, defined_name in wb.defined_names.items():
        destinations = []
        try:
            destinations = [
                {"sheet": sheet, "range": coord}
                for sheet, coord in defined_name.destinations
            ]
        except Exception:
            destinations = []

        names.append(
            {
                "name": str(name),
                "attr_text": getattr(defined_name, "attr_text", None),
                "hidden": bool(getattr(defined_name, "hidden", False)),
                "destinations": destinations[:sample_limit],
            }
        )
    return names


def _to_markdown(report: dict[str, Any]) -> str:
    lines = [
        f"# Workbook Inspection: {Path(report['workbook']).name}",
        "",
        "## Summary",
        "",
        f"- Sheets: {report['sheet_count']}",
        f"- Named ranges: {len(report['named_ranges'])}",
        f"- Formula cells: {report['summary']['formula_cell_count']}",
        f"- Tables: {report['summary']['table_count']}",
        f"- Merged ranges: {report['summary']['merged_range_count']}",
        f"- Data validations: {report['summary']['data_validation_count']}",
        f"- Conditional formatting entries: {report['summary']['conditional_formatting_count']}",
        f"- Hidden sheets: {report['summary']['hidden_sheet_count']}",
        "",
        "## Sheets",
        "",
        "| Sheet | State | Dimensions | Formulas | Tables | Merged | Validations | CF |",
        "|---|---|---:|---:|---:|---:|---:|---:|",
    ]
    for sheet in report["sheets"]:
        lines.append(
            "| {name} | {sheet_state} | {dimensions} | {formula_cell_count} | "
            "{table_count} | {merged_range_count} | {data_validation_count} | "
            "{conditional_formatting_count} |".format(**sheet)
        )
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Summarize workbook structure before an agent edits or validates it."
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--sample-limit", type=int, default=20)
    parser.add_argument(
        "--format",
        choices=["json", "markdown"],
        default="json",
        help="Output format. Defaults to json for machine consumption.",
    )
    args = parser.parse_args()

    report = inspect_workbook(args.workbook, sample_limit=args.sample_limit)
    if args.format == "markdown":
        print(_to_markdown(report))
    else:
        print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
