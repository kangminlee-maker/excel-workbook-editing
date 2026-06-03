from __future__ import annotations

import argparse
import hashlib
import json
import time as monotonic_time
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SCHEMA_VERSION = "0.1"


def build_readonly_sample(
    workbook_path: Path,
    *,
    manifest_path: Path | None = None,
    include_structural_windows: bool = False,
    structural_window_padding: int = 2,
    max_structural_windows_per_sheet: int = 30,
    sheets: list[str] | None = None,
    row_windows: list[str] | None = None,
    default_max_rows: int = 20,
    max_columns: int = 50,
    preview_chars: int = 120,
) -> dict[str, Any]:
    workbook_path = workbook_path.expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"missing workbook: {workbook_path}")

    manifest = _read_json(manifest_path) if manifest_path else None
    started = monotonic_time.perf_counter()
    wb = load_workbook(
        workbook_path,
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    load_seconds = monotonic_time.perf_counter() - started
    try:
        selected_sheets = sheets or _manifest_sheet_names(manifest) or list(wb.sheetnames)
        windows_by_sheet = _windows_by_sheet(
            selected_sheets,
            row_windows or [],
            default_max_rows=default_max_rows,
        )
        if include_structural_windows:
            if manifest is None:
                raise ValueError("manifest_path is required for structural windows")
            _append_structural_windows(
                windows_by_sheet,
                manifest,
                selected_sheets=selected_sheets,
                padding=structural_window_padding,
                max_windows_per_sheet=max_structural_windows_per_sheet,
            )
        sheet_samples = []
        for sheet_name in selected_sheets:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"missing sheet: {sheet_name}")
            ws = wb[sheet_name]
            sheet_started = monotonic_time.perf_counter()
            windows = [
                _sample_window(
                    ws,
                    start_row=start,
                    end_row=end,
                    max_columns=max_columns,
                    preview_chars=preview_chars,
                )
                for start, end in windows_by_sheet[sheet_name]
            ]
            sheet_samples.append(
                {
                    "name": sheet_name,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "sample_seconds": round(
                        monotonic_time.perf_counter() - sheet_started,
                        6,
                    ),
                    "windows": windows,
                }
            )

        return {
            "schema_version": SCHEMA_VERSION,
            "generated_at": _utc_now(),
            "source": {
                "path": str(workbook_path),
                "file_name": workbook_path.name,
                "size_bytes": workbook_path.stat().st_size,
                "sha256": _sha256(workbook_path),
            },
            "engine": {
                "library": "openpyxl",
                "mode": "read_only",
                "data_only": False,
                "keep_links": False,
                "load_seconds": round(load_seconds, 6),
            },
            "limits": {
                "default_max_rows": default_max_rows,
                "max_columns": max_columns,
                "preview_chars": preview_chars,
            },
            "sheets": sheet_samples,
            "summary": _summary(sheet_samples),
            "parser_observations": [
                {
                    "level": "info",
                    "message": "Formula cells expose formula text, not recalculated formula results. Use the Excel engine for formula-dependent value validation.",
                },
                *(
                    [
                        {
                            "level": "info",
                            "message": "Sampling windows include manifest-derived pivot table and drawing anchor regions.",
                        }
                    ]
                    if include_structural_windows
                    else []
                ),
            ],
        }
    finally:
        wb.close()


def _windows_by_sheet(
    selected_sheets: list[str],
    row_windows: list[str],
    *,
    default_max_rows: int,
) -> dict[str, list[tuple[int, int]]]:
    windows = {sheet: [(1, default_max_rows)] for sheet in selected_sheets}
    if not row_windows:
        return windows

    windows = {sheet: [] for sheet in selected_sheets}
    for spec in row_windows:
        sheet_name, bounds = _parse_row_window(spec)
        if sheet_name not in windows:
            windows[sheet_name] = []
        windows[sheet_name].append(bounds)

    for sheet_name in selected_sheets:
        if not windows[sheet_name]:
            windows[sheet_name].append((1, default_max_rows))
    return windows


def _manifest_sheet_names(manifest: dict[str, Any] | None) -> list[str] | None:
    if manifest is None:
        return None
    return [sheet["name"] for sheet in manifest["workbook"]["sheets"]]


def _append_structural_windows(
    windows_by_sheet: dict[str, list[tuple[int, int]]],
    manifest: dict[str, Any],
    *,
    selected_sheets: list[str],
    padding: int,
    max_windows_per_sheet: int,
) -> None:
    manifest_sheets = {
        sheet["name"]: sheet for sheet in manifest["workbook"]["sheets"]
    }
    for sheet_name in selected_sheets:
        sheet = manifest_sheets.get(sheet_name)
        if sheet is None:
            continue
        bounds = sheet.get("dimension_bounds") or {}
        max_row = bounds.get("max_row")
        candidates = list(windows_by_sheet.get(sheet_name, []))
        for pivot in sheet.get("pivot_tables", []):
            location_bounds = pivot.get("location", {}).get("bounds")
            if location_bounds:
                candidates.append(
                    _padded_window(
                        location_bounds["min_row"],
                        location_bounds["max_row"],
                        padding=padding,
                        max_row=max_row,
                    )
                )
        for drawing in sheet.get("drawing_objects", []):
            start = drawing.get("from")
            end = drawing.get("to")
            if start and end:
                candidates.append(
                    _padded_window(
                        start["row"],
                        end["row"],
                        padding=padding,
                        max_row=max_row,
                    )
                )
        windows_by_sheet[sheet_name] = _merge_windows(
            candidates,
            max_windows=max_windows_per_sheet,
        )


def _padded_window(
    start_row: int,
    end_row: int,
    *,
    padding: int,
    max_row: int | None,
) -> tuple[int, int]:
    start = max(1, start_row - padding)
    end = end_row + padding
    if max_row is not None:
        end = min(end, max_row)
    return start, max(start, end)


def _merge_windows(
    windows: list[tuple[int, int]],
    *,
    max_windows: int,
) -> list[tuple[int, int]]:
    merged: list[tuple[int, int]] = []
    for start, end in sorted(set(windows)):
        if not merged or start > merged[-1][1] + 1:
            merged.append((start, end))
            continue
        previous_start, previous_end = merged[-1]
        merged[-1] = (previous_start, max(previous_end, end))
    return merged[:max_windows]


def _parse_row_window(spec: str) -> tuple[str, tuple[int, int]]:
    if ":" not in spec or "-" not in spec:
        raise ValueError("row window must look like Sheet Name:1-20")
    sheet_name, row_bounds = spec.rsplit(":", 1)
    start_text, end_text = row_bounds.split("-", 1)
    start = int(start_text)
    end = int(end_text)
    if start < 1 or end < start:
        raise ValueError("row window bounds must be positive and ordered")
    return sheet_name, (start, end)


def _sample_window(
    ws: Any,
    *,
    start_row: int,
    end_row: int,
    max_columns: int,
    preview_chars: int,
) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    formula_count = 0
    non_empty_count = 0

    for row_number, row in enumerate(
        ws.iter_rows(
            min_row=start_row,
            max_row=end_row,
            max_col=max_columns,
            values_only=False,
        ),
        start=start_row,
    ):
        cells: list[dict[str, Any]] = []
        row_non_empty_count = 0
        row_formula_count = 0
        first_non_empty_column: int | None = None
        last_non_empty_column: int | None = None

        for column_index, cell in enumerate(row, start=1):
            value = cell.value
            is_formula = _is_formula(value)
            if value is None and not is_formula:
                continue

            row_non_empty_count += 1
            non_empty_count += 1
            if is_formula:
                row_formula_count += 1
                formula_count += 1
            first_non_empty_column = (
                column_index
                if first_non_empty_column is None
                else min(first_non_empty_column, column_index)
            )
            last_non_empty_column = (
                column_index
                if last_non_empty_column is None
                else max(last_non_empty_column, column_index)
            )
            cells.append(
                {
                    "cell": f"{get_column_letter(column_index)}{row_number}",
                    "row": row_number,
                    "column": column_index,
                    "value_type": _value_type(value, is_formula=is_formula),
                    "value_preview": _preview(value, preview_chars),
                    "formula": str(value) if is_formula else None,
                }
            )

        rows.append(
            {
                "row": row_number,
                "non_empty_count": row_non_empty_count,
                "formula_count": row_formula_count,
                "first_non_empty_column": first_non_empty_column,
                "last_non_empty_column": last_non_empty_column,
                "cells": cells,
            }
        )

    return {
        "start_row": start_row,
        "end_row": end_row,
        "max_columns": max_columns,
        "row_count": len(rows),
        "non_empty_cell_count": non_empty_count,
        "formula_cell_count": formula_count,
        "rows": rows,
    }


def _summary(sheets: list[dict[str, Any]]) -> dict[str, int]:
    return {
        "sheet_count": len(sheets),
        "window_count": sum(len(sheet["windows"]) for sheet in sheets),
        "sampled_row_count": sum(
            window["row_count"]
            for sheet in sheets
            for window in sheet["windows"]
        ),
        "non_empty_cell_count": sum(
            window["non_empty_cell_count"]
            for sheet in sheets
            for window in sheet["windows"]
        ),
        "formula_cell_count": sum(
            window["formula_cell_count"]
            for sheet in sheets
            for window in sheet["windows"]
        ),
    }


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _value_type(value: Any, *, is_formula: bool) -> str:
    if is_formula:
        return "formula"
    if value is None:
        return "blank"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, (datetime, date, time)):
        return "datetime"
    return "string"


def _preview(value: Any, preview_chars: int) -> str | None:
    if value is None:
        return None
    if isinstance(value, (datetime, date, time)):
        text = value.isoformat()
    else:
        text = str(value)
    if len(text) <= preview_chars:
        return text
    return text[: max(preview_chars - 3, 0)] + "..."


def _read_json(path: Path | None) -> dict[str, Any] | None:
    if path is None:
        return None
    return json.loads(path.expanduser().resolve().read_text(encoding="utf-8"))


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Sample workbook rows through openpyxl read_only mode for source-driven evidence tuning."
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--sheet", action="append", dest="sheets")
    parser.add_argument(
        "--row-window",
        action="append",
        default=[],
        help="Target row window in the form 'Sheet Name:1-20'. May be repeated.",
    )
    parser.add_argument("--default-max-rows", type=int, default=20)
    parser.add_argument("--max-columns", type=int, default=50)
    parser.add_argument("--preview-chars", type=int, default=120)
    parser.add_argument("--manifest", type=Path)
    parser.add_argument("--include-structural-windows", action="store_true")
    parser.add_argument("--structural-window-padding", type=int, default=2)
    parser.add_argument("--max-structural-windows-per-sheet", type=int, default=30)
    args = parser.parse_args()

    sample = build_readonly_sample(
        args.workbook,
        manifest_path=args.manifest,
        include_structural_windows=args.include_structural_windows,
        structural_window_padding=args.structural_window_padding,
        max_structural_windows_per_sheet=args.max_structural_windows_per_sheet,
        sheets=args.sheets,
        row_windows=args.row_window,
        default_max_rows=args.default_max_rows,
        max_columns=args.max_columns,
        preview_chars=args.preview_chars,
    )
    payload = json.dumps(sample, ensure_ascii=False, indent=2)
    if args.output:
        args.output.write_text(payload + "\n", encoding="utf-8")
    else:
        print(payload)


if __name__ == "__main__":
    main()
