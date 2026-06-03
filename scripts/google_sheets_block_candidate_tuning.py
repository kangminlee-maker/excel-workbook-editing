from __future__ import annotations

import argparse
import html
import json
import re
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl.utils import get_column_letter, range_boundaries

from google_sheets_live_manifest import render_live_manifest_html


SCHEMA_VERSION = "0.1"


def build_block_candidate_tuning(
    *,
    live_block_candidates_path: Path,
    bounded_window_sample_path: Path,
) -> dict[str, Any]:
    live_block_candidates_path = live_block_candidates_path.expanduser().resolve()
    bounded_window_sample_path = bounded_window_sample_path.expanduser().resolve()
    block_candidates = _read_json(live_block_candidates_path)
    bounded_sample = _read_json(bounded_window_sample_path)
    sampled_regions = _sampled_regions(bounded_sample)
    tuning_actions = _tuning_actions(sampled_regions)
    remaining_read_queue = _remaining_read_queue(block_candidates, bounded_sample)
    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at": _utc_now(),
        "source": {
            "spreadsheet_id": block_candidates["source"]["spreadsheet_id"],
            "title": block_candidates["source"]["title"],
            "source_artifacts": {
                "live_block_candidates": str(live_block_candidates_path),
                "bounded_window_sample": str(bounded_window_sample_path),
            },
        },
        "authority": {
            "source_document": "live_google_sheet",
            "input_authority": "candidate_artifact_plus_broker_bounded_sample",
            "tuning_status": "tuning_packet_not_accepted_graph_claim",
            "formula_result_authority": "not_established",
        },
        "sampled_regions": sampled_regions,
        "tuning_actions": tuning_actions,
        "remaining_read_queue": remaining_read_queue,
        "summary": _summary(sampled_regions, tuning_actions, remaining_read_queue),
        "parser_observations": _parser_observations(sampled_regions, tuning_actions),
    }


def write_block_candidate_tuning_package(
    *,
    out_dir: Path,
    access_preflight_path: Path,
    live_manifest_path: Path,
    live_view_formula_profile_path: Path,
    live_block_candidates_path: Path,
    bounded_window_sample_path: Path,
    tuning: dict[str, Any],
) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    tuning_path = out_dir / "live-block-candidate-tuning.json"
    tuning_path.write_text(
        json.dumps(tuning, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    access_preflight = _read_json(access_preflight_path)
    manifest = _read_json(live_manifest_path)
    view_formula_profile = _read_json(live_view_formula_profile_path)
    block_candidates = _read_json(live_block_candidates_path)
    bounded_sample = _read_json(bounded_window_sample_path)
    (out_dir / "index.html").write_text(
        render_live_manifest_html(
            access_preflight=access_preflight,
            manifest=manifest,
            live_view_formula_profile=view_formula_profile,
            live_block_candidates=block_candidates,
            live_bounded_window_sample=bounded_sample,
            live_block_candidate_tuning=tuning,
        ),
        encoding="utf-8",
    )


def _sampled_regions(bounded_sample: dict[str, Any]) -> list[dict[str, Any]]:
    regions = []
    for response in bounded_sample.get("broker_responses", []):
        payload = response.get("payload", {})
        operation = response["operation"]
        for window in payload.get("windows", []) or []:
            regions.extend(_regions_for_window(operation, window))
    return regions


def _regions_for_window(operation: str, window: dict[str, Any]) -> list[dict[str, Any]]:
    range_text = window.get("range", "")
    sheet, a1_range = _split_sheet_range(range_text)
    min_col, min_row, _, _ = range_boundaries(a1_range)
    values = window.get("values", []) or []
    groups = _row_groups(values, start_row=min_row, start_column=min_col)
    regions = []
    for index, rows in enumerate(groups, start=1):
        bounds = _bounds_for_rows(rows)
        metrics = _metrics(rows)
        subtype = _region_subtype(operation, metrics)
        regions.append(
            {
                "id": f"sampled_region_{_slug(sheet)}_{_slug(operation)}_{index:03d}",
                "operation": operation,
                "sheet": sheet,
                "range": range_text,
                "subtype": subtype,
                "bounds": bounds,
                "metrics": metrics,
                "preview": _preview(rows),
                "evidence": [f"broker_window:{operation}:{range_text}"],
                "tuning_effect": _tuning_effect(subtype, metrics),
            }
        )
    return regions


def _row_groups(values: list[list[Any]], *, start_row: int, start_column: int) -> list[list[dict[str, Any]]]:
    groups = []
    current = []
    for row_offset, row in enumerate(values):
        cells = [
            {
                "row": start_row + row_offset,
                "column": start_column + column_offset,
                "value": value,
            }
            for column_offset, value in enumerate(row)
            if value not in ("", None)
        ]
        if cells:
            current.append({"row": start_row + row_offset, "cells": cells})
        elif current:
            groups.append(current)
            current = []
    if current:
        groups.append(current)
    return groups


def _tuning_actions(sampled_regions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    actions = []
    for region in sampled_regions:
        metrics = region["metrics"]
        if metrics["non_empty_cell_count"] == 0:
            continue
        if region["operation"] == "inspect.values_window":
            actions.append(
                {
                    "id": f"action_{region['id']}",
                    "type": "display_region_candidate_update",
                    "sheet": region["sheet"],
                    "target_range": region["bounds"]["a1_range"],
                    "effect": region["tuning_effect"],
                    "evidence_refs": [region["id"]],
                    "status": "candidate_update",
                }
            )
        if region["operation"] == "inspect.formula_window" and metrics["formula_cell_count"]:
            actions.append(
                {
                    "id": f"action_formula_{region['id']}",
                    "type": "formula_surface_confirmation",
                    "sheet": region["sheet"],
                    "target_range": region["bounds"]["a1_range"],
                    "effect": "raise_formula_surface_confidence_keep_formula_result_unestablished",
                    "evidence_refs": [region["id"]],
                    "status": "candidate_update",
                }
            )
        if metrics["url_cell_count"]:
            actions.append(
                {
                    "id": f"action_url_{region['id']}",
                    "type": "external_source_url_candidate",
                    "sheet": region["sheet"],
                    "target_range": region["bounds"]["a1_range"],
                    "effect": "use_url_as_candidate_source_id_for_importrange_review",
                    "evidence_refs": [region["id"]],
                    "status": "requires_source_allowlist_review",
                }
            )
        if metrics["error_display_count"]:
            actions.append(
                {
                    "id": f"action_error_{region['id']}",
                    "type": "formula_error_annotation",
                    "sheet": region["sheet"],
                    "target_range": region["bounds"]["a1_range"],
                    "effect": "keep_formula_result_authority_unestablished",
                    "evidence_refs": [region["id"]],
                    "status": "requires_formula_result_review",
                }
            )
    return actions


def _remaining_read_queue(
    block_candidates: dict[str, Any],
    bounded_sample: dict[str, Any],
) -> list[dict[str, Any]]:
    sampled_ranges = {
        range_text
        for request in bounded_sample.get("sampling_plan", {}).get("planned_requests", [])
        for range_text in request.get("ranges", [])
    }
    queue = []
    for sheet in block_candidates["sheets"]:
        for candidate in sheet["read_candidates"]:
            if candidate["range"] in sampled_ranges:
                continue
            queue.append(
                {
                    "id": candidate["id"],
                    "sheet": sheet["name"],
                    "operation": candidate["operation"],
                    "range": candidate["range"],
                    "reason": candidate["reason"],
                    "status": "pending_bounded_sampling",
                }
            )
    return queue


def _summary(
    sampled_regions: list[dict[str, Any]],
    tuning_actions: list[dict[str, Any]],
    remaining_read_queue: list[dict[str, Any]],
) -> dict[str, Any]:
    return {
        "sampled_region_count": len(sampled_regions),
        "tuning_action_count": len(tuning_actions),
        "display_region_update_count": sum(
            1 for item in tuning_actions if item["type"] == "display_region_candidate_update"
        ),
        "formula_surface_confirmation_count": sum(
            1 for item in tuning_actions if item["type"] == "formula_surface_confirmation"
        ),
        "external_source_url_candidate_count": sum(
            1 for item in tuning_actions if item["type"] == "external_source_url_candidate"
        ),
        "formula_error_annotation_count": sum(
            1 for item in tuning_actions if item["type"] == "formula_error_annotation"
        ),
        "remaining_read_queue_count": len(remaining_read_queue),
        "tuning_status": "generated_candidate_updates",
    }


def _parser_observations(
    sampled_regions: list[dict[str, Any]],
    tuning_actions: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    observations = [
        {
            "level": "info",
            "message": "Tuning actions are candidate updates only; they are not accepted graph claims.",
        }
    ]
    if any(item["type"] == "external_source_url_candidate" for item in tuning_actions):
        observations.append(
            {
                "level": "warning",
                "message": "URL candidates can guide IMPORTRANGE source review but do not authorize source spreadsheet reads.",
            }
        )
    if any(item["type"] == "formula_error_annotation" for item in tuning_actions):
        observations.append(
            {
                "level": "warning",
                "message": "Displayed errors were observed in bounded samples; formula-result authority remains unestablished.",
            }
        )
    if not sampled_regions:
        observations.append(
            {
                "level": "warning",
                "message": "No sampled regions were derived from bounded windows.",
            }
        )
    return observations


def _split_sheet_range(range_text: str) -> tuple[str, str]:
    if "!" not in range_text:
        return "", range_text
    sheet, a1 = range_text.split("!", 1)
    return sheet.strip("'"), a1


def _bounds_for_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    row_numbers = [row["row"] for row in rows]
    columns = [cell["column"] for row in rows for cell in row["cells"]]
    return _bounds(min(row_numbers), max(row_numbers), min(columns), max(columns))


def _bounds(start_row: int, end_row: int, start_column: int, end_column: int) -> dict[str, Any]:
    start_cell = f"{get_column_letter(start_column)}{start_row}"
    end_cell = f"{get_column_letter(end_column)}{end_row}"
    return {
        "start_row": start_row,
        "end_row": end_row,
        "start_column": start_column,
        "end_column": end_column,
        "a1_range": f"{start_cell}:{end_cell}" if start_cell != end_cell else start_cell,
    }


def _metrics(rows: list[dict[str, Any]]) -> dict[str, int]:
    values = [cell["value"] for row in rows for cell in row["cells"]]
    return {
        "row_count": len(rows),
        "non_empty_cell_count": len(values),
        "formula_cell_count": sum(1 for value in values if isinstance(value, str) and value.startswith("=")),
        "error_display_count": sum(1 for value in values if isinstance(value, str) and value.startswith("#")),
        "url_cell_count": sum(1 for value in values if isinstance(value, str) and value.startswith("http")),
        "number_like_cell_count": sum(1 for value in values if _is_number_like(str(value))),
    }


def _region_subtype(operation: str, metrics: dict[str, int]) -> str:
    if operation == "inspect.formula_window":
        return "sampled_formula_region"
    if metrics["url_cell_count"]:
        return "sampled_external_source_region"
    if metrics["number_like_cell_count"] >= 4:
        return "sampled_table_region"
    return "sampled_display_region"


def _tuning_effect(subtype: str, metrics: dict[str, int]) -> str:
    if subtype == "sampled_formula_region":
        return "confirm_formula_surface"
    if subtype == "sampled_external_source_region":
        return "add_external_source_candidate"
    if subtype == "sampled_table_region":
        return "extend_table_candidate_evidence"
    if metrics["error_display_count"]:
        return "annotate_error_state"
    return "add_display_evidence"


def _preview(rows: list[dict[str, Any]]) -> list[str]:
    preview = []
    for row in rows[:6]:
        values = [str(cell["value"]).replace("\n", " ").strip() for cell in row["cells"]]
        preview.append(f"R{row['row']}: " + " | ".join(values[:8]))
    return preview


def _is_number_like(value: str) -> bool:
    text = value.strip().replace(",", "").replace("%", "")
    if not text:
        return False
    try:
        float(text)
    except ValueError:
        return False
    return True


def _slug(value: Any) -> str:
    text = str(value or "none")
    text = re.sub(r"[^A-Za-z0-9가-힣]+", "_", text).strip("_").lower()
    return text or "none"


def _read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _utc_now() -> str:
    return datetime.now(UTC).isoformat().replace("+00:00", "Z")


def render_block_candidate_tuning_section(tuning: dict[str, Any]) -> str:
    metrics = "".join(
        f'<div class="metric"><div class="label">{_esc(key)}</div>'
        f'<div class="value">{_esc(value)}</div></div>'
        for key, value in tuning["summary"].items()
    )
    region_rows = "".join(
        "<tr>"
        f"<td>{_esc(item['sheet'])}</td>"
        f"<td>{_esc(item['operation'])}</td>"
        f"<td>{_esc(item['subtype'])}</td>"
        f"<td>{_esc(item['bounds']['a1_range'])}</td>"
        f"<td>{_esc(item['metrics'])}</td>"
        f"<td>{_esc(' / '.join(item['preview'][:2]))}</td>"
        "</tr>"
        for item in tuning["sampled_regions"][:40]
    )
    action_rows = "".join(
        "<tr>"
        f"<td>{_esc(item['type'])}</td>"
        f"<td>{_esc(item['sheet'])}</td>"
        f"<td>{_esc(item['target_range'])}</td>"
        f"<td>{_esc(item['status'])}</td>"
        f"<td>{_esc(item['effect'])}</td>"
        "</tr>"
        for item in tuning["tuning_actions"][:60]
    )
    observation_rows = "".join(
        "<tr>"
        f"<td>{_esc(item['level'])}</td>"
        f"<td>{_esc(item['message'])}</td>"
        "</tr>"
        for item in tuning["parser_observations"]
    )
    return f"""
  <h2>Block / Region Candidate Tuning</h2>
  <section class="grid">{metrics}</section>
  <h2>Sampled Regions</h2>
  <section class="panel"><table><thead><tr><th>Sheet</th><th>Operation</th><th>Subtype</th><th>Range</th><th>Metrics</th><th>Preview</th></tr></thead><tbody>{region_rows}</tbody></table></section>
  <h2>Tuning Actions</h2>
  <section class="panel"><table><thead><tr><th>Type</th><th>Sheet</th><th>Range</th><th>Status</th><th>Effect</th></tr></thead><tbody>{action_rows}</tbody></table></section>
  <h2>Tuning Observations</h2>
  <section class="panel"><table><thead><tr><th>Level</th><th>Message</th></tr></thead><tbody>{observation_rows}</tbody></table></section>
"""


def _esc(value: Any) -> str:
    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False)
    return html.escape(str(value))


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Tune Google Sheets block/region candidates with bounded sample windows."
    )
    parser.add_argument("--out-dir", type=Path, required=True)
    parser.add_argument("--access-preflight", type=Path, required=True)
    parser.add_argument("--live-manifest", type=Path, required=True)
    parser.add_argument("--live-view-formula-profile", type=Path, required=True)
    parser.add_argument("--live-block-candidates", type=Path, required=True)
    parser.add_argument("--bounded-window-sample", type=Path, required=True)
    args = parser.parse_args()

    tuning = build_block_candidate_tuning(
        live_block_candidates_path=args.live_block_candidates,
        bounded_window_sample_path=args.bounded_window_sample,
    )
    write_block_candidate_tuning_package(
        out_dir=args.out_dir,
        access_preflight_path=args.access_preflight,
        live_manifest_path=args.live_manifest,
        live_view_formula_profile_path=args.live_view_formula_profile,
        live_block_candidates_path=args.live_block_candidates,
        bounded_window_sample_path=args.bounded_window_sample,
        tuning=tuning,
    )


if __name__ == "__main__":
    main()
