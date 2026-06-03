from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path

import jsonschema
from openpyxl import Workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "scripts"))

from workbook_formula_pattern_profile import build_formula_pattern_profile  # noqa: E402


class WorkbookFormulaPatternProfileTest(unittest.TestCase):
    def test_builds_schema_valid_formula_signature_profile(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "formula_profile.xlsx"
            self._write_sample_workbook(workbook_path)

            profile = build_formula_pattern_profile(
                workbook_path,
                sheets=["누적"],
                row_windows=["누적:1-6"],
                max_columns=8,
            )

        schema = json.loads(
            (
                REPO_ROOT
                / "schemas"
                / "workbook-formula-pattern-profile.schema.json"
            ).read_text(encoding="utf-8")
        )
        jsonschema.Draft202012Validator(schema).validate(profile)

        window = profile["sheets"][0]["windows"][0]
        self.assertEqual(window["structure_hint"], "summary_formula_band")
        self.assertEqual(window["formula_count"], 2)
        self.assertEqual(window["repeated_signature_group_count"], 1)
        self.assertEqual(window["signature_groups"][0]["sample_cells"], ["E1", "F1"])

    @staticmethod
    def _write_sample_workbook(path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "누적"
        ws["E1"] = "=SUBTOTAL(9,E5:E100)"
        ws["F1"] = "=SUBTOTAL(9,F5:F100)"
        ws["A4"] = "ID"
        ws["E4"] = "매출"
        ws["F4"] = "수수료"
        ws["A5"] = "T1"
        ws["E5"] = 100
        ws["F5"] = 10
        wb.save(path)


if __name__ == "__main__":
    unittest.main()
