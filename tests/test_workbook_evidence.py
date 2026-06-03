from __future__ import annotations

import json
import sys
import tempfile
import unittest
from copy import copy
from pathlib import Path

import jsonschema
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Side, Border
from openpyxl.worksheet.table import Table, TableStyleInfo

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "scripts"))

from workbook_evidence import (  # noqa: E402
    build_evidence_package,
    build_evidence_package_from_artifacts,
)


class WorkbookEvidenceTest(unittest.TestCase):
    def test_builds_schema_valid_phase_one_evidence_package(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "document_shaped.xlsx"
            self._write_sample_workbook(workbook_path)

            package = build_evidence_package(workbook_path, max_cell_observations=100)

        schema = json.loads(
            (REPO_ROOT / "schemas" / "workbook-evidence.schema.json").read_text(
                encoding="utf-8"
            )
        )
        jsonschema.Draft202012Validator(schema).validate(package)

        self.assertEqual(package["schema_version"], "0.1")
        self.assertEqual(package["package_kind"], "direct_workbook_observation")
        self.assertEqual(package["summary"]["sheet_count"], 1)
        self.assertEqual(package["summary"]["merged_range_count"], 1)
        self.assertEqual(package["summary"]["table_count"], 1)
        self.assertEqual(package["summary"]["formula_cell_count"], 2)
        self.assertEqual(package["evidence_layers"]["rendered_visual"]["status"], "not_captured")
        self.assertEqual(package["coordinate_maps"]["status"], "grid_only")

        sheet = package["sheets"][0]
        self.assertEqual(sheet["merged_ranges"][0]["range"], "B2:D2")
        self.assertEqual(sheet["tables"][0]["range"], "B4:D6")

        title_cell = next(cell for cell in sheet["cells"] if cell["cell"] == "B2")
        self.assertTrue(title_cell["style"]["bold"])
        self.assertEqual(title_cell["alignment"]["horizontal"], "center")

        formulas = {item["cell"]: item["formula"] for item in sheet["formula_observations"]}
        self.assertEqual(formulas["D5"], "=C5*2")
        self.assertEqual(formulas["D6"], "=C6*2")

        seed_types = {seed["type"] for seed in sheet["visual_feature_seeds"]}
        self.assertIn("merged_range", seed_types)
        self.assertIn("declared_table", seed_types)

    def test_assembles_artifact_first_evidence_package(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            paths = _write_artifact_fixture(root)

            package = build_evidence_package_from_artifacts(**paths)

        schema = json.loads(
            (REPO_ROOT / "schemas" / "workbook-evidence.schema.json").read_text(
                encoding="utf-8"
            )
        )
        jsonschema.Draft202012Validator(schema).validate(package)

        self.assertEqual(package["package_kind"], "artifact_assembled_workbook_understanding")
        self.assertEqual(package["summary"]["artifact_count"], 19)
        self.assertEqual(package["summary"]["accepted_gate_count"], 1)
        self.assertEqual(package["summary"]["review_queue_count"], 2)
        self.assertEqual(package["coordinate_maps"]["status"], "normalized_captures_available")
        self.assertEqual(
            package["decision_indexes"]["accepted_pipeline_role_validation_ids"],
            ["role_validation_pipeline_summary"],
        )

    @staticmethod
    def _write_sample_workbook(path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Inspection"

        ws.merge_cells("B2:D2")
        ws["B2"] = "Inspection Result"
        ws["B2"].font = Font(bold=True, size=16)
        ws["B2"].fill = PatternFill("solid", fgColor="FFE699")
        alignment = copy(ws["B2"].alignment)
        alignment.horizontal = "center"
        ws["B2"].alignment = alignment

        headers = ["Item", "Measured", "Score"]
        for offset, header in enumerate(headers):
            cell = ws.cell(row=4, column=2 + offset, value=header)
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(style="thin"))

        ws["B5"] = "Length"
        ws["C5"] = 12
        ws["D5"] = "=C5*2"
        ws["B6"] = "Width"
        ws["C6"] = 8
        ws["D6"] = "=C6*2"
        ws["B8"] = "The table above contains sample inspection metrics."
        ws.column_dimensions["B"].width = 24
        ws.row_dimensions[2].height = 28

        table = Table(displayName="InspectionTable", ref="B4:D6")
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        wb.save(path)


def _write_artifact_fixture(root: Path) -> dict:
    common_source = {
        "path": str(root / "source.xlsx"),
        "file_name": "source.xlsx",
        "size_bytes": 123,
        "sha256": "0" * 64,
    }
    artifacts = {
        "manifest_path": {
            "schema_version": "0.1",
            "generated_at": "2026-06-01T00:00:00Z",
            "source": common_source,
            "workbook": {
                "sheet_count": 1,
                "external_links": [],
                "sheets": [
                    {
                        "name": "Summary",
                        "index": 0,
                        "state": "visible",
                        "dimension": "A1:C3",
                        "dimension_bounds": {
                            "min_row": 1,
                            "min_column": 1,
                            "max_row": 3,
                            "max_column": 3,
                        },
                        "detail_status": "scanned",
                        "drawing_objects": [],
                        "relationships": {},
                    }
                ],
            },
            "summary": {"sheet_count": 1},
        },
        "readonly_sample_path": {
            "schema_version": "0.1",
            "generated_at": "2026-06-01T00:00:00Z",
            "source": common_source,
            "limits": {"default_max_rows": 10, "preview_chars": 40},
            "sheets": [
                {
                    "name": "Summary",
                    "windows": [
                        {
                            "rows": [
                                {
                                    "row": 1,
                                    "cells": [
                                        {
                                            "cell": "A1",
                                            "row": 1,
                                            "column": 1,
                                            "value_type": "string",
                                            "value_preview": "Label",
                                            "formula": None,
                                        },
                                        {
                                            "cell": "B1",
                                            "row": 1,
                                            "column": 2,
                                            "value_type": "formula",
                                            "value_preview": "=SUM(Raw!A:A)",
                                            "formula": "=SUM(Raw!A:A)",
                                        },
                                    ],
                                }
                            ]
                        }
                    ],
                }
            ],
            "summary": {"formula_cell_count": 1},
        },
        "structural_style_profile_path": {
            "schema_version": "0.1",
            "generated_at": "2026-06-01T00:00:00Z",
            "source": common_source,
            "sheets": [
                {
                    "name": "Summary",
                    "merge_ranges": [],
                    "row_dimensions": [{"row": 1, "height": None, "hidden": False}],
                    "column_dimensions": [
                        {
                            "min_column": 1,
                            "max_column": 2,
                            "width": 12,
                            "hidden": False,
                        }
                    ],
                }
            ],
            "summary": {},
        },
        "view_state_preflight_path": {
            "schema_version": "0.1",
            "generated_at": "2026-06-01T00:00:00Z",
            "sheets": [
                {
                    "name": "Summary",
                    "sheet_state": "visible",
                    "sheet_format_pr": {"defaultRowHeight": "16.5"},
                    "hidden_row_spans": [],
                    "hidden_column_spans": [],
                    "panes": [],
                    "auto_filters": [],
                    "summary": {"hidden_row_count": 0, "hidden_column_count": 0},
                }
            ],
            "summary": {},
        },
        "coordinate_normalization_path": {
            "schema_version": "0.1",
            "generated_at": "2026-06-01T00:00:00Z",
            "coordinate_mappings": [
                {
                    "id": "coord_1",
                    "status": "normalized_visible_range",
                    "sheet": "Summary",
                    "cell_range": "A1:C3",
                    "capture_id": "capture_1",
                    "target_id": "target_1",
                    "quality_status": "usable",
                    "view_state_classification": "no_material_view_state_signal",
                }
            ],
            "summary": {},
        },
        "gate_execution_path": {
            "schema_version": "0.1",
            "generated_at": "2026-06-01T00:00:00Z",
            "gate_results": [
                {
                    "id": "gate_result_1",
                    "status": "accepted",
                    "reason": "ok",
                    "evidence_refs": ["gate_1"],
                },
                {
                    "id": "gate_result_2",
                    "status": "review_required",
                    "reason": "capture_required",
                    "evidence_refs": ["gate_2"],
                },
            ],
            "summary": {"accepted_count": 1, "review_required_count": 1},
        },
        "boundary_decisions_path": {
            "schema_version": "0.1",
            "generated_at": "2026-06-01T00:00:00Z",
            "boundary_decisions": [],
            "summary": {"accepted_count": 0, "review_required_count": 0},
        },
        "pipeline_role_validation_path": {
            "schema_version": "0.1",
            "generated_at": "2026-06-01T00:00:00Z",
            "role_validations": [
                {
                    "id": "role_validation_pipeline_summary",
                    "status": "accepted",
                    "reason": "summary_formula_role_supported",
                    "output_ref": {"sheet": "Summary", "range": "A1:C3"},
                    "evidence_refs": ["pipeline_summary"],
                },
                {
                    "id": "role_validation_pipeline_review",
                    "status": "review_required",
                    "reason": "unresolved_input_region",
                    "output_ref": {"sheet": "Summary", "range": "E1:E3"},
                    "evidence_refs": ["pipeline_review"],
                },
            ],
            "summary": {"accepted_count": 1, "review_required_count": 1},
        },
    }
    generic_names = [
        "formula_patterns_path",
        "block_candidates_path",
        "table_io_pipelines_path",
        "cross_validation_plan_path",
        "render_captures_path",
        "capture_quality_path",
        "recapture_candidate_plan_path",
        "recapture_candidate_captures_path",
        "recapture_candidate_quality_path",
        "view_state_profile_path",
        "visual_features_path",
    ]
    for name in generic_names:
        artifacts[name] = {
            "schema_version": "0.1",
            "generated_at": "2026-06-01T00:00:00Z",
            "summary": {},
        }
    artifacts["render_captures_path"]["summary"] = {"capture_count": 1}
    artifacts["visual_features_path"]["summary"] = {
        "feature_result_count": 1,
        "detected_count": 1,
    }

    paths = {}
    for arg_name, payload in artifacts.items():
        path = root / f"{arg_name}.json"
        path.write_text(json.dumps(payload), encoding="utf-8")
        paths[arg_name] = path
    return paths


if __name__ == "__main__":
    unittest.main()
