from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path

import jsonschema
from openpyxl import Workbook
from openpyxl.styles import Font

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "scripts"))

from workbook_manifest import build_workbook_manifest  # noqa: E402


class WorkbookManifestTest(unittest.TestCase):
    def test_builds_schema_valid_manifest_without_full_workbook_model(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "manifest_sample.xlsx"
            self._write_sample_workbook(workbook_path)

            manifest = build_workbook_manifest(
                workbook_path,
                sample_limit=5,
                max_sheet_xml_bytes=10_000_000,
            )

        schema = json.loads(
            (REPO_ROOT / "schemas" / "workbook-manifest.schema.json").read_text(
                encoding="utf-8"
            )
        )
        jsonschema.Draft202012Validator(schema).validate(manifest)

        self.assertEqual(manifest["schema_version"], "0.1")
        self.assertIn("shared_strings", manifest)
        self.assertIn("pivot_caches", manifest["workbook"])
        self.assertIn("external_links", manifest["workbook"])
        self.assertEqual(manifest["workbook"]["external_links"], [])
        self.assertEqual(manifest["summary"]["sheet_count"], 1)
        self.assertEqual(manifest["summary"]["scanned_sheet_count"], 1)
        self.assertEqual(manifest["summary"]["skipped_large_sheet_count"], 0)

        sheet = manifest["workbook"]["sheets"][0]
        self.assertEqual(sheet["detail_status"], "scanned")
        self.assertEqual(sheet["dimension"], "B2:D5")
        self.assertEqual(
            sheet["dimension_bounds"],
            {"min_row": 2, "min_column": 2, "max_row": 5, "max_column": 4},
        )
        self.assertEqual(sheet["counts"]["formula_elements"], 1)
        self.assertEqual(sheet["samples"]["cells"][0]["value_preview"], "Document Header")
        self.assertEqual(sheet["samples"]["formulas"][0]["cell"], "D5")
        self.assertEqual(sheet["samples"]["merged_ranges"], ["B2:D2"])
        self.assertEqual(sheet["drawing_objects"], [])
        self.assertEqual(sheet["pivot_tables"], [])

    @staticmethod
    def _write_sample_workbook(path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Manifest"
        ws.merge_cells("B2:D2")
        ws["B2"] = "Document Header"
        ws["B2"].font = Font(bold=True)
        ws["B4"] = "Metric"
        ws["C4"] = "Value"
        ws["D4"] = "Calculated"
        ws["B5"] = "Revenue"
        ws["C5"] = 100
        ws["D5"] = "=C5*2"
        ws.column_dimensions["B"].width = 22
        wb.save(path)


if __name__ == "__main__":
    unittest.main()
