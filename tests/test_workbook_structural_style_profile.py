from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path

import jsonschema
from openpyxl import Workbook
from openpyxl.styles import PatternFill

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "scripts"))

from workbook_manifest import build_workbook_manifest  # noqa: E402
from workbook_readonly_sample import build_readonly_sample  # noqa: E402
from workbook_structural_style_profile import build_structural_style_profile  # noqa: E402


class WorkbookStructuralStyleProfileTest(unittest.TestCase):
    def test_extracts_merged_ranges_and_style_boundaries(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            workbook_path = root / "sample.xlsx"
            manifest_path = root / "manifest.json"
            sample_path = root / "sample.json"
            _write_style_workbook(workbook_path)

            manifest = build_workbook_manifest(workbook_path)
            sample = build_readonly_sample(
                workbook_path,
                sheets=["Sheet1"],
                default_max_rows=5,
                max_columns=5,
            )
            manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
            sample_path.write_text(json.dumps(sample), encoding="utf-8")

            profile = build_structural_style_profile(
                manifest_path,
                sample_path,
                sheets=["Sheet1"],
            )

        schema = json.loads(
            (
                REPO_ROOT
                / "schemas"
                / "workbook-structural-style-profile.schema.json"
            ).read_text(encoding="utf-8")
        )
        jsonschema.Draft202012Validator(schema).validate(profile)

        sheet = profile["sheets"][0]
        self.assertEqual(profile["summary"]["sheet_count"], 1)
        self.assertEqual(sheet["summary"]["merge_range_count"], 1)
        self.assertGreaterEqual(sheet["summary"]["style_boundary_count"], 1)
        self.assertEqual(sheet["merge_ranges"][0]["range"], "A1:B1")


def _write_style_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.merge_cells("A1:B1")
    ws["A1"] = "Merged Title"
    ws["A2"] = "Left"
    ws["B2"] = "Right"
    ws["A3"] = 10
    ws["B3"] = 20
    ws["A2"].fill = PatternFill("solid", fgColor="FFEEAA")
    ws["A3"].fill = PatternFill("solid", fgColor="FFEEAA")
    ws["B2"].fill = PatternFill("solid", fgColor="AADDEE")
    ws["B3"].fill = PatternFill("solid", fgColor="AADDEE")
    wb.save(path)


if __name__ == "__main__":
    unittest.main()
